# -*- coding: utf-8 -*-
"""
PJ Board — 社内向けセルフホスト・プロジェクト管理ツール
FastAPI + SQLite 単一プロセス構成。閉域ネットワークで動作（外部CDN依存なし）。

起動:  python -m uvicorn app:app --host 0.0.0.0 --port 8100
"""
import contextvars
import csv
import hashlib
import io
import json
import os
import re
import secrets
import sqlite3
import zipfile
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from typing import Any, Optional

from fastapi import FastAPI, HTTPException, Request, Response
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.environ.get("PJBOARD_DB", os.path.join(DATA_DIR, "pjboard.db"))

# デバッグ機能（右上のログイン完全切替）。本番では PJBOARD_DEBUG=0 で無効化する。
DEBUG_FEATURES = os.environ.get("PJBOARD_DEBUG", "1") != "0"

os.makedirs(DATA_DIR, exist_ok=True)

app = FastAPI(
    title="PJ Board API",
    version="0.3.1",
    description=(
        "社内プロジェクト管理ツール PJ Board の REST API。\n\n"
        "AIエージェント向けの要点:\n"
        "- まず GET /api/ai/help を読むと全体像と操作手順が分かります\n"
        "- GET /api/ai/context で全データの要約(JSON)を取得できます\n"
        "- 書き込み系は body/query の actor_id にユーザーIDを渡すと権限チェックが働きます"
        "（省略時はチェックなし＝管理操作扱い）\n"
        "- 日付は YYYY-MM-DD 形式の文字列です"
    ),
)


# ---------------------------------------------------------------- DB helpers

@contextmanager
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def rows_to_dicts(rows) -> list[dict]:
    out = []
    for r in rows:
        d = dict(r)
        d.pop("password_hash", None)   # 認証情報はAPIレスポンスに出さない
        out.append(d)
    return out


def now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


SCHEMA = """
CREATE TABLE IF NOT EXISTS orgs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    color TEXT DEFAULT '#64748b',
    created_at TEXT
);
CREATE TABLE IF NOT EXISTS project_members (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    member_id INTEGER NOT NULL REFERENCES members(id) ON DELETE CASCADE,
    role TEXT DEFAULT 'member',            -- leader / member / advisor / external
    can_view_comments INTEGER DEFAULT 1,   -- 主に external 用の閲覧制限フラグ
    can_view_detail INTEGER DEFAULT 1,
    added_at TEXT,
    UNIQUE(project_id, member_id)
);
CREATE TABLE IF NOT EXISTS project_notes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    category TEXT DEFAULT 'その他',        -- 環境 / 体制 / ルール / その他 など自由
    title TEXT NOT NULL,
    content TEXT DEFAULT '',
    pinned INTEGER DEFAULT 0,
    sort_order INTEGER DEFAULT 0,
    updated_by INTEGER REFERENCES members(id) ON DELETE SET NULL,
    created_at TEXT, updated_at TEXT
);
CREATE TABLE IF NOT EXISTS projects (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    description TEXT DEFAULT '',
    color TEXT DEFAULT '#4f6ef7',
    status TEXT DEFAULT 'active',          -- active / archived
    start_date TEXT, end_date TEXT,
    custom_fields TEXT DEFAULT '[]',       -- [{key,label,type,options[]}]
    settings TEXT DEFAULT '{}',            -- 表示・権限・テンプレート等のPJ別設定
    created_at TEXT
);
CREATE TABLE IF NOT EXISTS members (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    role TEXT DEFAULT '',
    color TEXT DEFAULT '#7c8db5',
    org_id INTEGER REFERENCES orgs(id) ON DELETE SET NULL,
    account_type TEXT DEFAULT 'internal',  -- internal / external
    org_role TEXT DEFAULT 'staff',         -- 組織権限: manager / site_admin / professional / staff
    password_hash TEXT,                    -- "salt$sha256"。NULL は初期状態（空パスワードでログイン可）
    email TEXT,                            -- ログインID（必須運用・小文字で一意）
    active INTEGER DEFAULT 1
);
CREATE TABLE IF NOT EXISTS sessions (
    token TEXT PRIMARY KEY,
    member_id INTEGER NOT NULL REFERENCES members(id) ON DELETE CASCADE,
    created_at TEXT, last_seen TEXT
);
CREATE TABLE IF NOT EXISTS statuses (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    name TEXT NOT NULL,
    color TEXT DEFAULT '#8b95a7',
    sort_order INTEGER DEFAULT 0,
    is_done INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    parent_id INTEGER REFERENCES tasks(id) ON DELETE SET NULL,
    title TEXT NOT NULL,
    description TEXT DEFAULT '',
    status_id INTEGER REFERENCES statuses(id) ON DELETE SET NULL,
    assignee_id INTEGER REFERENCES members(id) ON DELETE SET NULL,
    assignee_label TEXT,                   -- 仮想担当（メンバー外の選択肢。例: 顧客）
    priority TEXT DEFAULT 'medium',        -- highest/high/medium/low
    start_date TEXT, due_date TEXT,
    progress INTEGER DEFAULT 0,
    estimate_h REAL, actual_h REAL,
    milestone INTEGER DEFAULT 0,
    tags TEXT DEFAULT '[]',
    deps TEXT DEFAULT '[]',                -- 先行タスクID配列
    custom_values TEXT DEFAULT '{}',
    sort_order INTEGER DEFAULT 0,
    created_at TEXT, updated_at TEXT
);
CREATE TABLE IF NOT EXISTS comments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
    author_id INTEGER REFERENCES members(id) ON DELETE SET NULL,
    body TEXT NOT NULL,
    created_at TEXT
);
CREATE TABLE IF NOT EXISTS task_links (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
    title TEXT NOT NULL,
    url TEXT DEFAULT '',
    kind TEXT DEFAULT 'link'               -- link / doc / design / repo
);
CREATE TABLE IF NOT EXISTS meta (
    key TEXT PRIMARY KEY, value TEXT
);
CREATE TABLE IF NOT EXISTS user_prefs (
    member_id INTEGER NOT NULL REFERENCES members(id) ON DELETE CASCADE,
    key TEXT NOT NULL,
    value TEXT DEFAULT '{}',
    PRIMARY KEY (member_id, key)
);
CREATE TABLE IF NOT EXISTS qa_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    seq INTEGER,                           -- PJ内連番（QA-001 表示用）
    title TEXT NOT NULL,                   -- 件名
    question TEXT DEFAULT '',              -- 質問の詳細
    answer TEXT DEFAULT '',                -- 回答
    decision TEXT DEFAULT '',              -- 決定事項
    note TEXT DEFAULT '',                  -- 備考
    category TEXT DEFAULT '',              -- 分類（仕様/環境/データ 等 自由）
    priority TEXT DEFAULT 'medium',        -- high / medium / low
    status TEXT DEFAULT 'open',            -- open(回答待ち)/pending(保留)/answered(回答済み)/closed(クローズ)
    asker_name TEXT DEFAULT '',            -- 質問者（顧客名等の自由記入）
    assignee_id INTEGER REFERENCES members(id) ON DELETE SET NULL,  -- 回答担当
    task_id INTEGER REFERENCES tasks(id) ON DELETE SET NULL,        -- 関連タスク
    asked_at TEXT,                         -- 質問日
    due_date TEXT,                         -- 回答期限
    answered_at TEXT,                      -- 回答日
    created_at TEXT, updated_at TEXT
);
CREATE TABLE IF NOT EXISTS issues (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    seq INTEGER,                           -- PJ内連番（ISS-001 表示用）
    title TEXT NOT NULL,                   -- 件名
    description TEXT DEFAULT '',           -- 課題の内容・背景
    policy TEXT DEFAULT '',                -- 方針
    action_plan TEXT DEFAULT '',           -- 実行内容
    category TEXT DEFAULT '',              -- 分類（自由）
    priority TEXT DEFAULT 'medium',        -- highest/high/medium/low（重要度）
    status TEXT DEFAULT 'open',            -- open(未対応)/doing(対応中)/resolved(解決済み)/closed(クローズ)
    assignee_id INTEGER REFERENCES members(id) ON DELETE SET NULL,  -- 担当者
    raised_by TEXT DEFAULT '',             -- 起票者（自由記入）
    raised_at TEXT,                        -- 起票日
    due_date TEXT,                         -- 対応期限
    resolved_at TEXT,                      -- 解決日
    created_at TEXT, updated_at TEXT
);
CREATE TABLE IF NOT EXISTS issue_tasks (   -- 課題⇔タスクの関連（多対多）
    issue_id INTEGER NOT NULL REFERENCES issues(id) ON DELETE CASCADE,
    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
    PRIMARY KEY (issue_id, task_id)
);
CREATE TABLE IF NOT EXISTS issue_comments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    issue_id INTEGER NOT NULL REFERENCES issues(id) ON DELETE CASCADE,
    author_id INTEGER REFERENCES members(id) ON DELETE SET NULL,
    author_name TEXT DEFAULT '',
    body TEXT NOT NULL,
    created_at TEXT
);
CREATE TABLE IF NOT EXISTS qa_comments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    qa_id INTEGER NOT NULL REFERENCES qa_items(id) ON DELETE CASCADE,
    author_id INTEGER REFERENCES members(id) ON DELETE SET NULL,
    author_name TEXT DEFAULT '',           -- 表示名（Excel取込・顧客発言など非メンバー用）
    body TEXT NOT NULL,
    created_at TEXT
);
CREATE TABLE IF NOT EXISTS notifications (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL REFERENCES members(id) ON DELETE CASCADE,
    type TEXT NOT NULL,                    -- mention / assign / comment / status / due / watch / system
    project_id INTEGER, task_id INTEGER, actor_id INTEGER,
    message TEXT DEFAULT '',
    read INTEGER DEFAULT 0,
    created_at TEXT
);
CREATE TABLE IF NOT EXISTS watchers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
    member_id INTEGER NOT NULL REFERENCES members(id) ON DELETE CASCADE,
    UNIQUE(task_id, member_id)
);
CREATE TABLE IF NOT EXISTS attachments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    target_type TEXT NOT NULL,             -- task / note
    target_id INTEGER NOT NULL,
    filename TEXT NOT NULL,
    stored_name TEXT NOT NULL,
    size INTEGER DEFAULT 0,
    content_type TEXT DEFAULT 'application/octet-stream',
    uploaded_by INTEGER REFERENCES members(id) ON DELETE SET NULL,
    created_at TEXT
);
CREATE TABLE IF NOT EXISTS login_logs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    member_id INTEGER, name TEXT, success INTEGER,
    ip TEXT, ua TEXT, created_at TEXT
);
CREATE TABLE IF NOT EXISTS api_tokens (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    member_id INTEGER NOT NULL REFERENCES members(id) ON DELETE CASCADE,
    token_hash TEXT NOT NULL,
    label TEXT DEFAULT '',
    created_at TEXT, last_used TEXT
);
CREATE TABLE IF NOT EXISTS reactions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    comment_id INTEGER NOT NULL REFERENCES comments(id) ON DELETE CASCADE,
    member_id INTEGER NOT NULL REFERENCES members(id) ON DELETE CASCADE,
    emoji TEXT NOT NULL,
    UNIQUE(comment_id, member_id, emoji)
);
CREATE TABLE IF NOT EXISTS task_relations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
    other_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
    kind TEXT DEFAULT 'relates',           -- relates / blocks
    UNIQUE(task_id, other_id, kind)
);
CREATE TABLE IF NOT EXISTS saved_filters (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    member_id INTEGER NOT NULL REFERENCES members(id) ON DELETE CASCADE,
    project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    name TEXT NOT NULL,
    filters TEXT DEFAULT '{}',
    created_at TEXT
);
CREATE TABLE IF NOT EXISTS baselines (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    name TEXT DEFAULT '',
    snapshot TEXT DEFAULT '[]',            -- [{task_id, start, due}]
    created_at TEXT
);
CREATE TABLE IF NOT EXISTS activities (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    task_id INTEGER REFERENCES tasks(id) ON DELETE CASCADE,
    actor_id INTEGER REFERENCES members(id) ON DELETE SET NULL,
    action TEXT NOT NULL,
    detail TEXT DEFAULT '',
    created_at TEXT
);
"""

DEFAULT_STATUSES = [
    ("未着手", "#8b95a7", 0, 0),
    ("進行中", "#4f6ef7", 1, 0),
    ("レビュー中", "#f59e0b", 2, 0),
    ("完了", "#22c55e", 3, 1),
]

# プロジェクト別設定の既定値（projects.settings にJSONで差分保存し、読み出し時にマージ）
DEFAULT_NOTE_TEMPLATES = [
    {"category": "環境", "title": "検証環境一覧・アクセス方法"},
    {"category": "体制", "title": "社内体制・担当者"},
    {"category": "体制", "title": "先方担当者・窓口"},
    {"category": "ルール", "title": "定例・会議体"},
    {"category": "ルール", "title": "開発ルール（ブランチ運用・レビュー等）"},
    {"category": "ツール", "title": "使用ツール・リンク集"},
    {"category": "その他", "title": "決定事項ログ"},
    {"category": "その他", "title": "障害・緊急時の連絡フロー"},
]

DEFAULT_PROJECT_SETTINGS = {
    # --- 表示設定（タスクの項目をどこまで出すか） ---
    "show_estimate": True,       # 見積h・実績h
    "show_priority": True,
    "show_tags": True,
    "show_custom_fields": True,
    # --- 権限の細部調整 ---
    "member_can_edit_own_schedule": False,  # memberが自分の担当タスクの開始/期限/見積を変更可
    "member_can_edit_notes": True,
    "member_can_create_tasks": True,
    "advisor_can_comment": True,            # ご意見番のコメント投稿
    "unassigned_can_comment": True,         # 未アサイン社内ユーザーのコメント投稿（横断的な意見募集の肝）
    # --- 外部ユーザーの既定・制限 ---
    "external_default_view_comments": False,
    "external_default_view_detail": False,
    "external_can_export": False,
    # 外部パートナーに公開するタブ（変更はマネージャー/サイト管理者＋確認入力が必要）
    # ※課題(kadai)は社内の管理情報のため、既定では外部に公開しない
    "external_visible_tabs": ["dashboard", "board", "table", "gantt",
                              "calendar", "qa", "issues", "notes"],
    # --- 担当者の追加選択肢（アサイン済みメンバー以外。例: 顧客・ベンダーA） ---
    "virtual_assignees": [],
    # --- 通知（Teams/Slack Incoming Webhook） ---
    "webhook_url": "",
    "webhook_events": ["mention", "assign", "due"],   # mention/assign/due/status/comment
    # --- ノートテンプレート ---
    "note_templates": DEFAULT_NOTE_TEMPLATES,
}


def merge_settings(raw: Optional[str]) -> dict:
    s = dict(DEFAULT_PROJECT_SETTINGS)
    try:
        s.update(json.loads(raw or "{}"))
    except (ValueError, TypeError):
        pass
    return s


def migrate(conn: sqlite3.Connection):
    """既存DBを新スキーマへ段階的に移行する。"""
    cols = {r["name"] for r in conn.execute("PRAGMA table_info(members)")}
    if "org_id" not in cols:
        conn.execute("ALTER TABLE members ADD COLUMN org_id INTEGER REFERENCES orgs(id)")
    if "account_type" not in cols:
        conn.execute("ALTER TABLE members ADD COLUMN account_type TEXT DEFAULT 'internal'")
    if "org_role" not in cols:
        conn.execute("ALTER TABLE members ADD COLUMN org_role TEXT DEFAULT 'staff'")
    if "password_hash" not in cols:
        conn.execute("ALTER TABLE members ADD COLUMN password_hash TEXT")
    if "email" not in cols:
        conn.execute("ALTER TABLE members ADD COLUMN email TEXT")   # ログインID / SSOヘッダー連携用
    # メールログイン化に伴い、未登録メンバーへダミーメールを補完（毎起動・冪等）
    conn.execute("UPDATE members SET email='user'||id||'@example.com'"
                 " WHERE email IS NULL OR email=''")
    t_cols = {r["name"] for r in conn.execute("PRAGMA table_info(tasks)")}
    if "deleted_at" not in t_cols:
        conn.execute("ALTER TABLE tasks ADD COLUMN deleted_at TEXT")       # ゴミ箱（論理削除）
    if "recur" not in t_cols:
        conn.execute("ALTER TABLE tasks ADD COLUMN recur TEXT DEFAULT ''")  # 繰り返し: weekly/biweekly/monthly
    if "assignee_label" not in t_cols:
        conn.execute("ALTER TABLE tasks ADD COLUMN assignee_label TEXT")    # 仮想担当
    n_cols = {r["name"] for r in conn.execute("PRAGMA table_info(project_notes)")}
    if "deleted_at" not in n_cols:
        conn.execute("ALTER TABLE project_notes ADD COLUMN deleted_at TEXT")
    c_cols = {r["name"] for r in conn.execute("PRAGMA table_info(comments)")}
    if "updated_at" not in c_cols:
        conn.execute("ALTER TABLE comments ADD COLUMN updated_at TEXT")
    # 期限切れセッションのパージ（30日）
    conn.execute("DELETE FROM sessions WHERE last_seen < datetime('now', '-30 days')")
    pm_cols = {r["name"] for r in conn.execute("PRAGMA table_info(project_members)")}
    if pm_cols and "role" not in pm_cols:
        conn.execute("ALTER TABLE project_members ADD COLUMN role TEXT DEFAULT 'member'")
        conn.execute("ALTER TABLE project_members ADD COLUMN can_view_comments INTEGER DEFAULT 1")
        conn.execute("ALTER TABLE project_members ADD COLUMN can_view_detail INTEGER DEFAULT 1")
        # 権限導入前からのメンバーは従来どおり全操作できるよう admin(現leader) に引き上げる
        conn.execute("UPDATE project_members SET role='admin'")
    # 旧ロール名 → 新ロール名（admin→leader / viewer→advisor）
    conn.execute("UPDATE project_members SET role='leader' WHERE role='admin'")
    conn.execute("UPDATE project_members SET role='advisor' WHERE role='viewer'")
    p_cols = {r["name"] for r in conn.execute("PRAGMA table_info(projects)")}
    if "settings" not in p_cols:
        conn.execute("ALTER TABLE projects ADD COLUMN settings TEXT DEFAULT '{}'")
    # 組織が1つも無ければ既定組織を作り、無所属ユーザーを割り当てる
    if conn.execute("SELECT COUNT(*) c FROM orgs").fetchone()["c"] == 0 and \
       conn.execute("SELECT COUNT(*) c FROM members").fetchone()["c"] > 0:
        cur = conn.execute("INSERT INTO orgs(name, color, created_at) VALUES(?,?,?)",
                           ("既定の組織", "#4f6ef7", now()))
        conn.execute("UPDATE members SET org_id=? WHERE org_id IS NULL", (cur.lastrowid,))
    # QA: 決定事項・備考カラム（2026-08-30追加）
    qa_cols = {r["name"] for r in conn.execute("PRAGMA table_info(qa_items)")}
    if qa_cols and "decision" not in qa_cols:
        conn.execute("ALTER TABLE qa_items ADD COLUMN decision TEXT DEFAULT ''")
        conn.execute("ALTER TABLE qa_items ADD COLUMN note TEXT DEFAULT ''")
    # 親タスク進捗の自動算出を既存データにも適用（毎起動・冪等）
    for r in conn.execute("SELECT id FROM projects"):
        recalc_parent_progress(conn, r["id"])
    # project_members が空なら、従来の全員共有の挙動を維持するため
    # 既存の全プロジェクトに全アクティブユーザーをアサインする
    if conn.execute("SELECT COUNT(*) c FROM project_members").fetchone()["c"] == 0:
        pids = [r["id"] for r in conn.execute("SELECT id FROM projects")]
        mids = [r["id"] for r in conn.execute("SELECT id FROM members WHERE active=1")]
        for pid in pids:
            for mid in mids:
                conn.execute(
                    "INSERT OR IGNORE INTO project_members(project_id, member_id, added_at)"
                    " VALUES(?,?,?)", (pid, mid, now()))


def init_db():
    with db() as conn:
        conn.executescript(SCHEMA)
        migrate(conn)
        if conn.execute("SELECT COUNT(*) c FROM projects").fetchone()["c"] == 0:
            seed_demo(conn)


def seed_demo(conn: sqlite3.Connection):
    """初回起動時のデモデータ。使い方の見本を兼ねる。"""
    t = now()
    org1 = conn.execute("INSERT INTO orgs(name, color, created_at) VALUES(?,?,?)",
                        ("情報システム部", "#4f6ef7", t)).lastrowid
    org2 = conn.execute("INSERT INTO orgs(name, color, created_at) VALUES(?,?,?)",
                        ("外部パートナー", "#8b5cf6", t)).lastrowid
    members = [
        ("田中 太郎", "PM", "#4f6ef7", org1, "internal", "manager"),
        ("佐藤 花子", "デザイナー", "#ec4899", org1, "internal", "staff"),
        ("鈴木 一郎", "エンジニア", "#22c55e", org1, "internal", "staff"),
        ("山田 美咲", "エンジニア", "#f59e0b", org2, "external", "staff"),
    ]
    mids = []
    for name, role, color, org, acct, orole in members:
        cur = conn.execute(
            "INSERT INTO members(name, role, color, org_id, account_type, org_role)"
            " VALUES(?,?,?,?,?,?)", (name, role, color, org, acct, orole))
        mids.append(cur.lastrowid)
    # プロジェクト未アサインの外部ユーザー（アサイン操作のデモ用）
    conn.execute("INSERT INTO members(name, role, color, org_id, account_type)"
                 " VALUES(?,?,?,?,?)",
                 ("高橋 健", "エンジニア", "#06b6d4", org2, "external"))
    # ログインID用のダミーメール（migrate() と同形式）
    conn.execute("UPDATE members SET email='user'||id||'@example.com'"
                 " WHERE email IS NULL OR email=''")

    cf = json.dumps([
        {"key": "env", "label": "対象環境", "type": "select",
         "options": ["本番", "ステージング", "開発"]},
        {"key": "ticket", "label": "社内チケットNo", "type": "text", "options": []},
    ], ensure_ascii=False)
    cur = conn.execute(
        "INSERT INTO projects(name, description, color, start_date, end_date, custom_fields, created_at)"
        " VALUES(?,?,?,?,?,?,?)",
        ("社内ポータル刷新", "老朽化した社内ポータルのリニューアル案件。デモ用サンプルデータです。",
         "#4f6ef7", "2026-08-03", "2026-10-30", cf, t))
    pid = cur.lastrowid
    seed_roles = ["leader", "member", "member", "external"]
    for mid, role in zip(mids, seed_roles):
        conn.execute(
            "INSERT INTO project_members(project_id, member_id, role,"
            " can_view_comments, can_view_detail, added_at) VALUES(?,?,?,?,?,?)",
            (pid, mid, role, 0 if role == "external" else 1,
             0 if role == "external" else 1, t))
    notes = [
        ("環境", "検証環境一覧・アクセス方法",
         "■ ステージング\nURL: https://stg.portal.example.com\nSSH: stg-portal01 (踏み台 bastion01 経由)\nDB: portal-stg (RDS, 開発VPC内)\n\n■ 開発\nURL: http://dev.portal.example.local\n各自の開発機から直接アクセス可", 1),
        ("体制", "担当者・役職",
         "PM: 田中 太郎（情報システム部 課長）\nデザイン: 佐藤 花子\n実装リード: 鈴木 一郎\n実装（外部）: 山田 美咲（外部パートナー）\n\n顧客側窓口: 総務部 高橋様", 1),
        ("ルール", "定例・使用ツール",
         "定例MTG: 毎週月曜 10:00-10:30（会議室B / Teams併用）\n週報: 金曜17時までに本ツールの進捗を最新化\n連絡: Teams「ポータル刷新」チャネル\nソース管理: 社内GitLab portal-renewal リポジトリ", 0),
    ]
    for i, (cat, title, content, pin) in enumerate(notes):
        conn.execute(
            "INSERT INTO project_notes(project_id, category, title, content, pinned,"
            " sort_order, updated_by, created_at, updated_at) VALUES(?,?,?,?,?,?,?,?,?)",
            (pid, cat, title, content, pin, i, mids[0], t, t))

    sids = {}
    for name, color, order, is_done in DEFAULT_STATUSES:
        c = conn.execute(
            "INSERT INTO statuses(project_id, name, color, sort_order, is_done) VALUES(?,?,?,?,?)",
            (pid, name, color, order, is_done))
        sids[name] = c.lastrowid

    def task(title, status, assignee, priority, start, due, progress,
             parent=None, tags=None, desc="", est=None, mile=0, deps=None, cv=None):
        c = conn.execute(
            "INSERT INTO tasks(project_id, parent_id, title, description, status_id, assignee_id,"
            " priority, start_date, due_date, progress, estimate_h, milestone, tags, deps,"
            " custom_values, sort_order, created_at, updated_at)"
            " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (pid, parent, title, desc, sids[status],
             mids[assignee] if assignee is not None else None,
             priority, start, due, progress, est, mile,
             json.dumps(tags or [], ensure_ascii=False),
             json.dumps(deps or []),
             json.dumps(cv or {}, ensure_ascii=False),
             task.order, t, t))
        task.order += 1
        return c.lastrowid
    task.order = 0

    t1 = task("要件定義", "完了", 0, "high", "2026-08-03", "2026-08-14", 100,
              tags=["フェーズ1"], desc="現行ポータルの課題整理と新要件の確定。")
    t1a = task("現行システム棚卸し", "完了", 2, "medium", "2026-08-03", "2026-08-07", 100, parent=t1)
    t1b = task("ヒアリング・要件確定", "完了", 0, "high", "2026-08-08", "2026-08-14", 100, parent=t1)
    t2 = task("基本設計", "進行中", 2, "high", "2026-08-17", "2026-09-04", 60,
              tags=["フェーズ1"], deps=[t1], desc="画面設計・API設計・DB設計。")
    t2a = task("画面デザイン作成", "レビュー中", 1, "high", "2026-08-17", "2026-08-28", 80,
               parent=t2, tags=["デザイン"], cv={"env": "開発"})
    t2b = task("API設計書作成", "進行中", 2, "medium", "2026-08-20", "2026-09-04", 40, parent=t2)
    t3 = task("実装", "未着手", 3, "highest", "2026-09-07", "2026-10-09", 0,
              tags=["フェーズ2"], deps=[t2], est=160)
    task("認証基盤の実装", "未着手", 3, "highest", "2026-09-07", "2026-09-18", 0,
         parent=t3, cv={"env": "開発", "ticket": "IT-1024"})
    task("お知らせ機能の実装", "未着手", 2, "medium", "2026-09-14", "2026-09-30", 0, parent=t3)
    t4 = task("テスト・リリース", "未着手", 0, "high", "2026-10-12", "2026-10-30", 0,
              tags=["フェーズ2"], deps=[t3])
    task("リリース判定会", "未着手", 0, "high", "2026-10-28", "2026-10-28", 0,
         parent=t4, mile=1)

    conn.execute(
        "INSERT INTO comments(task_id, author_id, body, created_at) VALUES(?,?,?,?)",
        (t2a, 0 + mids[0], "トップページのワイヤー確認しました。ナビゲーションは現行踏襲でお願いします。",
         "2026-08-20 10:12:00"))
    conn.execute(
        "INSERT INTO comments(task_id, author_id, body, created_at) VALUES(?,?,?,?)",
        (t2a, mids[1], "承知しました。ヘッダー案を2パターン用意して明日共有します。",
         "2026-08-20 15:40:00"))
    conn.execute(
        "INSERT INTO task_links(task_id, title, url, kind) VALUES(?,?,?,?)",
        (t2a, "デザインFigma", "https://example.com/figma/portal", "design"))
    conn.execute(
        "INSERT INTO task_links(task_id, title, url, kind) VALUES(?,?,?,?)",
        (t2b, "API設計書（社内Wiki）", "https://example.com/wiki/api-spec", "doc"))
    conn.execute(
        "INSERT INTO activities(project_id, task_id, actor_id, action, detail, created_at)"
        " VALUES(?,?,?,?,?,?)",
        (pid, t2a, mids[1], "status", "進行中 → レビュー中", "2026-08-21 09:30:00"))


# ---------------------------------------------------------------- Pydantic models

class ProjectIn(BaseModel):
    name: str
    description: str = ""
    color: str = "#4f6ef7"
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    custom_fields: Optional[list] = None
    member_ids: list[int] = []      # 初期メンバー（作成者など）


class OrgIn(BaseModel):
    name: str
    color: str = "#64748b"


class MemberIn(BaseModel):
    name: str
    role: str = ""
    color: str = "#7c8db5"
    org_id: Optional[int] = None
    account_type: str = "internal"
    org_role: str = "staff"          # manager / site_admin / professional / staff
    email: str = ""                  # ログインID（必須・一意）。SSOヘッダー連携にも使用


class StatusIn(BaseModel):
    name: str
    color: str = "#8b95a7"
    sort_order: int = 0
    is_done: bool = False


class TaskIn(BaseModel):
    title: str
    description: str = ""
    status_id: Optional[int] = None
    assignee_id: Optional[int] = None
    priority: str = "medium"
    assignee_label: Optional[str] = None
    start_date: Optional[str] = None
    due_date: Optional[str] = None
    progress: int = 0
    estimate_h: Optional[float] = None
    actual_h: Optional[float] = None
    milestone: bool = False
    parent_id: Optional[int] = None
    tags: list = []
    deps: list = []
    custom_values: dict = {}
    recur: str = ""                  # '' / weekly / biweekly / monthly
    actor_id: Optional[int] = None


class TaskPatch(BaseModel):
    """部分更新。渡されたフィールドのみ反映。"""
    title: Optional[str] = None
    description: Optional[str] = None
    status_id: Optional[int] = None
    assignee_id: Optional[int] = None
    assignee_label: Optional[str] = None
    priority: Optional[str] = None
    start_date: Optional[str] = None
    due_date: Optional[str] = None
    progress: Optional[int] = None
    estimate_h: Optional[float] = None
    actual_h: Optional[float] = None
    milestone: Optional[bool] = None
    parent_id: Optional[int] = None
    tags: Optional[list] = None
    deps: Optional[list] = None
    custom_values: Optional[dict] = None
    sort_order: Optional[int] = None
    recur: Optional[str] = None
    expected_updated_at: Optional[str] = None   # 楽観ロック（一致しなければ409）
    actor_id: Optional[int] = None   # 変更者（アクティビティ記録用）


class CommentIn(BaseModel):
    body: str
    author_id: Optional[int] = None


class LinkIn(BaseModel):
    title: str
    url: str = ""
    kind: str = "link"


# ---------------------------------------------------------------- util

def task_row_to_dict(r) -> dict:
    d = dict(r)
    for k in ("tags", "deps"):
        d[k] = json.loads(d.get(k) or "[]")
    d["custom_values"] = json.loads(d.get("custom_values") or "{}")
    d["milestone"] = bool(d.get("milestone"))
    return d


def project_row_to_dict(r) -> dict:
    d = dict(r)
    d["custom_fields"] = json.loads(d.get("custom_fields") or "[]")
    d["settings"] = merge_settings(d.get("settings"))
    return d


def record_activity(conn, project_id, task_id, actor_id, action, detail=""):
    conn.execute(
        "INSERT INTO activities(project_id, task_id, actor_id, action, detail, created_at)"
        " VALUES(?,?,?,?,?,?)",
        (project_id, task_id, actor_id, action, detail, now()))


# ---------------------------------------------------------------- 通知・Webhook

FILES_DIR = os.path.join(DATA_DIR, "files")
BACKUP_DIR = os.path.join(DATA_DIR, "backups")
os.makedirs(FILES_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

ALLOWED_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg", ".pdf", ".txt", ".md",
                ".csv", ".xlsx", ".xls", ".docx", ".pptx", ".zip", ".log", ".json",
                ".yaml", ".yml", ".drawio", ".msg", ".eml"}
MAX_FILE_SIZE = 20 * 1024 * 1024   # 20MB


def get_user_pref(conn, member_id, key, default=None):
    r = conn.execute("SELECT value FROM user_prefs WHERE member_id=? AND key=?",
                     (member_id, key)).fetchone()
    if not r:
        return default
    try:
        return json.loads(r["value"])
    except ValueError:
        return default


def detect_webhook_provider(url: str) -> str:
    """Webhook URL から送信先サービスを判定する。"""
    u = (url or "").lower()
    if "hooks.slack.com" in u:
        return "slack"
    if "chat.googleapis.com" in u:
        return "googlechat"
    if "discord.com/api/webhooks" in u or "discordapp.com/api/webhooks" in u:
        return "discord"
    if ".webhook.office.com" in u or ".logic.azure.com" in u:
        return "teams"
    return "text"


def webhook_payload(provider: str, text: str) -> dict:
    """サービスごとの Incoming Webhook ペイロード。
    slack / googlechat / 汎用 = {"text"} ・ discord = {"content"} ・
    teams = Workflows 用 Adaptive Card。"""
    if provider == "discord":
        return {"content": text[:1900]}   # Discordは2000文字制限
    if provider == "teams":
        return {"type": "message", "attachments": [{
            "contentType": "application/vnd.microsoft.card.adaptive",
            "content": {
                "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
                "type": "AdaptiveCard", "version": "1.2",
                "body": [{"type": "TextBlock", "text": text, "wrap": True}],
            }}]}
    return {"text": text}   # slack / googlechat / 汎用


def post_webhook_async(url, text, provider=None):
    """Slack / Google Chat / Teams / Discord 等の Incoming Webhook へ非同期送信
    （失敗しても本処理を止めない）。provider 未指定・auto はURLから自動判定。"""
    if not provider or provider == "auto":
        provider = detect_webhook_provider(url)
    import threading
    import urllib.request

    def _post():
        try:
            body = json.dumps(webhook_payload(provider, text), ensure_ascii=False).encode()
            req = urllib.request.Request(
                url, data=body, headers={"Content-Type": "application/json"})
            urllib.request.urlopen(req, timeout=10)
        except Exception:
            pass

    threading.Thread(target=_post, daemon=True).start()


NOTIFY_TYPE_LABEL = {"mention": "メンション", "assign": "担当割当", "comment": "コメント",
                     "status": "ステータス変更", "due": "期限リマインド",
                     "watch": "ウォッチ", "system": "システム"}


def notify(conn, user_id, ntype, project_id=None, task_id=None, actor_id=None, message=""):
    """アプリ内通知。自分自身の操作は通知しない。
    ユーザー設定（notify_webhook）があれば外部チャット（Teams/Slack DM等）へも転送する。"""
    if user_id is None or user_id == actor_id:
        return
    conn.execute(
        "INSERT INTO notifications(user_id, type, project_id, task_id, actor_id,"
        " message, created_at) VALUES(?,?,?,?,?,?,?)",
        (user_id, ntype, project_id, task_id, actor_id, message, now()))
    # 個人設定の外部転送（ユーザー設定ページで登録した Incoming Webhook 宛て）
    try:
        pw = get_user_pref(conn, user_id, "notify_webhook") or {}
        url = (pw.get("url") or "").strip()
        events = pw.get("events")
        if pw.get("enabled") and url and (not events or ntype in events):
            pname = ""
            if project_id:
                pr = conn.execute("SELECT name FROM projects WHERE id=?",
                                  (project_id,)).fetchone()
                pname = f"［{pr['name']}］" if pr else ""
            post_webhook_async(
                url, f"🔔 PJ Board {NOTIFY_TYPE_LABEL.get(ntype, ntype)}{pname} {message}",
                provider=pw.get("provider"))
    except Exception:
        pass


def send_webhook(conn, project_id, event, text):
    """PJ設定の Teams/Slack Incoming Webhook へ非同期送信（設定があり、対象イベントの場合のみ）。"""
    s = get_settings(conn, project_id)
    url = (s.get("webhook_url") or "").strip()
    if not url or event not in (s.get("webhook_events") or []):
        return
    post_webhook_async(url, text)


def norm_name(s: str) -> str:
    return (s or "").replace(" ", "").replace("　", "").lower()


def mentioned_user_ids(conn, body: str) -> list[int]:
    """コメント本文の @名前 からユーザーIDを解決する。"""
    ids = []
    text = norm_name(body)
    for r in conn.execute("SELECT id, name FROM members WHERE active=1"):
        if "@" + norm_name(r["name"]) in text:
            ids.append(r["id"])
    return ids


def watchers_of(conn, task_id) -> set[int]:
    return {r["member_id"] for r in conn.execute(
        "SELECT member_id FROM watchers WHERE task_id=?", (task_id,))}


def get_task_or_404(conn, task_id) -> dict:
    r = conn.execute(
        "SELECT * FROM tasks WHERE id=? AND deleted_at IS NULL", (task_id,)).fetchone()
    if not r:
        raise HTTPException(404, "task not found")
    return task_row_to_dict(r)


# ---------------------------------------------------------------- 認証（セッション）
# ログイン前提の運用。Cookie セッションでログインユーザーを特定する。
# 未ログイン（Cookieなし）のアクセスは従来どおり素通し＝スクリプト・MCP向けの緩め運用。

CURRENT_USER: contextvars.ContextVar = contextvars.ContextVar("current_user", default=None)

ORG_RANK = {"manager": 4, "site_admin": 3, "professional": 2, "staff": 1}


def org_rank(u) -> int:
    return ORG_RANK.get(((u or {}).get("org_role")) or "staff", 1)


def hash_password(password: str) -> str:
    salt = secrets.token_hex(8)
    h = hashlib.sha256((salt + password).encode()).hexdigest()
    return f"{salt}${h}"


def verify_password(stored: Optional[str], password: str) -> bool:
    if not stored:
        return True          # 初期状態（パスワード未設定）はそのままログイン可
    salt, h = stored.split("$", 1)
    return hashlib.sha256((salt + password).encode()).hexdigest() == h


SSO_EMAIL_HEADER = os.environ.get("PJBOARD_SSO_EMAIL_HEADER", "")  # 例: x-amzn-oidc-identity


@app.middleware("http")
async def session_middleware(request: Request, call_next):
    user = None
    token = request.cookies.get("pjb_session")
    if token:
        with db() as conn:
            r = conn.execute(
                "SELECT m.* FROM sessions s JOIN members m ON m.id=s.member_id"
                " WHERE s.token=? AND m.active=1", (token,)).fetchone()
            if r:
                user = dict(r)
                conn.execute("UPDATE sessions SET last_seen=? WHERE token=?",
                             (now(), token))
    # APIトークン（Authorization: Bearer）— 自動化・MCP向けの正規経路
    if user is None:
        auth = request.headers.get("authorization", "")
        if auth.startswith("Bearer "):
            th = hashlib.sha256(auth[7:].strip().encode()).hexdigest()
            with db() as conn:
                r = conn.execute(
                    "SELECT m.* FROM api_tokens t JOIN members m ON m.id=t.member_id"
                    " WHERE t.token_hash=? AND m.active=1", (th,)).fetchone()
                if r:
                    user = dict(r)
                    conn.execute("UPDATE api_tokens SET last_used=? WHERE token_hash=?",
                                 (now(), th))
    # SSOヘッダー連携（ALB+Cognito等。メールアドレスでユーザー特定）
    if user is None and SSO_EMAIL_HEADER:
        email = request.headers.get(SSO_EMAIL_HEADER)
        if email:
            with db() as conn:
                r = conn.execute(
                    "SELECT * FROM members WHERE email=? AND active=1", (email,)).fetchone()
                if r:
                    user = dict(r)
    # CSRF防御: Cookieセッションでの書き込みはfetch由来ヘッダーを必須にする
    # （SameSite=Lax と二重化。Bearer/未ログインには適用しない）
    if (token and user is not None
            and request.method in ("POST", "PATCH", "PUT", "DELETE")
            and request.headers.get("x-requested-with") != "fetch"):
        return Response(json.dumps({"detail": "CSRF: X-Requested-With ヘッダーが必要です"}),
                        status_code=403, media_type="application/json")
    cv_token = CURRENT_USER.set(user)
    try:
        return await call_next(request)
    finally:
        CURRENT_USER.reset(cv_token)


def resolve_uid(claimed: Optional[int]) -> Optional[int]:
    """閲覧/操作ユーザーIDを確定する。
    - 未ログイン: 渡された値をそのまま信用（緩め・自動化向け）
    - ログイン中: 原則自分自身。別ユーザーの指定（デバッグ偽装）はサイト管理者以上のみ。"""
    su = CURRENT_USER.get()
    if su is None:
        return claimed
    if claimed is None or claimed == su["id"]:
        return su["id"]
    if org_rank(su) >= ORG_RANK["site_admin"]:
        return claimed
    raise HTTPException(403, "別ユーザーとしての閲覧・操作はサイト管理者以上のデバッグ機能です")


def check_site_admin(msg="この操作にはサイト管理者以上の権限が必要です"):
    su = CURRENT_USER.get()
    if su is not None and org_rank(su) < ORG_RANK["site_admin"]:
        raise HTTPException(403, msg)


# ---------------------------------------------------------------- 権限
# 【2層モデル】
#  組織ロール(members.org_role): manager / professional / staff
#    → manager・professional は全プロジェクトに「暗黙の管理者」（アサイン表示は不要）
#  プロジェクトロール(project_members.role): leader / member / advisor(ご意見番) / external
#
# 実効権限(effective role):
#   admin   = 組織のmanager/professional、またはPJのleader … 全操作
#   member  = PJのmember … 自分の担当タスクの状態・進捗等のみ（日程はPJ設定で緩和可）
#   advisor = PJのadvisor(ご意見番) … 閲覧＋コメント
#   viewer  = 未アサインの社内ユーザー … 全PJ閲覧可＋コメント可（PJ設定でOFF可）←横断的に意見を聞ける肝
#   external= 外部アカウント … アサインされたPJのみ・閲覧範囲もフラグで制限
# 方針は「緩め」: actor_id が渡されない場合（スクリプト・API直叩き）はチェックしない。

MEMBER_EDITABLE_FIELDS = {"status_id", "progress", "actual_h", "description",
                          "tags", "custom_values", "title"}
MEMBER_SCHEDULE_FIELDS = {"start_date", "due_date", "estimate_h"}


def get_membership(conn, pid, uid):
    if uid is None:
        return None
    return conn.execute(
        "SELECT * FROM project_members WHERE project_id=? AND member_id=?",
        (pid, uid)).fetchone()


def get_settings(conn, pid) -> dict:
    r = conn.execute("SELECT settings FROM projects WHERE id=?", (pid,)).fetchone()
    return merge_settings(r["settings"] if r else None)


def effective_role(conn, pid, uid) -> Optional[str]:
    """組織ロール＋プロジェクトロールから実効権限を求める。uid=None は素通し(admin扱い)。"""
    if uid is None:
        return "admin"
    m = conn.execute("SELECT * FROM members WHERE id=?", (uid,)).fetchone()
    if not m:
        return None
    pm = get_membership(conn, pid, uid)
    if m["account_type"] == "external":
        return "external" if pm else None      # 未アサインの外部はアクセス不可
    if (m["org_role"] or "staff") in ("manager", "professional"):
        return "admin"                          # 暗黙の管理者
    if pm:
        return {"leader": "admin", "member": "member",
                "advisor": "advisor"}.get(pm["role"], "member")
    return "viewer"                             # 社内・未アサイン＝閲覧＋コメント


def can_comment_in(conn, pid, uid) -> bool:
    role = effective_role(conn, pid, uid)
    if role in ("admin", "member"):
        return True
    s = get_settings(conn, pid)
    if role == "advisor":
        return bool(s.get("advisor_can_comment", True))
    if role == "viewer":
        return bool(s.get("unassigned_can_comment", True))
    if role == "external":
        pm = get_membership(conn, pid, uid)
        return bool(pm and pm["can_view_comments"])
    return False


def check_admin(conn, pid, actor_id, msg):
    if actor_id is None:
        return
    if effective_role(conn, pid, actor_id) != "admin":
        raise HTTPException(403, msg)


def check_role(conn, pid, actor_id, allowed: set, msg: str):
    """actor_id が指定されている場合のみ、実効権限が allowed に含まれるか検査する。"""
    if actor_id is None:
        return
    if effective_role(conn, pid, actor_id) not in allowed:
        raise HTTPException(403, msg)


# --- 外部ユーザー防御（緩め運用の例外: 外部アカウントには常に厳格に適用する） ---

TAB_LABELS = {"dashboard": "ダッシュボード", "board": "ボード", "table": "テーブル",
              "gantt": "WBSガント", "calendar": "カレンダー", "qa": "QA",
              "kadai": "課題", "issues": "コメント一覧", "notes": "ノート"}


def external_guard(pid: int, tab: Optional[str] = None):
    """ログイン中の外部ユーザーに対する防御。
    - 未アサインのプロジェクトへのアクセスを拒否
    - tab 指定時は、PJ設定 external_visible_tabs に無い機能を拒否
    社内ユーザー・未ログイン（自動化）はここでは制限しない。"""
    su = CURRENT_USER.get()
    if su is None or su.get("account_type") != "external":
        return
    with db() as conn:
        if not get_membership(conn, pid, su["id"]):
            raise HTTPException(403, "このプロジェクトへのアクセス権がありません")
        if tab:
            tabs = get_settings(conn, pid).get("external_visible_tabs") or []
            if tab not in tabs:
                raise HTTPException(
                    403, f"このプロジェクトでは外部ユーザーに「{TAB_LABELS.get(tab, tab)}」は公開されていません")


def _actor_rank(conn, actor_id) -> Optional[int]:
    """操作者の組織ランク。ログイン中はセッションを優先。None=素通し（自動化）。"""
    su = CURRENT_USER.get()
    if su is not None:
        return org_rank(su)
    if actor_id is None:
        return None
    r = conn.execute("SELECT * FROM members WHERE id=?", (actor_id,)).fetchone()
    return org_rank(dict(r)) if r else 1


# ---------------------------------------------------------------- API: auth

class LoginIn(BaseModel):
    email: str
    password: str = ""


@app.get("/api/auth/config")
def auth_config():
    """ログイン画面用の公開設定（デバッグ機能の有無）。"""
    return {"debug": DEBUG_FEATURES}


@app.post("/api/auth/login")
def auth_login(body: LoginIn, request: Request, response: Response):
    ip = request.client.host if request.client else ""
    ua = request.headers.get("user-agent", "")[:200]
    email = body.email.strip().lower()
    if not email:
        raise HTTPException(400, "メールアドレスを入力してください")
    with db() as conn:
        m = conn.execute(
            "SELECT * FROM members WHERE lower(email)=? AND active=1",
            (email,)).fetchone()
        if not m:
            # 存在有無を悟らせないためパスワード誤りと同じメッセージにする
            raise HTTPException(401, "メールアドレスまたはパスワードが違います")
        fails = conn.execute(
            "SELECT COUNT(*) c FROM login_logs WHERE member_id=? AND success=0"
            " AND created_at > datetime('now', 'localtime', '-10 minutes')",
            (m["id"],)).fetchone()["c"]
        if fails >= 5:
            raise HTTPException(429, "ログイン試行が多すぎます。10分後に再試行してください")
        if not verify_password(m["password_hash"], body.password):
            conn.execute(
                "INSERT INTO login_logs(member_id, name, success, ip, ua, created_at)"
                " VALUES(?,?,?,?,?,?)", (m["id"], m["name"], 0, ip, ua, now()))
            conn.commit()
            raise HTTPException(401, "メールアドレスまたはパスワードが違います")
        conn.execute(
            "INSERT INTO login_logs(member_id, name, success, ip, ua, created_at)"
            " VALUES(?,?,?,?,?,?)", (m["id"], m["name"], 1, ip, ua, now()))
        token = secrets.token_hex(24)
        conn.execute("INSERT INTO sessions(token, member_id, created_at, last_seen)"
                     " VALUES(?,?,?,?)", (token, m["id"], now(), now()))
    response.set_cookie("pjb_session", token, httponly=True, samesite="lax",
                        max_age=60 * 60 * 24 * 30)
    d = dict(m)
    d.pop("password_hash", None)
    return {"user": d, "needs_password": m["password_hash"] is None}


@app.post("/api/auth/logout")
def auth_logout(request: Request, response: Response):
    token = request.cookies.get("pjb_session")
    if token:
        with db() as conn:
            conn.execute("DELETE FROM sessions WHERE token=?", (token,))
    response.delete_cookie("pjb_session")
    return {"ok": True}


@app.post("/api/auth/debug-login")
def auth_debug_login(body: dict, request: Request, response: Response):
    """開発時専用: ログインユーザーの完全切替（パスワード不要・全ユーザー可）。
    member_id か name で対象を指定。ログイン画面のデバッグボタンからも使うため
    未ログインでも呼び出せる。セッションごと張り替えるため、以後は完全に
    そのユーザーとして動作する。
    ※本番運用では PJBOARD_DEBUG=0 を設定してこの機能を無効化すること。"""
    if not DEBUG_FEATURES:
        raise HTTPException(403, "本番環境ではデバッグ切替は無効化されています")
    mid = body.get("member_id")
    name = body.get("name")
    with db() as conn:
        if mid is not None:
            m = conn.execute("SELECT * FROM members WHERE id=? AND active=1",
                             (mid,)).fetchone()
        elif name:
            m = conn.execute("SELECT * FROM members WHERE name=? AND active=1",
                             (name,)).fetchone()
        else:
            raise HTTPException(400, "member_id か name を指定してください")
        if not m:
            raise HTTPException(404, "ユーザーが見つかりません")
        old_token = request.cookies.get("pjb_session")
        if old_token:
            conn.execute("DELETE FROM sessions WHERE token=?", (old_token,))
        token = secrets.token_hex(24)
        conn.execute("INSERT INTO sessions(token, member_id, created_at, last_seen)"
                     " VALUES(?,?,?,?)", (token, m["id"], now(), now()))
    response.set_cookie("pjb_session", token, httponly=True, samesite="lax",
                        max_age=60 * 60 * 24 * 30)
    d = dict(m)
    d.pop("password_hash", None)
    return {"user": d}


@app.get("/api/auth/me")
def auth_me():
    su = CURRENT_USER.get()
    if su is None:
        raise HTTPException(401, "not logged in")
    d = dict(su)
    d.pop("password_hash", None)
    d["org_rank"] = org_rank(su)
    d["debug_enabled"] = DEBUG_FEATURES
    return d


class PasswordIn(BaseModel):
    current_password: str = ""
    new_password: str


@app.post("/api/auth/password")
def change_own_password(body: PasswordIn, request: Request):
    su = CURRENT_USER.get()
    if su is None:
        raise HTTPException(401, "ログインが必要です")
    with db() as conn:
        m = conn.execute("SELECT * FROM members WHERE id=?", (su["id"],)).fetchone()
        if not verify_password(m["password_hash"], body.current_password):
            raise HTTPException(401, "現在のパスワードが違います")
        conn.execute("UPDATE members SET password_hash=? WHERE id=?",
                     (hash_password(body.new_password) if body.new_password else None,
                      su["id"]))
        # 全端末ログアウト（現在のセッションだけ残す）
        cur = request.cookies.get("pjb_session") or ""
        conn.execute("DELETE FROM sessions WHERE member_id=? AND token != ?",
                     (su["id"], cur))
    return {"ok": True}


@app.post("/api/members/{mid}/reset-password")
def reset_password(mid: int, body: dict):
    """サイト管理者以上によるパスワードリセット。new_password 空で「未設定」に戻す。"""
    check_site_admin("パスワードのリセットはサイト管理者以上のみ可能です")
    new = body.get("new_password") or ""
    with db() as conn:
        conn.execute("UPDATE members SET password_hash=? WHERE id=?",
                     (hash_password(new) if new else None, mid))
        conn.execute("DELETE FROM sessions WHERE member_id=?", (mid,))
    return {"ok": True}


# ---------------------------------------------------------------- API: bootstrap

@app.get("/api/bootstrap")
def bootstrap(user_id: Optional[int] = None):
    user_id = resolve_uid(user_id)
    with db() as conn:
        users = rows_to_dicts(conn.execute(
            "SELECT * FROM members WHERE active=1 ORDER BY id"))
        me = next((u for u in users if u["id"] == user_id), None)
        if me and me["account_type"] == "external":
            # 外部ユーザーにはアサインされたプロジェクトのみ見せる
            projects = [project_row_to_dict(r) for r in conn.execute(
                "SELECT p.* FROM projects p"
                " JOIN project_members pm ON pm.project_id=p.id AND pm.member_id=?"
                " ORDER BY p.id", (user_id,))]
        else:
            projects = [project_row_to_dict(r) for r in conn.execute(
                "SELECT * FROM projects ORDER BY id")]
        orgs = rows_to_dicts(conn.execute("SELECT * FROM orgs ORDER BY id"))
        pm = rows_to_dicts(conn.execute(
            "SELECT project_id, member_id, role, can_view_comments, can_view_detail"
            " FROM project_members"))
    return {"projects": projects, "users": users, "orgs": orgs,
            "project_members": pm}


# ---------------------------------------------------------------- API: projects

@app.post("/api/projects")
def create_project(p: ProjectIn):
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO projects(name, description, color, start_date, end_date,"
            " custom_fields, created_at) VALUES(?,?,?,?,?,?,?)",
            (p.name, p.description, p.color, p.start_date, p.end_date,
             json.dumps(p.custom_fields or [], ensure_ascii=False), now()))
        pid = cur.lastrowid
        for name, color, order, is_done in DEFAULT_STATUSES:
            conn.execute(
                "INSERT INTO statuses(project_id, name, color, sort_order, is_done)"
                " VALUES(?,?,?,?,?)", (pid, name, color, order, is_done))
        for i, mid in enumerate(p.member_ids):
            conn.execute(
                "INSERT OR IGNORE INTO project_members(project_id, member_id, role, added_at)"
                " VALUES(?,?,?,?)", (pid, mid, "leader" if i == 0 else "member", now()))
        apply_note_templates(conn, pid, p.member_ids[0] if p.member_ids else None)
        r = conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    return project_row_to_dict(r)


EXTERNAL_SETTING_KEYS = ("external_visible_tabs", "external_can_export",
                         "external_default_view_comments", "external_default_view_detail")


@app.patch("/api/projects/{pid}")
def update_project(pid: int, p: dict, actor_id: Optional[int] = None):
    actor_id = resolve_uid(actor_id)
    with db() as conn:
        check_admin(conn, pid, actor_id,
                    "プロジェクト設定の変更はリーダーまたは組織の上位者のみ行えます")
        # 外部パートナー公開設定の変更は「マネージャー／サイト管理者＋プロジェクト名の確認入力」を常に必須にする
        # （チェックボックス1つで情報が漏れないよう、堅牢な二重確認）
        if "settings" in p:
            pr0 = conn.execute("SELECT name, settings FROM projects WHERE id=?",
                               (pid,)).fetchone()
            if not pr0:
                raise HTTPException(404, "project not found")
            old_s = merge_settings(pr0["settings"])
            new_s = merge_settings(json.dumps(p["settings"], ensure_ascii=False))
            changed = [k for k in EXTERNAL_SETTING_KEYS if old_s.get(k) != new_s.get(k)]
            if changed:
                rank = _actor_rank(conn, actor_id)
                if rank is not None and rank < ORG_RANK["site_admin"]:
                    raise HTTPException(
                        403, "外部パートナーの公開設定の変更はマネージャー／サイト管理者のみ行えます")
                if (p.get("confirm_text") or "").strip() != pr0["name"]:
                    raise HTTPException(
                        400, "外部公開設定の変更には、確認のためプロジェクト名の入力が必要です")
                record_activity(conn, pid, None, actor_id, "settings",
                                f"外部公開設定を変更: {', '.join(changed)}")
    allowed = {"name", "description", "color", "status", "start_date", "end_date",
               "custom_fields", "settings"}
    sets, vals = [], []
    for k, v in p.items():
        if k not in allowed:
            continue
        if k in ("custom_fields", "settings"):
            v = json.dumps(v, ensure_ascii=False)
        sets.append(f"{k}=?")
        vals.append(v)
    if not sets:
        raise HTTPException(400, "no valid fields")
    vals.append(pid)
    with db() as conn:
        conn.execute(f"UPDATE projects SET {', '.join(sets)} WHERE id=?", vals)
        r = conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
        if not r:
            raise HTTPException(404, "project not found")
    return project_row_to_dict(r)


@app.delete("/api/projects/{pid}")
def delete_project(pid: int):
    with db() as conn:
        conn.execute("DELETE FROM projects WHERE id=?", (pid,))
    return {"ok": True}


@app.get("/api/projects/{pid}/data")
def project_data(pid: int, user_id: Optional[int] = None):
    """1プロジェクト分の全ビュー用データをまとめて返す。"""
    external_guard(pid)
    user_id = resolve_uid(user_id)
    with db() as conn:
        pr = conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
        if not pr:
            raise HTTPException(404, "project not found")
        statuses = rows_to_dicts(conn.execute(
            "SELECT * FROM statuses WHERE project_id=? ORDER BY sort_order, id", (pid,)))
        tasks = [task_row_to_dict(r) for r in conn.execute(
            "SELECT * FROM tasks WHERE project_id=? AND deleted_at IS NULL"
            " ORDER BY sort_order, id", (pid,))]
        members = rows_to_dicts(conn.execute(
            "SELECT m.* FROM members m"
            " JOIN project_members pm ON pm.member_id=m.id AND pm.project_id=?"
            " WHERE m.active=1 ORDER BY m.id", (pid,)))
        counts = {r["task_id"]: r["c"] for r in conn.execute(
            "SELECT task_id, COUNT(*) c FROM comments"
            " WHERE task_id IN (SELECT id FROM tasks WHERE project_id=?)"
            " GROUP BY task_id", (pid,))}
        link_counts = {r["task_id"]: r["c"] for r in conn.execute(
            "SELECT task_id, COUNT(*) c FROM task_links"
            " WHERE task_id IN (SELECT id FROM tasks WHERE project_id=?)"
            " GROUP BY task_id", (pid,))}
        # タスクに紐づくオープン課題数（クローズ済みは数えない）
        issue_counts = {r["task_id"]: r["c"] for r in conn.execute(
            "SELECT it.task_id, COUNT(*) c FROM issue_tasks it"
            " JOIN issues i ON i.id=it.issue_id"
            " WHERE i.project_id=? AND i.status != 'closed'"
            " GROUP BY it.task_id", (pid,))}
        activities = rows_to_dicts(conn.execute(
            "SELECT a.*, m.name actor_name, t.title task_title FROM activities a"
            " LEFT JOIN members m ON m.id=a.actor_id"
            " LEFT JOIN tasks t ON t.id=a.task_id"
            " WHERE a.project_id=? ORDER BY a.id DESC LIMIT 30", (pid,)))
        membership = get_membership(conn, pid, user_id)
        my_role = effective_role(conn, pid, user_id) if user_id else None
        my_can_comment = can_comment_in(conn, pid, user_id) if user_id else True
    for t in tasks:
        t["comment_count"] = counts.get(t["id"], 0)
        t["link_count"] = link_counts.get(t["id"], 0)
        t["issue_count"] = issue_counts.get(t["id"], 0)

    my_flags = {"can_view_comments": 1, "can_view_detail": 1}
    my_project_role = membership["role"] if membership else None
    if membership:
        my_flags = {"can_view_comments": membership["can_view_comments"],
                    "can_view_detail": membership["can_view_detail"]}
    # 外部ユーザーの閲覧制限: コメント不可視なら件数も見せない。
    # 詳細不可視なら説明文・アクティビティも隠す（WBS・進捗・担当は見える）
    if my_role == "external":
        if not my_flags["can_view_comments"]:
            for t in tasks:
                t["comment_count"] = 0
            activities = [a for a in activities if a["action"] != "comment"]
        if not my_flags["can_view_detail"]:
            for t in tasks:
                t["description"] = ""
            activities = []
        # タブ単位の公開範囲: タスク系タブが1つも公開されていなければタスクを返さない
        su = CURRENT_USER.get()
        if su is not None and su.get("account_type") == "external":
            with db() as conn:
                tabs = get_settings(conn, pid).get("external_visible_tabs") or []
            if not any(t2 in tabs for t2 in
                       ("dashboard", "board", "table", "gantt", "calendar")):
                tasks = []
                activities = []
    return {"project": project_row_to_dict(pr), "statuses": statuses,
            "tasks": tasks, "members": members, "activities": activities,
            "my_role": my_role, "my_project_role": my_project_role,
            "my_flags": my_flags, "my_can_comment": my_can_comment}


# ---------------------------------------------------------------- API: orgs

@app.post("/api/orgs")
def create_org(o: OrgIn):
    check_site_admin()
    with db() as conn:
        cur = conn.execute("INSERT INTO orgs(name, color, created_at) VALUES(?,?,?)",
                           (o.name, o.color, now()))
        r = conn.execute("SELECT * FROM orgs WHERE id=?", (cur.lastrowid,)).fetchone()
    return dict(r)


@app.patch("/api/orgs/{oid}")
def update_org(oid: int, o: dict):
    check_site_admin()
    allowed = {"name", "color"}
    sets = [f"{k}=?" for k in o if k in allowed]
    vals = [o[k] for k in o if k in allowed] + [oid]
    if not sets:
        raise HTTPException(400, "no valid fields")
    with db() as conn:
        conn.execute(f"UPDATE orgs SET {', '.join(sets)} WHERE id=?", vals)
        r = conn.execute("SELECT * FROM orgs WHERE id=?", (oid,)).fetchone()
        if not r:
            raise HTTPException(404, "org not found")
    return dict(r)


@app.delete("/api/orgs/{oid}")
def delete_org(oid: int):
    check_site_admin()
    with db() as conn:
        used = conn.execute(
            "SELECT COUNT(*) c FROM members WHERE org_id=? AND active=1",
            (oid,)).fetchone()["c"]
        if used:
            raise HTTPException(400, f"この組織には {used} 名のユーザーが所属しています")
        conn.execute("DELETE FROM orgs WHERE id=?", (oid,))
    return {"ok": True}


# ---------------------------------------------------------------- API: project members

@app.post("/api/projects/{pid}/members")
def assign_member(pid: int, body: dict):
    body["actor_id"] = resolve_uid(body.get("actor_id"))
    mid = body.get("member_id")
    if not mid:
        raise HTTPException(400, "member_id required")
    with db() as conn:
        if not conn.execute("SELECT 1 FROM projects WHERE id=?", (pid,)).fetchone():
            raise HTTPException(404, "project not found")
        check_admin(conn, pid, body["actor_id"],
                    "メンバーのアサインはリーダーまたは組織の上位者のみ行えます")
        m = conn.execute("SELECT * FROM members WHERE id=? AND active=1", (mid,)).fetchone()
        if not m:
            raise HTTPException(404, "user not found")
        # 外部アカウントは external ロール固定（社内ロールとの入替不可）、社内既定は member
        s = get_settings(conn, pid)
        is_ext = m["account_type"] == "external"
        # 外部ユーザーのアサインは誤操作防止のため二重の保護:
        #  1) マネージャー／サイト管理者のみ  2) 氏名の完全一致入力（confirm_name）を常に必須
        if is_ext:
            rank = _actor_rank(conn, body["actor_id"])
            if rank is not None and rank < ORG_RANK["site_admin"]:
                raise HTTPException(
                    403, "外部ユーザーのアサインはマネージャー／サイト管理者のみ、管理画面から行えます")
            if (body.get("confirm_name") or "").strip() != m["name"]:
                raise HTTPException(
                    400, "確認のため、アサインする外部ユーザーの氏名を正確に入力してください")
        role = body.get("role") or ("external" if is_ext else "member")
        if is_ext and role != "external":
            raise HTTPException(400, "外部アカウントには external 以外のロールを設定できません")
        if not is_ext and role == "external":
            raise HTTPException(400, "社内アカウントに external ロールは設定できません")
        cvc = body.get("can_view_comments",
                       1 if not is_ext else int(bool(s.get("external_default_view_comments"))))
        cvd = body.get("can_view_detail",
                       1 if not is_ext else int(bool(s.get("external_default_view_detail"))))
        conn.execute(
            "INSERT OR IGNORE INTO project_members(project_id, member_id, role,"
            " can_view_comments, can_view_detail, added_at) VALUES(?,?,?,?,?,?)",
            (pid, mid, role, cvc, cvd, now()))
        record_activity(conn, pid, None, body.get("actor_id"), "member_add", m["name"])
    return {"ok": True}


@app.patch("/api/projects/{pid}/members/{mid}")
def update_project_member(pid: int, mid: int, body: dict):
    """ロール・閲覧フラグの変更（外部の閲覧制限はここ＝PJ管理からのみ）。"""
    body["actor_id"] = resolve_uid(body.get("actor_id"))
    with db() as conn:
        check_admin(conn, pid, body.get("actor_id"),
                    "ロール変更はリーダーまたは組織の上位者のみ行えます")
        m = conn.execute("SELECT account_type FROM members WHERE id=?",
                         (mid,)).fetchone()
        is_ext = m and m["account_type"] == "external"
        # 外部ユーザーの閲覧フラグ緩和はマネージャー／サイト管理者のみ
        if is_ext and ("can_view_comments" in body or "can_view_detail" in body):
            rank = _actor_rank(conn, body.get("actor_id"))
            if rank is not None and rank < ORG_RANK["site_admin"]:
                raise HTTPException(
                    403, "外部ユーザーの閲覧範囲の変更はマネージャー／サイト管理者のみ行えます")
        if "role" in body:
            if is_ext and body["role"] != "external":
                raise HTTPException(400, "外部アカウントには external 以外のロールを設定できません")
            if not is_ext and body["role"] == "external":
                raise HTTPException(400, "社内アカウントに external ロールは設定できません")
        allowed = {"role", "can_view_comments", "can_view_detail"}
        sets = [f"{k}=?" for k in body if k in allowed]
        vals = [body[k] for k in body if k in allowed]
        if not sets:
            raise HTTPException(400, "no valid fields")
        vals += [pid, mid]
        conn.execute(
            f"UPDATE project_members SET {', '.join(sets)}"
            " WHERE project_id=? AND member_id=?", vals)
        r = conn.execute(
            "SELECT * FROM project_members WHERE project_id=? AND member_id=?",
            (pid, mid)).fetchone()
        if not r:
            raise HTTPException(404, "membership not found")
    return dict(r)


@app.delete("/api/projects/{pid}/members/{mid}")
def unassign_member(pid: int, mid: int):
    actor = resolve_uid(None)
    with db() as conn:
        check_admin(conn, pid, actor,
                    "メンバーの解除はリーダーまたは組織の上位者のみ行えます")
        m = conn.execute("SELECT * FROM members WHERE id=?", (mid,)).fetchone()
        conn.execute(
            "DELETE FROM project_members WHERE project_id=? AND member_id=?", (pid, mid))
        # 担当タスクは残す（表示は全ユーザー辞書で解決）。必要なら手動で付け替え
        if m:
            record_activity(conn, pid, None, None, "member_remove", m["name"])
    return {"ok": True}


# ---------------------------------------------------------------- API: project notes（ルール・メモ共有）

class NoteIn(BaseModel):
    title: str
    content: str = ""
    category: str = "その他"
    pinned: bool = False
    actor_id: Optional[int] = None


@app.get("/api/projects/{pid}/notes")
def list_notes(pid: int):
    external_guard(pid, "notes")
    with db() as conn:
        rows = conn.execute(
            "SELECT n.*, m.name updated_by_name FROM project_notes n"
            " LEFT JOIN members m ON m.id=n.updated_by"
            " WHERE n.project_id=? AND n.deleted_at IS NULL"
            " ORDER BY n.pinned DESC, n.sort_order, n.id", (pid,)).fetchall()
        out = rows_to_dicts(rows)
        for n in out:
            n["attachments"] = attachments_for(conn, "note", n["id"])
    return out


@app.post("/api/projects/{pid}/notes")
def create_note(pid: int, n: NoteIn):
    n.actor_id = resolve_uid(n.actor_id)
    with db() as conn:
        n_allowed = {"admin", "member"} if get_settings(conn, pid).get("member_can_edit_notes", True) else {"admin"}
        check_role(conn, pid, n.actor_id, n_allowed,
                   "このプロジェクトでノートを編集する権限がありません")
        cur = conn.execute(
            "INSERT INTO project_notes(project_id, category, title, content, pinned,"
            " sort_order, updated_by, created_at, updated_at)"
            " VALUES(?,?,?,?,?,"
            " (SELECT COALESCE(MAX(sort_order),-1)+1 FROM project_notes WHERE project_id=?),"
            " ?,?,?)",
            (pid, n.category, n.title, n.content, int(n.pinned), pid,
             n.actor_id, now(), now()))
        r = conn.execute("SELECT * FROM project_notes WHERE id=?",
                         (cur.lastrowid,)).fetchone()
    return dict(r)


@app.patch("/api/notes/{nid}")
def update_note(nid: int, body: dict):
    body["actor_id"] = resolve_uid(body.get("actor_id"))
    with db() as conn:
        old = conn.execute("SELECT * FROM project_notes WHERE id=?", (nid,)).fetchone()
        if not old:
            raise HTTPException(404, "note not found")
        actor = body.get("actor_id")
        n_allowed = {"admin", "member"} if get_settings(conn, old["project_id"]).get("member_can_edit_notes", True) else {"admin"}
        check_role(conn, old["project_id"], actor, n_allowed,
                   "このプロジェクトでノートを編集する権限がありません")
        allowed = {"title", "content", "category", "pinned", "sort_order"}
        sets = [f"{k}=?" for k in body if k in allowed]
        vals = [body[k] for k in body if k in allowed]
        if not sets:
            raise HTTPException(400, "no valid fields")
        sets += ["updated_by=?", "updated_at=?"]
        vals += [actor, now(), nid]
        conn.execute(f"UPDATE project_notes SET {', '.join(sets)} WHERE id=?", vals)
        r = conn.execute("SELECT * FROM project_notes WHERE id=?", (nid,)).fetchone()
    return dict(r)


@app.delete("/api/notes/{nid}")
def delete_note(nid: int, actor_id: Optional[int] = None):
    actor_id = resolve_uid(actor_id)
    with db() as conn:
        old = conn.execute("SELECT * FROM project_notes WHERE id=?", (nid,)).fetchone()
        if old:
            n_allowed = {"admin", "member"} if get_settings(conn, old["project_id"]).get("member_can_edit_notes", True) else {"admin"}
            check_role(conn, old["project_id"], actor_id, n_allowed,
                       "このプロジェクトでノートを削除する権限がありません")
            conn.execute("UPDATE project_notes SET deleted_at=? WHERE id=?", (now(), nid))
    return {"ok": True}


NOTE_TEMPLATE_PLACEHOLDER = {
    "検証環境一覧・アクセス方法": "（未記入）\n\n記入例:\n■ ステージング\nURL:\n接続方法（踏み台・VPN等）:\nDB:\n\n■ 開発\nURL:\nアカウント払い出し:",
    "社内体制・担当者": "（未記入）\n\n記入例:\nPM:\nリード:\n担当（機能別）:\nエスカレーション先:",
    "先方担当者・窓口": "（未記入）\n\n記入例:\n窓口:  ○○様（部署・役職）\n決裁者:\n連絡手段・時間帯:",
    "定例・会議体": "（未記入）\n\n記入例:\n定例MTG: 毎週○曜 00:00-00:00（場所/URL）\n議事録の置き場所:\n進捗報告のルール:",
    "開発ルール（ブランチ運用・レビュー等）": "（未記入）\n\n記入例:\nブランチ運用:\nレビュー必須条件:\nコーディング規約:",
    "使用ツール・リンク集": "（未記入）\n\n記入例:\nソース管理:\nチャット:\n設計書:\nCI/CD:",
    "決定事項ログ": "（未記入）\n\n記入例:\n2026-08-23  ○○の方式はAで確定（決定者: △△）",
    "障害・緊急時の連絡フロー": "（未記入）\n\n記入例:\n一次連絡先:\nエスカレーション順:\n障害時の暫定対応手順:",
}


def apply_note_templates(conn, pid: int, actor_id=None) -> int:
    """PJ設定のテンプレートから、未作成のタイトルのノートを雛形として作成する。"""
    s = get_settings(conn, pid)
    existing = {r["title"] for r in conn.execute(
        "SELECT title FROM project_notes WHERE project_id=?", (pid,))}
    created = 0
    for i, tpl in enumerate(s.get("note_templates", [])):
        title = tpl.get("title", "").strip()
        if not title or title in existing:
            continue
        conn.execute(
            "INSERT INTO project_notes(project_id, category, title, content, pinned,"
            " sort_order, updated_by, created_at, updated_at) VALUES(?,?,?,?,?,?,?,?,?)",
            (pid, tpl.get("category", "その他"), title,
             NOTE_TEMPLATE_PLACEHOLDER.get(title, "（未記入）"),
             0, 100 + i, actor_id, now(), now()))
        created += 1
    return created


@app.post("/api/projects/{pid}/notes/apply-template")
def apply_template(pid: int, body: dict):
    with db() as conn:
        actor = resolve_uid(body.get("actor_id"))
        n_allowed = {"admin", "member"} if get_settings(conn, pid).get(
            "member_can_edit_notes", True) else {"admin"}
        check_role(conn, pid, actor, n_allowed, "ノートを編集する権限がありません")
        created = apply_note_templates(conn, pid, actor)
    return {"created": created}


# ---------------------------------------------------------------- API: discussions（イシュー一覧）

@app.get("/api/projects/{pid}/discussions")
def discussions(pid: int, user_id: Optional[int] = None):
    """コメントスレッドの一覧（イシュービュー用）。元データはタスクのコメントと同一。"""
    user_id = resolve_uid(user_id)
    with db() as conn:
        if user_id is not None:
            if effective_role(conn, pid, user_id) == "external":
                m = get_membership(conn, pid, user_id)
                if not m or not m["can_view_comments"]:
                    raise HTTPException(403, "外部ユーザーにはコメントの閲覧権限がありません")
        threads = rows_to_dicts(conn.execute(
            "SELECT t.id, t.title, t.status_id, t.assignee_id, t.priority, t.due_date,"
            "       COUNT(c.id) comment_count,"
            "       MAX(c.created_at) last_at"
            " FROM tasks t JOIN comments c ON c.task_id=t.id"
            " WHERE t.project_id=? AND t.deleted_at IS NULL"
            " GROUP BY t.id ORDER BY last_at DESC", (pid,)))
        # 各スレッドの最新コメント（投稿者・冒頭）を付与
        for th in threads:
            last = conn.execute(
                "SELECT c.body, c.created_at, m.name author_name FROM comments c"
                " LEFT JOIN members m ON m.id=c.author_id"
                " WHERE c.task_id=? ORDER BY c.id DESC LIMIT 1", (th["id"],)).fetchone()
            th["last_body"] = (last["body"][:80] if last else "")
            th["last_author"] = last["author_name"] if last else None
        recent = rows_to_dicts(conn.execute(
            "SELECT c.*, m.name author_name, m.color author_color, t.title task_title"
            " FROM comments c"
            " LEFT JOIN members m ON m.id=c.author_id"
            " JOIN tasks t ON t.id=c.task_id"
            " WHERE t.project_id=? ORDER BY c.id DESC LIMIT 25", (pid,)))
    return {"threads": threads, "recent": recent}


# ---------------------------------------------------------------- API: overview（横断ダッシュボード）

@app.get("/api/overview")
def overview(user_id: Optional[int] = None):
    """自分が関与する全プロジェクトの横断サマリー。"""
    user_id = resolve_uid(user_id)
    today = date.today().isoformat()
    with db() as conn:
        if user_id:
            proj_rows = conn.execute(
                "SELECT p.* FROM projects p"
                " JOIN project_members pm ON pm.project_id=p.id AND pm.member_id=?"
                " WHERE p.status='active' ORDER BY p.id", (user_id,)).fetchall()
        else:
            proj_rows = conn.execute(
                "SELECT * FROM projects WHERE status='active' ORDER BY id").fetchall()
        pids = [r["id"] for r in proj_rows]

        projects = []
        for pr in proj_rows:
            pid = pr["id"]
            statuses = rows_to_dicts(conn.execute(
                "SELECT * FROM statuses WHERE project_id=? ORDER BY sort_order", (pid,)))
            done_ids = {s["id"] for s in statuses if s["is_done"]}
            tasks = [task_row_to_dict(r) for r in conn.execute(
                "SELECT * FROM tasks WHERE project_id=? AND deleted_at IS NULL", (pid,))]
            total = len(tasks)
            done = sum(1 for t in tasks if t["status_id"] in done_ids)
            overdue = sum(1 for t in tasks if t["due_date"] and t["due_date"] < today
                          and t["status_id"] not in done_ids)
            my_open = sum(1 for t in tasks if user_id and t["assignee_id"] == user_id
                          and t["status_id"] not in done_ids)
            member_count = conn.execute(
                "SELECT COUNT(*) c FROM project_members WHERE project_id=?",
                (pid,)).fetchone()["c"]
            projects.append({
                "project": project_row_to_dict(pr),
                "total": total, "done": done, "overdue": overdue, "my_open": my_open,
                "member_count": member_count,
                "progress_avg": round(sum(t["progress"] for t in tasks) / total) if total else 0,
                "status_dist": [
                    {"name": s["name"], "color": s["color"],
                     "count": sum(1 for t in tasks if t["status_id"] == s["id"])}
                    for s in statuses],
            })

        my_tasks = []
        if user_id:
            rows = conn.execute(
                "SELECT t.*, p.name project_name, p.color project_color,"
                " s.name status_name, s.color status_color, s.is_done"
                " FROM tasks t JOIN projects p ON p.id=t.project_id"
                " LEFT JOIN statuses s ON s.id=t.status_id"
                " WHERE t.assignee_id=? AND t.deleted_at IS NULL AND COALESCE(s.is_done, 0)=0"
                " ORDER BY t.due_date IS NULL, t.due_date LIMIT 30", (user_id,)).fetchall()
            for r in rows:
                d = task_row_to_dict(r)
                for k in ("project_name", "project_color", "status_name",
                          "status_color", "is_done"):
                    d[k] = r[k]
                my_tasks.append(d)

        activities = []
        if pids:
            ph = ",".join("?" * len(pids))
            activities = rows_to_dicts(conn.execute(
                f"SELECT a.*, m.name actor_name, t.title task_title, p.name project_name"
                f" FROM activities a"
                f" LEFT JOIN members m ON m.id=a.actor_id"
                f" LEFT JOIN tasks t ON t.id=a.task_id"
                f" JOIN projects p ON p.id=a.project_id"
                f" WHERE a.project_id IN ({ph}) ORDER BY a.id DESC LIMIT 20", pids))
    return {"projects": projects, "my_tasks": my_tasks, "activities": activities,
            "today": today}


# ---------------------------------------------------------------- API: 管理画面（マネージャー／サイト管理者）

@app.get("/api/admin/projects")
def admin_projects():
    """管理画面のプロジェクト管理タブ用: 全PJの進捗・健全性・メンバー構成。"""
    check_site_admin("管理画面はマネージャー／サイト管理者のみ利用できます")
    today = date.today()
    ts = today.isoformat()
    stale_line = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")
    out = []
    with db() as conn:
        for pr in conn.execute("SELECT * FROM projects ORDER BY status='archived', id"):
            pid = pr["id"]
            done_ids = {s["id"] for s in conn.execute(
                "SELECT id FROM statuses WHERE project_id=? AND is_done=1", (pid,))}
            tasks = list(conn.execute(
                "SELECT status_id, progress, due_date FROM tasks"
                " WHERE project_id=? AND deleted_at IS NULL", (pid,)))
            total = len(tasks)
            done = sum(1 for t in tasks if t["status_id"] in done_ids)
            overdue = sum(1 for t in tasks if t["due_date"] and t["due_date"] < ts
                          and t["status_id"] not in done_ids)
            avg = round(sum(t["progress"] or 0 for t in tasks) / total) if total else 0
            qa_open = conn.execute(
                "SELECT COUNT(*) c FROM qa_items WHERE project_id=?"
                " AND status IN ('open','pending')", (pid,)).fetchone()["c"]
            last_act = conn.execute(
                "SELECT MAX(created_at) m FROM activities WHERE project_id=?",
                (pid,)).fetchone()["m"]
            members = rows_to_dicts(conn.execute(
                "SELECT m.id, m.name, m.color, m.account_type, m.org_role, pm.role,"
                " pm.can_view_comments, pm.can_view_detail"
                " FROM project_members pm JOIN members m ON m.id=pm.member_id"
                " WHERE pm.project_id=? AND m.active=1"
                " ORDER BY pm.role='external', m.id", (pid,)))
            p2 = project_row_to_dict(pr)
            elapsed = None
            if p2.get("start_date") and p2.get("end_date") and \
                    p2["end_date"] > p2["start_date"]:
                sd = datetime.strptime(p2["start_date"], "%Y-%m-%d").date()
                ed = datetime.strptime(p2["end_date"], "%Y-%m-%d").date()
                elapsed = max(0, min(100, round((today - sd).days /
                                                max((ed - sd).days, 1) * 100)))
            out.append({
                "project": p2, "total": total, "done": done, "overdue": overdue,
                "progress_avg": avg, "qa_open": qa_open, "last_activity": last_act,
                "elapsed_pct": elapsed,
                "stalled": bool(p2["status"] == "active" and last_act
                                and last_act < stale_line),
                "external_tabs": merge_settings(pr["settings"]).get("external_visible_tabs"),
                "members": members,
            })
    return out


@app.get("/api/admin/analytics")
def admin_analytics():
    """管理画面の横断分析タブ用: メンバー負荷・週次スループット・超過ワースト・外部アクセス。"""
    check_site_admin("管理画面はマネージャー／サイト管理者のみ利用できます")
    today = date.today()
    ts = today.isoformat()
    week_end = (today + timedelta(days=7)).isoformat()
    with db() as conn:
        done_ids = {r["id"] for r in conn.execute(
            "SELECT id FROM statuses WHERE is_done=1")}
        active_pids = {r["id"] for r in conn.execute(
            "SELECT id FROM projects WHERE status='active'")}
        pname = {r["id"]: r["name"] for r in conn.execute("SELECT id, name FROM projects")}
        parent_ids = {r["parent_id"] for r in conn.execute(
            "SELECT DISTINCT parent_id FROM tasks"
            " WHERE parent_id IS NOT NULL AND deleted_at IS NULL")}
        tasks = [dict(r) for r in conn.execute(
            "SELECT id, project_id, title, status_id, assignee_id, due_date,"
            " estimate_h, updated_at FROM tasks WHERE deleted_at IS NULL")]
        tasks = [t for t in tasks if t["project_id"] in active_pids]
        members = rows_to_dicts(conn.execute("SELECT * FROM members WHERE active=1"))
        # --- メンバー負荷（親タスクは除外して実作業のみ）
        workload = []
        d30 = (today - timedelta(days=30)).isoformat()
        for m in members:
            mine = [t for t in tasks
                    if t["assignee_id"] == m["id"] and t["id"] not in parent_ids]
            open_t = [t for t in mine if t["status_id"] not in done_ids]
            workload.append({
                "id": m["id"], "name": m["name"], "color": m["color"],
                "account_type": m["account_type"], "org_role": m["org_role"],
                "open": len(open_t),
                "overdue": sum(1 for t in open_t
                               if t["due_date"] and t["due_date"] < ts),
                "due_week": sum(1 for t in open_t
                                if t["due_date"] and ts <= t["due_date"] <= week_end),
                "est_open": round(sum(t["estimate_h"] or 0 for t in open_t), 1),
                "projects": len({t["project_id"] for t in open_t}),
                "done_30d": sum(1 for t in mine if t["status_id"] in done_ids
                                and (t["updated_at"] or "") >= d30),
            })
        workload.sort(key=lambda w: (-w["overdue"], -w["open"]))
        # --- 週次スループット（完了ステータスのタスクの最終更新週・直近8週）
        weekly = []
        monday = today - timedelta(days=today.weekday())
        for i in range(7, -1, -1):
            ws_ = monday - timedelta(weeks=i)
            we_ = (ws_ + timedelta(days=6)).isoformat()
            weekly.append({
                "label": f"{ws_.month}/{ws_.day}",
                "done": sum(1 for t in tasks if t["status_id"] in done_ids
                            and ws_.isoformat() <= (t["updated_at"] or "")[:10] <= we_)})
        # --- 期限超過ワースト10
        overdue_tasks = sorted(
            [{"id": t["id"], "title": t["title"], "project_id": t["project_id"],
              "project_name": pname.get(t["project_id"], ""),
              "due_date": t["due_date"],
              "assignee": next((m["name"] for m in members
                                if m["id"] == t["assignee_id"]), None),
              "days": (today - datetime.strptime(t["due_date"], "%Y-%m-%d").date()).days}
             for t in tasks if t["due_date"] and t["due_date"] < ts
             and t["status_id"] not in done_ids and t["id"] not in parent_ids],
            key=lambda x: -x["days"])[:10]
        # --- 外部ユーザーのアクセス状況（セキュリティレビュー用）
        externals = []
        for m in [x for x in members if x["account_type"] == "external"]:
            pms = rows_to_dicts(conn.execute(
                "SELECT pm.*, p.name project_name, p.settings FROM project_members pm"
                " JOIN projects p ON p.id=pm.project_id WHERE pm.member_id=?",
                (m["id"],)))
            last_login = conn.execute(
                "SELECT MAX(created_at) m FROM login_logs"
                " WHERE member_id=? AND success=1", (m["id"],)).fetchone()["m"]
            externals.append({
                "id": m["id"], "name": m["name"], "email": m["email"],
                "last_login": last_login,
                "projects": [{
                    "project_id": x["project_id"], "name": x["project_name"],
                    "can_view_comments": x["can_view_comments"],
                    "can_view_detail": x["can_view_detail"],
                    "tabs": merge_settings(x["settings"]).get("external_visible_tabs"),
                } for x in pms],
            })
    return {"workload": workload, "weekly_done": weekly,
            "overdue_tasks": overdue_tasks, "externals": externals}


# ---------------------------------------------------------------- API: members

def _check_email(conn, email: str, exclude_mid: Optional[int] = None) -> str:
    """メール必須・重複チェック（小文字化して返す）。"""
    email = (email or "").strip().lower()
    if not email or "@" not in email:
        raise HTTPException(400, "メールアドレスを入力してください")
    q = "SELECT id FROM members WHERE lower(email)=?"
    args: list = [email]
    if exclude_mid is not None:
        q += " AND id != ?"
        args.append(exclude_mid)
    if conn.execute(q, args).fetchone():
        raise HTTPException(400, "このメールアドレスは既に使われています")
    return email


@app.post("/api/members")
def create_member(m: MemberIn):
    check_site_admin()
    with db() as conn:
        email = _check_email(conn, m.email)
        cur = conn.execute(
            "INSERT INTO members(name, role, color, org_id, account_type, org_role, email)"
            " VALUES(?,?,?,?,?,?,?)",
            (m.name, m.role, m.color, m.org_id, m.account_type, m.org_role, email))
        r = conn.execute("SELECT * FROM members WHERE id=?", (cur.lastrowid,)).fetchone()
    d = dict(r); d.pop("password_hash", None)
    return d


@app.patch("/api/members/{mid}")
def update_member(mid: int, m: dict):
    check_site_admin()
    allowed = {"name", "role", "color", "active", "org_id", "account_type", "org_role", "email"}
    sets = [f"{k}=?" for k in m if k in allowed]
    vals = [m[k] for k in m if k in allowed] + [mid]
    if not sets:
        raise HTTPException(400, "no valid fields")
    with db() as conn:
        if "email" in m:
            m["email"] = _check_email(conn, m["email"], exclude_mid=mid)
            vals = [m[k] for k in m if k in allowed] + [mid]
        conn.execute(f"UPDATE members SET {', '.join(sets)} WHERE id=?", vals)
        r = conn.execute("SELECT * FROM members WHERE id=?", (mid,)).fetchone()
    d = dict(r); d.pop("password_hash", None)
    return d


@app.delete("/api/members/{mid}")
def delete_member(mid: int):
    check_site_admin()
    with db() as conn:
        conn.execute("UPDATE members SET active=0 WHERE id=?", (mid,))
    return {"ok": True}


# ---------------------------------------------------------------- API: statuses

@app.post("/api/projects/{pid}/statuses")
def create_status(pid: int, s: StatusIn):
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO statuses(project_id, name, color, sort_order, is_done)"
            " VALUES(?,?,?,?,?)",
            (pid, s.name, s.color, s.sort_order, int(s.is_done)))
        r = conn.execute("SELECT * FROM statuses WHERE id=?", (cur.lastrowid,)).fetchone()
    return dict(r)


@app.patch("/api/statuses/{sid}")
def update_status(sid: int, s: dict):
    allowed = {"name", "color", "sort_order", "is_done"}
    sets = [f"{k}=?" for k in s if k in allowed]
    vals = [s[k] for k in s if k in allowed] + [sid]
    if not sets:
        raise HTTPException(400, "no valid fields")
    with db() as conn:
        conn.execute(f"UPDATE statuses SET {', '.join(sets)} WHERE id=?", vals)
        r = conn.execute("SELECT * FROM statuses WHERE id=?", (sid,)).fetchone()
    return dict(r)


@app.delete("/api/statuses/{sid}")
def delete_status(sid: int):
    with db() as conn:
        used = conn.execute(
            "SELECT COUNT(*) c FROM tasks WHERE status_id=?", (sid,)).fetchone()["c"]
        if used:
            raise HTTPException(400, f"このステータスには {used} 件のタスクがあります")
        conn.execute("DELETE FROM statuses WHERE id=?", (sid,))
    return {"ok": True}


# ---------------------------------------------------------------- API: tasks

def _status_maps(conn: sqlite3.Connection, pid: int) -> Optional[dict]:
    """ステータス⇔進捗連動用の役割マップ。
    first=先頭（未着手扱い）/ done=完了群 / mid=進行中（名前に「進行」を含む→2番目の未完了）"""
    sts = conn.execute(
        "SELECT id, name, is_done FROM statuses WHERE project_id=?"
        " ORDER BY sort_order, id", (pid,)).fetchall()
    if not sts:
        return None
    first = sts[0]["id"]
    done = [s["id"] for s in sts if s["is_done"]]
    mid = next((s["id"] for s in sts if not s["is_done"] and "進行" in s["name"]), None)
    if mid is None:
        nd = [s["id"] for s in sts if not s["is_done"] and s["id"] != first]
        mid = nd[0] if nd else first
    return {"first": first, "done": done, "mid": mid}


def status_for_progress(maps, progress, current_status_id):
    """進捗の帯（0=未着手 / 1-99=進行中 / 100=完了）に合うステータスを返す。
    1-99 で既に中間ステータス（レビュー中等）にある場合はそれを尊重する。"""
    if maps is None:
        return current_status_id
    if progress >= 100:
        return (current_status_id if current_status_id in maps["done"]
                else (maps["done"][0] if maps["done"] else current_status_id))
    if progress <= 0:
        return maps["first"]
    if (current_status_id and current_status_id != maps["first"]
            and current_status_id not in maps["done"]):
        return current_status_id
    return maps["mid"]


def progress_for_status(maps, status_id, current_progress):
    """ステータス変更に合わせた進捗。完了=100 / 先頭=0 / 中間=（0か100なら）50。"""
    if maps is None or status_id is None:
        return current_progress
    if status_id in maps["done"]:
        return 100
    if status_id == maps["first"]:
        return 0
    return current_progress if 0 < current_progress < 100 else 50


def recalc_parent_progress(conn: sqlite3.Connection, pid: int):
    """サブタスクを持つタスクの進捗を、子タスク進捗の平均から自動算出して保存する。
    多階層はボトムアップで再帰計算（葉の値のみが入力）。updated_at は変えない
    （自動再計算で楽観ロックを壊さないため）。"""
    rows = conn.execute(
        "SELECT id, parent_id, progress, status_id FROM tasks"
        " WHERE project_id=? AND deleted_at IS NULL", (pid,)).fetchall()
    prog = {r["id"]: r["progress"] or 0 for r in rows}
    status = {r["id"]: r["status_id"] for r in rows}
    children: dict = {}
    for r in rows:
        children.setdefault(r["parent_id"], []).append(r["id"])
    memo: dict = {}

    def calc(tid):
        if tid in memo:
            return memo[tid]
        kids = [k for k in children.get(tid, []) if k in prog]
        v = prog[tid] if not kids else round(sum(calc(k) for k in kids) / len(kids))
        memo[tid] = v
        return v

    for tid in prog:
        calc(tid)
    maps = _status_maps(conn, pid)
    for tid, v in memo.items():
        if not children.get(tid):
            continue
        if v != prog[tid]:
            conn.execute("UPDATE tasks SET progress=? WHERE id=?", (v, tid))
        # 親のステータスも進捗の帯に合わせる
        ns = status_for_progress(maps, v, status[tid])
        if ns != status[tid]:
            conn.execute("UPDATE tasks SET status_id=? WHERE id=?", (ns, tid))


@app.post("/api/projects/{pid}/tasks")
def create_task(pid: int, t: TaskIn):
    t.actor_id = resolve_uid(t.actor_id)
    with db() as conn:
        allowed = {"admin", "member"}
        if not get_settings(conn, pid).get("member_can_create_tasks", True):
            allowed = {"admin"}
        check_role(conn, pid, t.actor_id, allowed,
                   "このプロジェクトでタスクを作成する権限がありません")
        max_order = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) m FROM tasks WHERE project_id=?",
            (pid,)).fetchone()["m"]
        status_id = t.status_id
        if status_id is None:
            first = conn.execute(
                "SELECT id FROM statuses WHERE project_id=? ORDER BY sort_order LIMIT 1",
                (pid,)).fetchone()
            status_id = first["id"] if first else None
            if t.progress:   # 進捗指定つき作成はステータスも帯に合わせる
                status_id = status_for_progress(_status_maps(conn, pid),
                                                t.progress, status_id)
        cur = conn.execute(
            "INSERT INTO tasks(project_id, parent_id, title, description, status_id,"
            " assignee_id, assignee_label, priority, start_date, due_date, progress,"
            " estimate_h, actual_h,"
            " milestone, tags, deps, custom_values, recur, sort_order, created_at, updated_at)"
            " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (pid, t.parent_id, t.title, t.description, status_id, t.assignee_id,
             t.assignee_label,
             t.priority, t.start_date, t.due_date, t.progress, t.estimate_h, t.actual_h,
             int(t.milestone), json.dumps(t.tags, ensure_ascii=False),
             json.dumps(t.deps), json.dumps(t.custom_values, ensure_ascii=False),
             t.recur, max_order + 1, now(), now()))
        record_activity(conn, pid, cur.lastrowid, t.actor_id, "create", t.title)
        if t.assignee_id and t.assignee_id != t.actor_id:
            notify(conn, t.assignee_id, "assign", pid, cur.lastrowid, t.actor_id,
                   f"「{t.title}」の担当になりました")
        recalc_parent_progress(conn, pid)
        r = conn.execute("SELECT * FROM tasks WHERE id=?", (cur.lastrowid,)).fetchone()
    return task_row_to_dict(r)


@app.patch("/api/tasks/{tid}")
def update_task(tid: int, patch: TaskPatch):
    data = patch.model_dump(exclude_unset=True)
    actor_id = resolve_uid(data.pop("actor_id", None))
    expected = data.pop("expected_updated_at", None)
    if not data:
        raise HTTPException(400, "no fields")
    with db() as conn:
        old = get_task_or_404(conn, tid)
        if expected is not None and old["updated_at"] != expected:
            raise HTTPException(409, "他のユーザーが先に更新しています。再読込してください")
        # 権限チェック（actor_id が渡された場合のみ）
        if actor_id is not None:
            role = effective_role(conn, old["project_id"], actor_id)
            if role == "admin":
                pass
            elif role == "member":
                s = get_settings(conn, old["project_id"])
                editable = set(MEMBER_EDITABLE_FIELDS)
                if s.get("member_can_edit_own_schedule"):
                    editable |= MEMBER_SCHEDULE_FIELDS
                fields = set(data.keys())
                is_self_assign = (fields == {"assignee_id"}
                                  and old["assignee_id"] is None
                                  and data["assignee_id"] == actor_id)
                if not is_self_assign:
                    if old["assignee_id"] != actor_id:
                        raise HTTPException(
                            403, "メンバーは自分の担当タスクのみ変更できます")
                    over = fields - editable
                    if over:
                        raise HTTPException(
                            403, "スケジュール・期限など全体に影響する変更は"
                                 "リーダーまたは組織の上位者のみ行えます")
            else:
                raise HTTPException(403, "このプロジェクトの編集権限がありません")
        # サブタスクを持つタスクの進捗は自動算出のため、直接指定は無視する
        has_kids = conn.execute(
            "SELECT 1 FROM tasks WHERE parent_id=? AND deleted_at IS NULL LIMIT 1",
            (tid,)).fetchone() is not None
        if "progress" in data and has_kids:
            data.pop("progress")
            if not data:
                return old   # get_task_or_404 は dict を返す
        # ステータス⇔進捗の連動（片方だけ指定されたとき他方を自動導出）
        maps = _status_maps(conn, old["project_id"])
        if "progress" in data and "status_id" not in data:
            ns = status_for_progress(maps, data["progress"], old["status_id"])
            if ns != old["status_id"]:
                data["status_id"] = ns
        elif "status_id" in data and "progress" not in data and not has_kids:
            np = progress_for_status(maps, data["status_id"], old["progress"])
            if np != old["progress"]:
                data["progress"] = np
        sets, vals = [], []
        for k, v in data.items():
            if k in ("tags", "deps", "custom_values"):
                v = json.dumps(v, ensure_ascii=False)
            elif k == "milestone":
                v = int(v)
            sets.append(f"{k}=?")
            vals.append(v)
        sets.append("updated_at=?")
        vals.append(now())
        vals.append(tid)
        conn.execute(f"UPDATE tasks SET {', '.join(sets)} WHERE id=?", vals)

        # 主要な変更をアクティビティに記録
        if "status_id" in data and data["status_id"] != old["status_id"]:
            names = {r["id"]: r["name"] for r in conn.execute(
                "SELECT id, name FROM statuses WHERE project_id=?", (old["project_id"],))}
            record_activity(conn, old["project_id"], tid, actor_id, "status",
                            f"{names.get(old['status_id'], '—')} → {names.get(data['status_id'], '—')}")
        if "assignee_id" in data and data["assignee_id"] != old["assignee_id"]:
            names = {r["id"]: r["name"] for r in conn.execute("SELECT id, name FROM members")}
            old_disp = names.get(old["assignee_id"]) or old.get("assignee_label") or "未割当"
            new_disp = names.get(data["assignee_id"]) or data.get("assignee_label") or "未割当"
            record_activity(conn, old["project_id"], tid, actor_id, "assignee",
                            f"{old_disp} → {new_disp}")
        elif "assignee_label" in data and data["assignee_label"] != old.get("assignee_label"):
            record_activity(conn, old["project_id"], tid, actor_id, "assignee",
                            f"{old.get('assignee_label') or '未割当'} → {data['assignee_label'] or '未割当'}")
        if "progress" in data and data["progress"] != old["progress"]:
            record_activity(conn, old["project_id"], tid, actor_id, "progress",
                            f"{old['progress']}% → {data['progress']}%")

        # ---- 通知・Webhook・繰り返しタスク
        pid = old["project_id"]
        actor_name = ""
        if actor_id:
            ar = conn.execute("SELECT name FROM members WHERE id=?", (actor_id,)).fetchone()
            actor_name = ar["name"] if ar else ""
        if "assignee_id" in data and data["assignee_id"] != old["assignee_id"]:
            notify(conn, data["assignee_id"], "assign", pid, tid, actor_id,
                   f"「{old['title']}」の担当になりました")
            send_webhook(conn, pid, "assign",
                         f"📌 {actor_name or 'システム'} が「{old['title']}」の担当を変更しました")
        if "status_id" in data and data["status_id"] != old["status_id"]:
            st = conn.execute("SELECT * FROM statuses WHERE id=?",
                              (data["status_id"],)).fetchone()
            st_name = st["name"] if st else "?"
            targets = watchers_of(conn, tid) | {old["assignee_id"]}
            for uid2 in targets:
                notify(conn, uid2, "status", pid, tid, actor_id,
                       f"「{old['title']}」→ {st_name}")
            send_webhook(conn, pid, "status",
                         f"🔄 「{old['title']}」が {st_name} になりました（{actor_name}）")
            # 繰り返しタスク: 完了列に移動したら次回分を自動作成
            if st and st["is_done"] and (old.get("recur") or "").strip():
                days = {"weekly": 7, "biweekly": 14, "monthly": 30}.get(old["recur"], 0)
                if days:
                    def shift(dstr):
                        if not dstr:
                            return None
                        from datetime import timedelta
                        d0 = datetime.strptime(dstr, "%Y-%m-%d")
                        return (d0 + timedelta(days=days)).strftime("%Y-%m-%d")
                    first = conn.execute(
                        "SELECT id FROM statuses WHERE project_id=?"
                        " ORDER BY sort_order LIMIT 1", (pid,)).fetchone()
                    cur2 = conn.execute(
                        "INSERT INTO tasks(project_id, parent_id, title, description,"
                        " status_id, assignee_id, priority, start_date, due_date, progress,"
                        " estimate_h, milestone, tags, deps, custom_values, recur,"
                        " sort_order, created_at, updated_at)"
                        " VALUES(?,?,?,?,?,?,?,?,?,0,?,?,?,?,?,?,?,?,?)",
                        (pid, old["parent_id"], old["title"], old["description"],
                         first["id"] if first else None, old["assignee_id"], old["priority"],
                         shift(old["start_date"]), shift(old["due_date"]),
                         old["estimate_h"], int(old["milestone"]),
                         json.dumps(old["tags"], ensure_ascii=False),
                         json.dumps(old["deps"]),
                         json.dumps(old["custom_values"], ensure_ascii=False),
                         old["recur"], old["sort_order"] + 1, now(), now()))
                    record_activity(conn, pid, cur2.lastrowid, None, "create",
                                    f"{old['title']}（繰り返し）")
                    notify(conn, old["assignee_id"], "system", pid, cur2.lastrowid, None,
                           f"繰り返しタスク「{old['title']}」の次回分を作成しました")
        recalc_parent_progress(conn, pid)
        r = conn.execute("SELECT * FROM tasks WHERE id=?", (tid,)).fetchone()
    return task_row_to_dict(r)


@app.delete("/api/tasks/{tid}")
def delete_task(tid: int, actor_id: Optional[int] = None):
    actor_id = resolve_uid(actor_id)
    with db() as conn:
        old = get_task_or_404(conn, tid)
        if actor_id is not None:
            role = effective_role(conn, old["project_id"], actor_id)
            # 緩め: admin は全タスク、member は自分の担当タスクのみ削除可
            if not (role == "admin" or
                    (role == "member" and old["assignee_id"] == actor_id)):
                raise HTTPException(403, "このタスクを削除する権限がありません")
        record_activity(conn, old["project_id"], None, actor_id, "delete", old["title"])
        conn.execute("UPDATE tasks SET deleted_at=? WHERE id=?", (now(), tid))
        recalc_parent_progress(conn, old["project_id"])
    return {"ok": True}


@app.post("/api/tasks/reorder")
def reorder_tasks(payload: dict):
    """[{id, sort_order, parent_id?}] を一括反映（ボード/WBSの並べ替え用）。"""
    items = payload.get("items", [])
    with db() as conn:
        for it in items:
            if "parent_id" in it:
                conn.execute("UPDATE tasks SET sort_order=?, parent_id=? WHERE id=?",
                             (it["sort_order"], it["parent_id"], it["id"]))
            else:
                conn.execute("UPDATE tasks SET sort_order=? WHERE id=?",
                             (it["sort_order"], it["id"]))
        if items:
            r = conn.execute("SELECT project_id FROM tasks WHERE id=?",
                             (items[0]["id"],)).fetchone()
            if r:
                recalc_parent_progress(conn, r["project_id"])
    return {"ok": True}


@app.get("/api/tasks/{tid}/detail")
def task_detail(tid: int, user_id: Optional[int] = None):
    user_id = resolve_uid(user_id)
    with db() as conn:
        t = get_task_or_404(conn, tid)
    external_guard(t["project_id"])   # 未アサインの外部を遮断（セッション基準）
    with db() as conn:
        hide_comments = False
        if user_id is not None:
            if effective_role(conn, t["project_id"], user_id) == "external":
                m = get_membership(conn, t["project_id"], user_id)
                if not m or not m["can_view_detail"]:
                    raise HTTPException(403, "外部ユーザーにはタスク詳細の閲覧権限がありません")
                hide_comments = not m["can_view_comments"]
        comments = rows_to_dicts(conn.execute(
            "SELECT c.*, m.name author_name, m.color author_color FROM comments c"
            " LEFT JOIN members m ON m.id=c.author_id"
            " WHERE c.task_id=? ORDER BY c.id", (tid,)))
        links = rows_to_dicts(conn.execute(
            "SELECT * FROM task_links WHERE task_id=? ORDER BY id", (tid,)))
        subtasks = [task_row_to_dict(r) for r in conn.execute(
            "SELECT * FROM tasks WHERE parent_id=? ORDER BY sort_order, id", (tid,))]
        activities = rows_to_dicts(conn.execute(
            "SELECT a.*, m.name actor_name FROM activities a"
            " LEFT JOIN members m ON m.id=a.actor_id"
            " WHERE a.task_id=? ORDER BY a.id DESC LIMIT 20", (tid,)))
    if hide_comments:
        comments = []
        activities = [a for a in activities if a["action"] != "comment"]
    with db() as conn:
        watching = bool(user_id and conn.execute(
            "SELECT 1 FROM watchers WHERE task_id=? AND member_id=?",
            (tid, user_id)).fetchone())
        attachments = attachments_for(conn, "task", tid)
        relations = rows_to_dicts(conn.execute(
            "SELECT r.*, t2.title other_title FROM task_relations r"
            " JOIN tasks t2 ON t2.id=r.other_id"
            " WHERE r.task_id=? AND t2.deleted_at IS NULL ORDER BY r.id", (tid,)))
        reactions = reactions_for_comments(
            conn, [c["id"] for c in comments], user_id)
        rel_issues = rows_to_dicts(conn.execute(
            "SELECT i.id, i.seq, i.title, i.status FROM issue_tasks it"
            " JOIN issues i ON i.id=it.issue_id WHERE it.task_id=?"
            " ORDER BY i.status='closed', i.seq", (tid,)))
    for c in comments:
        c["reactions"] = reactions.get(c["id"], [])
    return {"task": t, "comments": comments, "links": links,
            "subtasks": subtasks, "activities": activities,
            "comments_hidden": hide_comments, "watching": watching,
            "attachments": attachments, "relations": relations,
            "issues": rel_issues}


# ---------------------------------------------------------------- API: comments / links

@app.post("/api/tasks/{tid}/comments")
def add_comment(tid: int, c: CommentIn):
    c.author_id = resolve_uid(c.author_id)
    with db() as conn:
        t = get_task_or_404(conn, tid)
        if c.author_id is not None and not can_comment_in(conn, t["project_id"], c.author_id):
            raise HTTPException(403, "このプロジェクトでコメントする権限がありません")
        cur = conn.execute(
            "INSERT INTO comments(task_id, author_id, body, created_at) VALUES(?,?,?,?)",
            (tid, c.author_id, c.body, now()))
        record_activity(conn, t["project_id"], tid, c.author_id, "comment",
                        c.body[:60])
        # 通知: @メンション > 担当者・ウォッチャー
        author_name = ""
        if c.author_id:
            ar = conn.execute("SELECT name FROM members WHERE id=?",
                              (c.author_id,)).fetchone()
            author_name = ar["name"] if ar else ""
        mentioned = set(mentioned_user_ids(conn, c.body))
        for uid2 in mentioned:
            notify(conn, uid2, "mention", t["project_id"], tid, c.author_id,
                   f"{author_name} があなたをメンション: {c.body[:60]}")
        others = ({t["assignee_id"]} | watchers_of(conn, tid)) - mentioned
        for uid2 in others:
            notify(conn, uid2, "comment", t["project_id"], tid, c.author_id,
                   f"{author_name} が「{t['title']}」にコメント: {c.body[:60]}")
        if mentioned:
            send_webhook(conn, t["project_id"], "mention",
                         f"💬 {author_name} がメンション（{t['title']}）: {c.body[:120]}")
        else:
            send_webhook(conn, t["project_id"], "comment",
                         f"💬 {author_name} がコメント（{t['title']}）: {c.body[:120]}")
        r = conn.execute(
            "SELECT c.*, m.name author_name, m.color author_color FROM comments c"
            " LEFT JOIN members m ON m.id=c.author_id WHERE c.id=?",
            (cur.lastrowid,)).fetchone()
    return dict(r)


@app.delete("/api/comments/{cid}")
def delete_comment(cid: int, actor_id: Optional[int] = None):
    actor_id = resolve_uid(actor_id)
    with db() as conn:
        c = conn.execute("SELECT * FROM comments WHERE id=?", (cid,)).fetchone()
        if not c:
            return {"ok": True}
        if actor_id is not None and c["author_id"] != actor_id:
            t = conn.execute("SELECT project_id FROM tasks WHERE id=?",
                             (c["task_id"],)).fetchone()
            if t and effective_role(conn, t["project_id"], actor_id) != "admin":
                raise HTTPException(403, "他人のコメントを削除できるのは管理者のみです")
        conn.execute("DELETE FROM comments WHERE id=?", (cid,))
    return {"ok": True}


@app.post("/api/tasks/{tid}/links")
def add_link(tid: int, l: LinkIn):
    with db() as conn:
        get_task_or_404(conn, tid)
        cur = conn.execute(
            "INSERT INTO task_links(task_id, title, url, kind) VALUES(?,?,?,?)",
            (tid, l.title, l.url, l.kind))
        r = conn.execute("SELECT * FROM task_links WHERE id=?", (cur.lastrowid,)).fetchone()
    return dict(r)


@app.delete("/api/links/{lid}")
def delete_link(lid: int):
    with db() as conn:
        conn.execute("DELETE FROM task_links WHERE id=?", (lid,))
    return {"ok": True}


# ---------------------------------------------------------------- export helpers

def check_export_allowed(pid: int):
    """外部ユーザーのエクスポートはPJ設定で許可されている場合のみ。"""
    external_guard(pid)   # 未アサインの外部を遮断
    su = CURRENT_USER.get()
    if su is None:
        return
    with db() as conn:
        if effective_role(conn, pid, su["id"]) == "external" and \
                not get_settings(conn, pid).get("external_can_export"):
            raise HTTPException(403, "外部ユーザーのエクスポートはこのプロジェクトでは許可されていません")


EXCEL_FONT = "游ゴシック"   # 既定のMS Pゴシックより読みやすいフォントに統一


def XFont(**kw):
    """Excel出力用フォント（名前を游ゴシックに統一）。"""
    from openpyxl.styles import Font
    kw.setdefault("name", EXCEL_FONT)
    return Font(**kw)


def _wb_default_font(wb):
    """ブック既定フォント（index 0）を差し替え、フォント無指定のセルにも効かせる。
    ※Normalスタイルへの代入では既定セルに反映されない（openpyxl 3.1）。"""
    try:
        wb._fonts[0] = XFont(size=11)
    except Exception:
        pass


def _ignore_number_as_text(ws, sqref: str):
    """「数値が文字列として保存されています」の警告をセル範囲単位で抑止する。
    openpyxl は ignoredErrors を書き出せないため、範囲を控えて保存時にXML注入する。"""
    if not hasattr(ws, "_pjb_ignore"):
        ws._pjb_ignore = []
    ws._pjb_ignore.append(sqref)


def _save_wb_with_ignores(wb) -> io.BytesIO:
    """wb.save 後、控えた ignoredErrors をシートXMLへ注入して返す。"""
    bio = io.BytesIO()
    wb.save(bio)
    ignores = {i: getattr(ws, "_pjb_ignore", None)
               for i, ws in enumerate(wb.worksheets)}
    if not any(ignores.values()):
        bio.seek(0)
        return bio
    src = zipfile.ZipFile(bio)
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        for item in src.infolist():
            data = src.read(item.filename)
            m = re.match(r"xl/worksheets/sheet(\d+)\.xml$", item.filename)
            if m and ignores.get(int(m.group(1)) - 1):
                frag = "<ignoredErrors>" + "".join(
                    f'<ignoredError sqref="{s}" numberStoredAsText="1"/>'
                    for s in ignores[int(m.group(1)) - 1]) + "</ignoredErrors>"
                data = data.replace(b"</worksheet>", frag.encode() + b"</worksheet>")
            z.writestr(item, data)
    out.seek(0)
    return out


def collect_export_data(pid: int) -> dict:
    check_export_allowed(pid)
    with db() as conn:
        pr = conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
        if not pr:
            raise HTTPException(404, "project not found")
        project = project_row_to_dict(pr)
        statuses = rows_to_dicts(conn.execute(
            "SELECT * FROM statuses WHERE project_id=? ORDER BY sort_order", (pid,)))
        tasks = [task_row_to_dict(r) for r in conn.execute(
            "SELECT * FROM tasks WHERE project_id=? AND deleted_at IS NULL"
            " ORDER BY sort_order, id", (pid,))]
        members = rows_to_dicts(conn.execute("SELECT * FROM members"))
        comments = rows_to_dicts(conn.execute(
            "SELECT c.*, m.name author_name FROM comments c"
            " LEFT JOIN members m ON m.id=c.author_id"
            " WHERE c.task_id IN (SELECT id FROM tasks WHERE project_id=?)"
            " ORDER BY c.task_id, c.id", (pid,)))
        links = rows_to_dicts(conn.execute(
            "SELECT * FROM task_links WHERE task_id IN"
            " (SELECT id FROM tasks WHERE project_id=?) ORDER BY task_id, id", (pid,)))
        notes = rows_to_dicts(conn.execute(
            "SELECT n.*, m.name author_name FROM project_notes n"
            " LEFT JOIN members m ON m.id=n.updated_by"
            " WHERE n.project_id=? AND n.deleted_at IS NULL"
            " ORDER BY n.pinned DESC, n.sort_order, n.id", (pid,)))
        activities = rows_to_dicts(conn.execute(
            "SELECT a.*, m.name actor_name, t.title task_title FROM activities a"
            " LEFT JOIN members m ON m.id=a.actor_id"
            " LEFT JOIN tasks t ON t.id=a.task_id"
            " WHERE a.project_id=? ORDER BY a.id", (pid,)))
        qa = rows_to_dicts(conn.execute(
            "SELECT q.*, m.name assignee_name, t.title task_title FROM qa_items q"
            " LEFT JOIN members m ON m.id=q.assignee_id"
            " LEFT JOIN tasks t ON t.id=q.task_id"
            " WHERE q.project_id=? ORDER BY q.seq, q.id", (pid,)))
        qac = rows_to_dicts(conn.execute(
            "SELECT c.*, m.name member_name FROM qa_comments c"
            " LEFT JOIN members m ON m.id=c.author_id"
            " WHERE c.qa_id IN (SELECT id FROM qa_items WHERE project_id=?)"
            " ORDER BY c.qa_id, c.id", (pid,)))
        qac_by: dict = {}
        for c in qac:
            qac_by.setdefault(c["qa_id"], []).append(c)
        for q in qa:
            q["comments"] = qac_by.get(q["id"], [])
        issues = rows_to_dicts(conn.execute(
            "SELECT i.*, m.name assignee_name FROM issues i"
            " LEFT JOIN members m ON m.id=i.assignee_id"
            " WHERE i.project_id=? ORDER BY i.seq, i.id", (pid,)))
        itmap = _issue_tasks_map(conn, pid)
        ic = rows_to_dicts(conn.execute(
            "SELECT c.*, m.name member_name FROM issue_comments c"
            " LEFT JOIN members m ON m.id=c.author_id"
            " WHERE c.issue_id IN (SELECT id FROM issues WHERE project_id=?)"
            " ORDER BY c.issue_id, c.id", (pid,)))
        ic_by: dict = {}
        for c in ic:
            ic_by.setdefault(c["issue_id"], []).append(c)
        for i in issues:
            i["tasks"] = itmap.get(i["id"], [])
            i["comments"] = ic_by.get(i["id"], [])
    return {"project": project, "statuses": statuses, "tasks": tasks,
            "members": members, "comments": comments, "links": links,
            "notes": notes, "activities": activities, "qa": qa, "issues": issues}


def build_wbs_rows(tasks: list[dict]) -> list[dict]:
    """親子関係をたどり WBS 番号（1, 1.1, 1.2, 2 …）と深さを付与した順序付きリスト。"""
    children: dict[Any, list] = {}
    for t in tasks:
        children.setdefault(t["parent_id"], []).append(t)
    for v in children.values():
        v.sort(key=lambda x: (x["sort_order"], x["id"]))
    out = []

    def walk(parent_id, prefix, depth):
        for i, t in enumerate(children.get(parent_id, []), 1):
            num = f"{prefix}{i}"
            out.append({**t, "wbs": num, "depth": depth})
            walk(t["id"], num + ".", depth + 1)
    walk(None, "", 0)
    # 親が同一プロジェクト外/削除済みで孤児になったタスクも末尾に含める
    seen = {t["id"] for t in out}
    for t in tasks:
        if t["id"] not in seen:
            out.append({**t, "wbs": "-", "depth": 0})
    return out


PRIORITY_LABEL = {"highest": "最優先", "high": "高", "medium": "中", "low": "低"}


# ---------------------------------------------------------------- export: JSON

@app.get("/api/projects/{pid}/export.json")
def export_json(pid: int):
    """機械可読の全データダンプ（バックアップ・外部ツール連携用）。"""
    d = collect_export_data(pid)
    d["tasks"] = build_wbs_rows(d["tasks"])   # WBS番号・階層深さ付きで出力
    d["exported_at"] = now()
    body = json.dumps(d, ensure_ascii=False, indent=2)
    fname = f"project_{pid}_{date.today().isoformat()}.json"
    return Response(body, media_type="application/json; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------------------------------------------------------------- export: CSV

@app.get("/api/projects/{pid}/export.csv")
def export_csv(pid: int):
    d = collect_export_data(pid)
    smap = {s["id"]: s["name"] for s in d["statuses"]}
    mmap = {m["id"]: m["name"] for m in d["members"]}
    cf_defs = d["project"]["custom_fields"]

    buf = io.StringIO()
    w = csv.writer(buf)
    headers = ["WBS", "ID", "タスク名", "ステータス", "担当者", "優先度", "開始日", "期限",
               "進捗%", "見積h", "実績h", "マイルストーン", "タグ", "先行タスク", "説明"]
    headers += [f["label"] for f in cf_defs]
    w.writerow(headers)
    for t in build_wbs_rows(d["tasks"]):
        row = [t["wbs"], t["id"], ("　" * t["depth"]) + t["title"],
               smap.get(t["status_id"], ""), mmap.get(t["assignee_id"]) or t.get("assignee_label") or "",
               PRIORITY_LABEL.get(t["priority"], t["priority"]),
               t["start_date"] or "", t["due_date"] or "", t["progress"],
               t["estimate_h"] or "", t["actual_h"] or "",
               "○" if t["milestone"] else "", " / ".join(t["tags"]),
               " ".join(str(x) for x in t["deps"]), t["description"]]
        row += [t["custom_values"].get(f["key"], "") for f in cf_defs]
        w.writerow(row)
    data = "﻿" + buf.getvalue()   # BOM付きUTF-8（Excelでの文字化け防止）
    fname = f"tasks_{pid}_{date.today().isoformat()}.csv"
    return Response(data, media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------------------------------------------------------------- export: Excel

@app.get("/api/projects/{pid}/export.xlsx")
def export_xlsx(pid: int):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    d = collect_export_data(pid)
    smap = {s["id"]: s for s in d["statuses"]}
    mmap = {m["id"]: m["name"] for m in d["members"]}
    cf_defs = d["project"]["custom_fields"]
    wbs_rows = build_wbs_rows(d["tasks"])

    wb = Workbook()
    _wb_default_font(wb)
    head_fill = PatternFill("solid", fgColor="2F4870")
    head_font = XFont(color="FFFFFF", bold=True)

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    # --- サマリーシート
    ws = wb.active
    ws.title = "サマリー"
    total = len(d["tasks"])
    done = sum(1 for t in d["tasks"] if smap.get(t["status_id"], {}).get("is_done"))
    today = date.today().isoformat()
    overdue = sum(1 for t in d["tasks"]
                  if t["due_date"] and t["due_date"] < today
                  and not smap.get(t["status_id"], {}).get("is_done"))
    avg = round(sum(t["progress"] for t in d["tasks"]) / total, 1) if total else 0
    rows = [
        ("プロジェクト名", d["project"]["name"]),
        ("説明", d["project"]["description"]),
        ("期間", f"{d['project']['start_date'] or '-'} 〜 {d['project']['end_date'] or '-'}"),
        ("出力日", today),
        ("タスク総数", total), ("完了", done), ("期限超過", overdue),
        ("平均進捗", f"{avg}%"),
    ]
    for name, val in rows:
        ws.append([name, val])
    for r in range(1, len(rows) + 1):
        ws.cell(row=r, column=1).font = XFont(bold=True)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 60
    ws.append([])
    ws.append(["ステータス", "件数"])
    for s in d["statuses"]:
        ws.append([s["name"], sum(1 for t in d["tasks"] if t["status_id"] == s["id"])])
    _ignore_number_as_text(ws, "B1:B30")   # 「12.5%」等の表示用文字列

    # --- タスク一覧（WBS）シート
    ws2 = wb.create_sheet("タスク一覧(WBS)")
    headers = ["WBS", "ID", "タスク名", "ステータス", "担当者", "優先度", "開始日", "期限",
               "進捗%", "見積h", "実績h", "MS", "タグ", "説明"]
    headers += [f["label"] for f in cf_defs]
    ws2.append(headers)
    style_header(ws2, len(headers))
    for t in wbs_rows:
        row = [t["wbs"], t["id"], ("　" * t["depth"]) + t["title"],
               smap.get(t["status_id"], {}).get("name", ""),
               mmap.get(t["assignee_id"]) or t.get("assignee_label") or "",
               PRIORITY_LABEL.get(t["priority"], t["priority"]),
               t["start_date"] or "", t["due_date"] or "", t["progress"],
               t["estimate_h"], t["actual_h"], "○" if t["milestone"] else "",
               " / ".join(t["tags"]), t["description"]]
        row += [t["custom_values"].get(f["key"], "") for f in cf_defs]
        ws2.append(row)
    widths = [8, 6, 40, 12, 14, 8, 12, 12, 8, 8, 8, 5, 18, 50]
    for i, wd in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = wd
    if wbs_rows:
        _ignore_number_as_text(ws2, f"A2:A{1 + len(wbs_rows)}")   # WBS番号 "1.1" 等

    # --- コメントシート
    ws3 = wb.create_sheet("コメント")
    ws3.append(["タスクID", "タスク名", "投稿者", "日時", "本文"])
    style_header(ws3, 5)
    tmap = {t["id"]: t["title"] for t in d["tasks"]}
    for c in d["comments"]:
        ws3.append([c["task_id"], tmap.get(c["task_id"], ""),
                    c.get("author_name") or "-", c["created_at"], c["body"]])
    for col, wd in zip("ABCDE", [8, 32, 14, 20, 70]):
        ws3.column_dimensions[col].width = wd

    # --- 関連リンクシート
    ws4 = wb.create_sheet("関連リンク")
    ws4.append(["タスクID", "タスク名", "種別", "タイトル", "URL"])
    style_header(ws4, 5)
    for l in d["links"]:
        ws4.append([l["task_id"], tmap.get(l["task_id"], ""), l["kind"],
                    l["title"], l["url"]])
    for col, wd in zip("ABCDE", [8, 32, 10, 30, 60]):
        ws4.column_dimensions[col].width = wd

    # --- ノートシート
    ws5 = wb.create_sheet("ノート")
    ws5.append(["カテゴリ", "タイトル", "内容", "更新者", "更新日時"])
    style_header(ws5, 5)
    for n in d["notes"]:
        ws5.append([n["category"], n["title"], n["content"],
                    n.get("author_name") or "-", n["updated_at"] or n["created_at"]])
    for col, wd in zip("ABCDE", [12, 28, 70, 14, 20]):
        ws5.column_dimensions[col].width = wd

    # --- QAシート
    ws6 = wb.create_sheet("QA")
    ws6.append(["No", "分類", "件名", "質問", "質問者", "質問日", "回答担当",
                "回答期限", "ステータス", "回答", "回答日", "決定事項", "備考",
                "やり取り履歴"])
    style_header(ws6, 14)
    for q in d["qa"]:
        ws6.append([q["seq"], q["category"], q["title"], q["question"],
                    q["asker_name"], q["asked_at"], q.get("assignee_name") or "",
                    q["due_date"], QA_STATUS_LABEL.get(q["status"], q["status"]),
                    q["answer"], q["answered_at"], q.get("decision") or "",
                    q.get("note") or "",
                    "\n".join(_fmt_qa_comment_line(c) for c in q.get("comments", []))])
    for i, wd in enumerate([6, 10, 26, 44, 12, 11, 12, 11, 10, 44, 11, 30, 22, 48], 1):
        ws6.column_dimensions[get_column_letter(i)].width = wd

    # --- 課題シート
    ws7 = wb.create_sheet("課題")
    ws7.append(["No", "重要度", "分類", "件名", "課題内容", "起票者", "起票日", "担当",
                "対応期限", "状態", "方針", "実行内容", "解決日", "関連タスク"])
    style_header(ws7, 14)
    for q in d["issues"]:
        ws7.append([q["seq"], PRIORITY_LABEL.get(q["priority"], q["priority"]),
                    q["category"], q["title"], q["description"], q["raised_by"],
                    q["raised_at"], q.get("assignee_name") or "", q["due_date"],
                    ISSUE_STATUS_LABEL.get(q["status"], q["status"]),
                    q["policy"], q["action_plan"], q["resolved_at"],
                    " / ".join(t2["title"] for t2 in q.get("tasks", []))])
    for i, wd in enumerate([6, 8, 10, 26, 40, 11, 11, 12, 11, 10, 34, 34, 11, 30], 1):
        ws7.column_dimensions[get_column_letter(i)].width = wd

    bio = _save_wb_with_ignores(wb)
    fname = f"project_{pid}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------------------------------------------------------------- ビュー別 Excel

def _hex6(color, default="7C8DB5") -> str:
    c = str(color or "").lstrip("#")
    return c.upper() if len(c) == 6 else default


def _tint(hex6: str, ratio: float = 0.65) -> str:
    """白と混ぜて薄くした色（ratio=白の割合）。"""
    r, g, b = (int(hex6[i:i + 2], 16) for i in (0, 2, 4))
    return "".join(f"{int(v + (255 - v) * ratio):02X}" for v in (r, g, b))


def _assignee_hex(t, members_by_id, colors) -> str:
    if t.get("assignee_id"):
        m = members_by_id.get(t["assignee_id"])
        return _hex6(colors.get(f"u:{t['assignee_id']}")
                     or (m["color"] if m else None), "94A3B8")
    if t.get("assignee_label"):
        return _hex6(colors.get(f"v:{t['assignee_label']}") or "64748B", "64748B")
    return "94A3B8"


@app.post("/api/projects/{pid}/export/view.xlsx")
def export_view_xlsx(pid: int, body: dict):
    """画面単位のExcel出力（Web表示に近い見た目）。
    body: {view: table|wbs|board|calendar, ids?: 表示中タスクID（フィルタ反映）,
           colors?: 担当者色の上書きマップ, month?: [年, 月(0始まり)]}"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    view = body.get("view")
    if view not in ("table", "wbs", "board", "calendar", "qa", "kadai"):
        raise HTTPException(400, "view は table / wbs / board / calendar / qa / kadai を指定してください")
    colors = body.get("colors") or {}
    d = collect_export_data(pid)
    if body.get("ids"):
        keep = set(body["ids"])
        d["tasks"] = [t for t in d["tasks"] if t["id"] in keep]
    smap = {s["id"]: s for s in d["statuses"]}
    members_by_id = {m["id"]: m for m in d["members"]}
    mname = {m["id"]: m["name"] for m in d["members"]}
    today = date.today()

    thin = Side(style="thin", color="D8DEEA")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    vcenter = Alignment(vertical="center")
    head_fill = PatternFill("solid", fgColor="2F4870")
    head_font = XFont(bold=True, color="FFFFFF")

    wb = Workbook()
    _wb_default_font(wb)
    ws = wb.active

    def sfill(hex6):
        return PatternFill("solid", fgColor=hex6)

    def assignee_disp(t):
        return mname.get(t["assignee_id"]) or t.get("assignee_label") or ""

    # ---------- WBS / ガント（Excelガントチャート） ----------
    if view == "wbs":
        ws.title = "WBSガント"
        rows = build_wbs_rows(d["tasks"])
        dates = ([t["start_date"] for t in rows if t["start_date"]] +
                 [t["due_date"] for t in rows if t["due_date"]])
        pr = d["project"]
        if pr.get("start_date"):
            dates.append(pr["start_date"])
        if pr.get("end_date"):
            dates.append(pr["end_date"])
        if dates:
            d0 = datetime.strptime(min(dates), "%Y-%m-%d").date()
            d1 = datetime.strptime(max(dates), "%Y-%m-%d").date()
            ndays = min((d1 - d0).days + 1, 400)
        else:
            d0, ndays = today, 0
        days = [d0 + timedelta(days=i) for i in range(ndays)]
        LEFT = ["ID", "WBS", "タスク名", "担当", "開始", "期限", "進捗"]   # ID=取込時の突合キー
        NL = len(LEFT)
        # ヘッダー3段（月 / 日 / 曜日）
        for ci, label in enumerate(LEFT, 1):
            c = ws.cell(row=1, column=ci, value=label)
            c.fill, c.font, c.alignment, c.border = head_fill, head_font, center, grid
            ws.merge_cells(start_row=1, start_column=ci, end_row=3, end_column=ci)
        wd_jp = "月火水木金土日"
        month_start = 0
        for i, dt in enumerate(days):
            col = NL + 1 + i
            if i == 0 or dt.day == 1:
                if i > 0:
                    ws.merge_cells(start_row=1, start_column=NL + 1 + month_start,
                                   end_row=1, end_column=col - 1)
                mc = ws.cell(row=1, column=col, value=f"{dt.year}年{dt.month}月")
                mc.fill, mc.font, mc.alignment = head_fill, head_font, Alignment(horizontal="left", vertical="center")
                month_start = i
            dc = ws.cell(row=2, column=col, value=dt.day)
            wc = ws.cell(row=3, column=col, value=wd_jp[dt.weekday()])
            for c in (dc, wc):
                c.alignment, c.border = center, grid
                c.font = XFont(size=8, color="FFFFFF" if dt == today else
                              ("D04545" if dt.weekday() >= 5 else "475569"))
                if dt == today:
                    c.fill = sfill("E05252")
                elif dt.weekday() >= 5:
                    c.fill = sfill("EFF1F7")
                else:
                    c.fill = sfill("F8FAFC")
            ws.column_dimensions[get_column_letter(col)].width = 3.0
        if days:
            ws.merge_cells(start_row=1, start_column=NL + 1 + month_start,
                           end_row=1, end_column=NL + len(days))
        for ci, w in enumerate([5, 7, 38, 12, 11, 11, 7], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        # 明細行
        parents = {t["parent_id"] for t in d["tasks"] if t["parent_id"]}
        for ri, t in enumerate(rows, 4):
            is_parent = t["id"] in parents
            vals = [t["id"], t["wbs"],
                    ("　" * t["depth"]) + ("◆ " if t["milestone"] else "") + t["title"],
                    assignee_disp(t) or "—", t["start_date"] or "", t["due_date"] or "",
                    (t["progress"] or 0) / 100]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = grid
                c.alignment = vcenter if ci == 3 else center
                if ci == 1:
                    c.font = XFont(size=8, color="94A3B8")
                if ci == 7:
                    c.number_format = "0%"    # 数値として書き込み（文字列警告を出さない）
                if is_parent:
                    c.font = XFont(bold=True)
                    c.fill = sfill("EEF1F7")
                elif ci == 4 and (t["assignee_id"] or t.get("assignee_label")):
                    # 担当セルは画面同様に担当者色（淡色）で塗る
                    c.fill = sfill(_tint(_assignee_hex(t, members_by_id, colors), 0.72))
            ws.row_dimensions[ri].height = 18
            # 日付グリッド（週末シェード＋今日列＋バー）
            s = t["start_date"] or t["due_date"]
            e = t["due_date"] or t["start_date"]
            si = ei = None
            if s and days:
                sd = max(datetime.strptime(s, "%Y-%m-%d").date(), d0)
                ed = min(datetime.strptime(e, "%Y-%m-%d").date(), days[-1])
                if ed >= d0 and sd <= days[-1]:
                    si, ei = (sd - d0).days, (ed - d0).days
            bar_hex = "8B95A7" if is_parent else _assignee_hex(t, members_by_id, colors)
            done_cells = 0
            if si is not None:
                done_cells = round((ei - si + 1) * (t["progress"] or 0) / 100)
            for i, dt in enumerate(days):
                c = ws.cell(row=ri, column=NL + 1 + i)
                c.border = grid
                if t["milestone"] and e and dt.isoformat() == e:
                    c.value = "◆"
                    c.font = XFont(color="A855F7", bold=True)
                    c.alignment = center
                elif si is not None and si <= i <= ei:
                    # 進捗分は濃色・残りは淡色（Web のバー＋fill 表現に対応）
                    c.fill = sfill(bar_hex if i - si < done_cells else _tint(bar_hex, 0.62))
                elif dt == today:
                    c.fill = sfill("FDEAEA")
                elif dt.weekday() >= 5:
                    c.fill = sfill("F1F3F9")
        ws.freeze_panes = f"{get_column_letter(NL + 1)}4"
        if rows:
            _ignore_number_as_text(ws, f"B4:B{3 + len(rows)}")   # WBS番号 "1.1" 等

    # ---------- テーブル ----------
    elif view == "table":
        ws.title = "テーブル"
        cf_defs = d["project"].get("custom_fields") or []
        headers = ["ID", "WBS", "タスク名", "ステータス", "担当者", "優先度", "開始", "期限",
                   "進捗", "見積h", "実績h", "タグ"] + [f["label"] for f in cf_defs]
        for ci, hcell in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=hcell)
            c.fill, c.font, c.border, c.alignment = head_fill, head_font, grid, center
        rows = build_wbs_rows(d["tasks"])
        parents = {t["parent_id"] for t in d["tasks"] if t["parent_id"]}
        for ri, t in enumerate(rows, 2):
            st = smap.get(t["status_id"], {})
            done = bool(st.get("is_done"))
            vals = [t["id"], t["wbs"],
                    ("　" * t["depth"]) + ("◆ " if t["milestone"] else "") + t["title"],
                    st.get("name", "—"), assignee_disp(t) or "—",
                    PRIORITY_LABEL.get(t["priority"], t["priority"]),
                    t["start_date"] or "", t["due_date"] or "", (t["progress"] or 0) / 100,
                    t["estimate_h"], t["actual_h"], " / ".join(t["tags"])]
            vals += [t["custom_values"].get(f["key"], "") for f in cf_defs]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = grid
                c.alignment = vcenter if ci in (3, 12) else center
                if ci == 1:
                    c.font = XFont(size=8, color="94A3B8")
                if ci == 4 and st:                      # ステータス: 画面のバッジ同様に色付け
                    c.fill = sfill(_hex6(st.get("color"), "8B95A7"))
                    c.font = XFont(color="FFFFFF", bold=True, size=10)
                if ci == 5 and (t["assignee_id"] or t.get("assignee_label")):
                    c.fill = sfill(_tint(_assignee_hex(t, members_by_id, colors), 0.75))
                if ci == 8 and t["due_date"] and not done and (t["progress"] or 0) < 100 \
                        and t["due_date"] < today.isoformat():
                    c.font = XFont(color="DC2626", bold=True)   # 期限超過
                if ci == 9:
                    c.number_format = "0%"
                if t["id"] in parents and ci == 3:
                    c.font = XFont(bold=True)
        # 進捗のデータバー（画面のミニバー相当）
        if len(rows):
            from openpyxl.formatting.rule import DataBarRule
            ws.conditional_formatting.add(
                f"I2:I{len(rows) + 1}",
                DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                            color="4F6EF7", showValue=True))
        for ci, w in enumerate([5, 7, 40, 12, 14, 8, 11, 11, 9, 7, 7, 18], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"
        if rows:
            _ignore_number_as_text(ws, f"B2:B{1 + len(rows)}")   # WBS番号 "1.1" 等

    # ---------- ボード（カンバン） ----------
    elif view == "board":
        ws.title = "ボード"
        for ci, s in enumerate(d["statuses"], 1):
            col_tasks = [t for t in d["tasks"] if t["status_id"] == s["id"]]
            h = ws.cell(row=1, column=ci, value=f"{s['name']}（{len(col_tasks)}）")
            h.fill = sfill(_hex6(s["color"], "8B95A7"))
            h.font = XFont(bold=True, color="FFFFFF")
            h.alignment, h.border = center, grid
            ws.column_dimensions[get_column_letter(ci)].width = 34
            for ri, t in enumerate(col_tasks, 2):
                txt = (("◆ " if t["milestone"] else "") + t["title"] +
                       f"\n{assignee_disp(t) or '未割当'}"
                       f"{' / 期限 ' + t['due_date'] if t['due_date'] else ''}"
                       f" / {t['progress']}%")
                c = ws.cell(row=ri, column=ci, value=txt)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.border = grid
                c.fill = sfill(_tint(_assignee_hex(t, members_by_id, colors), 0.82))
                ws.row_dimensions[ri].height = max(
                    ws.row_dimensions[ri].height or 0, 34)

    # ---------- QA管理表（顧客提出用） ----------
    elif view == "qa":
        ws.title = "QA管理表"
        pr = d["project"]
        tcell = ws.cell(row=1, column=1, value=f"QA管理表 ― {pr['name']}")
        tcell.font = XFont(bold=True, size=14)
        total_qa = len(d["qa"])
        open_qa = sum(1 for q in d["qa"] if q["status"] in ("open", "pending"))
        mcell = ws.cell(row=2, column=1,
                        value=f"出力日: {today.isoformat()} ／ 全 {total_qa} 件"
                              f"（未回答・保留 {open_qa} 件）")
        mcell.font = XFont(color="64748B", size=10)
        headers = ["No", "分類", "件名・質問内容", "質問者", "質問日", "回答担当",
                   "回答期限", "ステータス", "回答", "回答日", "決定事項", "備考",
                   "やり取り履歴"]
        HR = 4
        for ci, hv in enumerate(headers, 1):
            c = ws.cell(row=HR, column=ci, value=hv)
            c.fill, c.font, c.border, c.alignment = head_fill, head_font, grid, center
        st_colors = {"open": "E05252", "pending": "F59E0B",
                     "answered": "4F6EF7", "closed": "16A34A"}
        wrap_top = Alignment(wrap_text=True, vertical="top")
        center_top = Alignment(horizontal="center", vertical="top", wrap_text=True)
        for ri, q in enumerate(d["qa"], HR + 1):
            qtext = q["title"] + (f"\n{q['question']}" if q["question"] else "")
            atext = q["answer"] or ""
            overdue = (q["due_date"] and q["status"] in ("open", "pending")
                       and q["due_date"] < today.isoformat())
            dtext = q.get("decision") or ""
            ntext = q.get("note") or ""
            ttext = "\n".join(_fmt_qa_comment_line(c) for c in q.get("comments", []))
            vals = [q["seq"], q["category"] or "", qtext, q["asker_name"] or "",
                    q["asked_at"] or "", q.get("assignee_name") or "",
                    q["due_date"] or "", QA_STATUS_LABEL.get(q["status"], q["status"]),
                    atext, q["answered_at"] or "", dtext, ntext, ttext]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = grid
                c.alignment = wrap_top if ci in (3, 9, 11, 12, 13) else center_top
                if ci == 8:
                    c.fill = sfill(st_colors.get(q["status"], "8B95A7"))
                    c.font = XFont(color="FFFFFF", bold=True, size=10)
                if ci == 13:
                    c.font = XFont(size=9, color="475569")
                if ci == 7 and overdue:
                    c.font = XFont(color="DC2626", bold=True)
            lines = max(qtext.count("\n") + 1, atext.count("\n") + 1,
                        dtext.count("\n") + 1, ntext.count("\n") + 1,
                        ttext.count("\n") + 1,
                        -(-len(qtext) // 28), -(-len(atext) // 32),
                        -(-len(dtext) // 18), -(-len(ntext) // 14), 1)
            ws.row_dimensions[ri].height = min(15 * lines + 6, 220)
        for ci, wd in enumerate([6, 10, 46, 12, 11, 12, 11, 11, 52, 11, 30, 22, 48], 1):
            ws.column_dimensions[get_column_letter(ci)].width = wd
        ws.freeze_panes = f"A{HR + 1}"

    # ---------- 課題管理表 ----------
    elif view == "kadai":
        ws.title = "課題管理表"
        pr = d["project"]
        tcell = ws.cell(row=1, column=1, value=f"課題管理表 ― {pr['name']}")
        tcell.font = XFont(bold=True, size=14)
        total_i = len(d["issues"])
        open_i = sum(1 for i in d["issues"] if i["status"] != "closed")
        mcell = ws.cell(row=2, column=1,
                        value=f"出力日: {today.isoformat()} ／ 全 {total_i} 件"
                              f"（オープン {open_i} 件）")
        mcell.font = XFont(color="64748B", size=10)
        headers = ["No", "重要度", "分類", "件名・課題内容", "起票者", "起票日",
                   "担当", "対応期限", "状態", "方針", "実行内容", "解決日",
                   "関連タスク", "コメント履歴"]
        HR = 4
        for ci, hv in enumerate(headers, 1):
            c = ws.cell(row=HR, column=ci, value=hv)
            c.fill, c.font, c.border, c.alignment = head_fill, head_font, grid, center
        ist_colors = {"open": "E05252", "doing": "F59E0B",
                      "resolved": "16A34A", "closed": "64748B"}
        wrap_top = Alignment(wrap_text=True, vertical="top")
        center_top = Alignment(horizontal="center", vertical="top", wrap_text=True)
        for ri, q in enumerate(d["issues"], HR + 1):
            qtext = q["title"] + (f"\n{q['description']}" if q["description"] else "")
            overdue = (q["due_date"] and q["status"] in ("open", "doing")
                       and q["due_date"] < today.isoformat())
            ttext = "\n".join(_fmt_qa_comment_line(c) for c in q.get("comments", []))
            reltext = "\n".join(t2["title"] for t2 in q.get("tasks", []))
            vals = [q["seq"], PRIORITY_LABEL.get(q["priority"], q["priority"]),
                    q["category"] or "", qtext, q["raised_by"] or "",
                    q["raised_at"] or "", q.get("assignee_name") or "",
                    q["due_date"] or "",
                    ISSUE_STATUS_LABEL.get(q["status"], q["status"]),
                    q["policy"] or "", q["action_plan"] or "",
                    q["resolved_at"] or "", reltext, ttext]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = grid
                c.alignment = wrap_top if ci in (4, 10, 11, 13, 14) else center_top
                if ci == 9:
                    c.fill = sfill(ist_colors.get(q["status"], "8B95A7"))
                    c.font = XFont(color="FFFFFF", bold=True, size=10)
                if ci == 14:
                    c.font = XFont(size=9, color="475569")
                if ci == 8 and overdue:
                    c.font = XFont(color="DC2626", bold=True)
                if ci == 2 and q["priority"] in ("highest", "high"):
                    c.font = XFont(color="DC2626", bold=True)
            texts = [qtext, q["policy"] or "", q["action_plan"] or "", ttext, reltext]
            lines = max([t3.count("\n") + 1 for t3 in texts] +
                        [-(-len(qtext) // 26), -(-len(q["policy"] or "") // 22),
                         -(-len(q["action_plan"] or "") // 22), 1])
            ws.row_dimensions[ri].height = min(15 * lines + 6, 220)
        for ci, wd in enumerate([6, 8, 10, 40, 11, 11, 11, 11, 10, 34, 34, 11, 24, 40], 1):
            ws.column_dimensions[get_column_letter(ci)].width = wd
        ws.freeze_panes = f"A{HR + 1}"

    # ---------- カレンダー（Google式の横断バー） ----------
    else:
        ws.title = "カレンダー"
        ym = body.get("month") or [today.year, today.month - 1]
        y, mo = int(ym[0]), int(ym[1]) + 1
        first = date(y, mo, 1)
        start_dow = first.weekday()               # 月曜=0
        days_in_month = (date(y + (mo == 12), mo % 12 + 1, 1) - timedelta(days=1)).day
        grid_start = first - timedelta(days=start_dow)
        n_weeks = -(-(start_dow + days_in_month) // 7)
        ws.cell(row=1, column=1, value=f"{y}年 {mo}月")
        ws.cell(row=1, column=1).font = XFont(bold=True, size=14)
        for ci, wd in enumerate("月火水木金土日", 1):
            c = ws.cell(row=2, column=ci, value=wd)
            c.fill, c.font, c.alignment, c.border = head_fill, head_font, center, grid
            ws.column_dimensions[get_column_letter(ci)].width = 22
        spans = []
        for t in d["tasks"]:
            s = t["start_date"] or t["due_date"]
            e = t["due_date"] or t["start_date"]
            if s:
                spans.append((s, e, t))
        spans.sort(key=lambda x: (x[0], x[1]))
        r = 3
        for w in range(n_weeks):
            ws_d = grid_start + timedelta(days=w * 7)
            we_d = ws_d + timedelta(days=6)
            # 日付行
            for i in range(7):
                dt = ws_d + timedelta(days=i)
                c = ws.cell(row=r, column=i + 1, value=dt.day)
                c.border = grid
                c.alignment = Alignment(horizontal="right", vertical="top")
                if dt == today:
                    c.fill = sfill("EEF3FF")
                    c.font = XFont(bold=True, color="4F6EF7")
                elif dt.month != mo:
                    c.font = XFont(color="B6BFCF")
                elif i >= 5:
                    c.fill = sfill("FAFBFE")
            # レーン割当（貪欲法）→ 横結合セルのバー
            lanes: list = []
            segs = []
            for s, e, t in spans:
                sd = datetime.strptime(s, "%Y-%m-%d").date()
                ed = datetime.strptime(e, "%Y-%m-%d").date()
                if ed < ws_d or sd > we_d:
                    continue
                si = 0 if sd <= ws_d else (sd - ws_d).days
                ei = 6 if ed >= we_d else (ed - ws_d).days
                lane = next((i for i, last in enumerate(lanes) if last < si), None)
                if lane is None:
                    lane = len(lanes)
                    lanes.append(-1)
                lanes[lane] = ei
                segs.append((lane, si, ei, t))
            for lane, si, ei, t in segs:
                rr = r + 1 + lane
                st = smap.get(t["status_id"], {})
                if ei > si:
                    ws.merge_cells(start_row=rr, start_column=si + 1,
                                   end_row=rr, end_column=ei + 1)
                c = ws.cell(row=rr, column=si + 1,
                            value=("◆ " if t["milestone"] else "") + t["title"])
                c.fill = sfill(_hex6(st.get("color"), "8B95A7"))
                c.font = XFont(color="FFFFFF", size=9,
                              strike=bool(st.get("is_done")))
                c.alignment = Alignment(vertical="center")
                for cc in range(si + 1, ei + 2):
                    ws.cell(row=rr, column=cc).border = grid
                ws.row_dimensions[rr].height = 15
            r += 1 + max(len(lanes), 3) + 1   # 日付行＋レーン行＋余白1行

    bio = _save_wb_with_ignores(wb)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="view_{view}_{pid}.xlsx"'})


# ---------------------------------------------------------------- export: HTML

@app.get("/api/projects/{pid}/export.html")
def export_html(pid: int):
    """案件終了後のアーカイブ用・完全スタンドアロンHTMLレポート。"""
    import html as h

    d = collect_export_data(pid)
    p = d["project"]
    smap = {s["id"]: s for s in d["statuses"]}
    mmap = {m["id"]: m["name"] for m in d["members"]}
    wbs_rows = build_wbs_rows(d["tasks"])
    today = date.today().isoformat()
    total = len(d["tasks"])
    done = sum(1 for t in d["tasks"] if smap.get(t["status_id"], {}).get("is_done"))
    avg = round(sum(t["progress"] for t in d["tasks"]) / total, 1) if total else 0

    comments_by_task: dict[int, list] = {}
    for c in d["comments"]:
        comments_by_task.setdefault(c["task_id"], []).append(c)
    links_by_task: dict[int, list] = {}
    for l in d["links"]:
        links_by_task.setdefault(l["task_id"], []).append(l)

    def esc(s):
        return h.escape(str(s if s is not None else ""))

    status_cells = "".join(
        f'<div class="stat"><div class="dot" style="background:{esc(s["color"])}"></div>'
        f'{esc(s["name"])}: <b>{sum(1 for t in d["tasks"] if t["status_id"] == s["id"])}</b></div>'
        for s in d["statuses"])

    kanban_cols = ""
    for s in d["statuses"]:
        cards = "".join(
            f'<div class="card"><b>{esc(t["title"])}</b>'
            f'<span>{esc(mmap.get(t["assignee_id"]) or t.get("assignee_label") or "未割当")} / '
            f'期限 {esc(t["due_date"] or "-")}</span></div>'
            for t in d["tasks"] if t["status_id"] == s["id"])
        kanban_cols += (
            f'<div class="col"><h4 style="border-color:{esc(s["color"])}">'
            f'{esc(s["name"])}</h4>{cards}</div>')

    table_rows = ""
    for t in wbs_rows:
        st = smap.get(t["status_id"], {})
        pad = t["depth"] * 18
        table_rows += (
            f'<tr><td>{esc(t["wbs"])}</td>'
            f'<td style="padding-left:{pad + 8}px">{esc(t["title"])}'
            f'{" ◆" if t["milestone"] else ""}</td>'
            f'<td><span class="badge" style="background:{esc(st.get("color", "#999"))}">'
            f'{esc(st.get("name", "-"))}</span></td>'
            f'<td>{esc(mmap.get(t["assignee_id"]) or t.get("assignee_label") or "未割当")}</td>'
            f'<td>{esc(PRIORITY_LABEL.get(t["priority"], t["priority"]))}</td>'
            f'<td>{esc(t["start_date"] or "-")}</td><td>{esc(t["due_date"] or "-")}</td>'
            f'<td><div class="pbar"><div style="width:{t["progress"]}%"></div></div>'
            f'{t["progress"]}%</td></tr>')

    qa_st_color = {"open": "#e05252", "pending": "#f59e0b",
                   "answered": "#4f6ef7", "closed": "#16a34a"}
    qa_rows = "".join(
        f'<tr><td>QA-{q["seq"]:03d}</td><td>{esc(q["category"])}</td>'
        f'<td><b>{esc(q["title"])}</b>'
        f'{"<div class=desc>" + esc(q["question"]) + "</div>" if q["question"] else ""}</td>'
        f'<td>{esc(q["asker_name"])}</td><td>{esc(q["asked_at"] or "")}</td>'
        f'<td>{esc(q.get("assignee_name") or "")}</td><td>{esc(q["due_date"] or "")}</td>'
        f'<td><span class="badge" style="background:{qa_st_color.get(q["status"], "#8b95a7")}">'
        f'{esc(QA_STATUS_LABEL.get(q["status"], q["status"]))}</span></td>'
        f'<td class="desc">{esc(q["answer"])}</td><td>{esc(q["answered_at"] or "")}</td>'
        f'<td class="desc">{esc(q.get("decision") or "")}</td>'
        f'<td class="desc">{esc(q.get("note") or "")}</td>'
        f'<td class="desc">{esc(chr(10).join(_fmt_qa_comment_line(c) for c in q.get("comments", [])))}</td></tr>'
        for q in d["qa"])

    iss_st_color = {"open": "#e05252", "doing": "#f59e0b",
                    "resolved": "#16a34a", "closed": "#64748b"}
    issue_rows = "".join(
        f'<tr><td>ISS-{q["seq"]:03d}</td>'
        f'<td>{esc(PRIORITY_LABEL.get(q["priority"], q["priority"]))}</td>'
        f'<td><b>{esc(q["title"])}</b>'
        f'{"<div class=desc>" + esc(q["description"]) + "</div>" if q["description"] else ""}</td>'
        f'<td>{esc(q["raised_by"])}</td><td>{esc(q["raised_at"] or "")}</td>'
        f'<td>{esc(q.get("assignee_name") or "")}</td><td>{esc(q["due_date"] or "")}</td>'
        f'<td><span class="badge" style="background:{iss_st_color.get(q["status"], "#8b95a7")}">'
        f'{esc(ISSUE_STATUS_LABEL.get(q["status"], q["status"]))}</span></td>'
        f'<td class="desc">{esc(q["policy"])}</td><td class="desc">{esc(q["action_plan"])}</td>'
        f'<td class="desc">{esc(chr(10).join(t2["title"] for t2 in q.get("tasks", [])))}</td></tr>'
        for q in d["issues"])

    notes_sections = ""
    for n in d["notes"]:
        notes_sections += (
            f'<details open><summary>{"📌 " if n["pinned"] else ""}'
            f'[{esc(n["category"])}] {esc(n["title"])}'
            f'<span class="meta">（{esc(n.get("author_name") or "-")} / '
            f'{esc((n["updated_at"] or n["created_at"] or "")[:10])}）</span></summary>'
            f'<p class="desc">{esc(n["content"])}</p></details>')

    act_rows = "".join(
        f'<tr><td>{esc((a["created_at"] or "")[:16])}</td>'
        f'<td>{esc(a.get("actor_name") or "システム")}</td>'
        f'<td>{esc(a["action"])}</td>'
        f'<td>{esc(a.get("task_title") or "")}</td><td>{esc(a["detail"])}</td></tr>'
        for a in d["activities"])

    detail_sections = ""
    for t in wbs_rows:
        cs = comments_by_task.get(t["id"], [])
        ls = links_by_task.get(t["id"], [])
        if not cs and not ls and not t["description"]:
            continue
        body = ""
        if t["description"]:
            body += f'<p class="desc">{esc(t["description"])}</p>'
        if ls:
            body += "<ul>" + "".join(
                f'<li>[{esc(l["kind"])}] {esc(l["title"])} — {esc(l["url"])}</li>'
                for l in ls) + "</ul>"
        for c in cs:
            body += (f'<div class="cmt"><b>{esc(c.get("author_name") or "-")}</b> '
                     f'<span>{esc(c["created_at"])}</span><p>{esc(c["body"])}</p></div>')
        detail_sections += (
            f'<details><summary>{esc(t["wbs"])} {esc(t["title"])} '
            f'（コメント {len(cs)} 件）</summary>{body}</details>')

    html_doc = f"""<!DOCTYPE html>
<html lang="ja"><head><meta charset="utf-8">
<title>{esc(p["name"])} - プロジェクトレポート</title>
<style>
body{{font-family:'Segoe UI','Hiragino Sans','Meiryo',sans-serif;margin:0;background:#f4f6fa;color:#1f2937}}
.wrap{{max-width:1100px;margin:0 auto;padding:32px 20px}}
h1{{border-left:6px solid {esc(p["color"])};padding-left:12px}}
h2{{margin-top:40px;border-bottom:2px solid #d8deea;padding-bottom:6px}}
.meta{{color:#64748b;font-size:14px}}
.stats{{display:flex;gap:18px;flex-wrap:wrap;margin:16px 0}}
.stat{{background:#fff;border:1px solid #e2e8f0;border-radius:8px;padding:10px 16px;display:flex;align-items:center;gap:8px}}
.dot{{width:10px;height:10px;border-radius:50%}}
.kanban{{display:flex;gap:12px;overflow-x:auto;padding-bottom:8px}}
.col{{background:#eef1f7;border-radius:8px;padding:10px;min-width:220px;flex:1}}
.col h4{{margin:0 0 8px;border-left:4px solid;padding-left:8px}}
.card{{background:#fff;border-radius:6px;padding:8px 10px;margin-bottom:8px;box-shadow:0 1px 2px rgba(0,0,0,.08)}}
.card span{{display:block;color:#64748b;font-size:12px;margin-top:2px}}
table{{width:100%;border-collapse:collapse;background:#fff;font-size:13px}}
th,td{{border:1px solid #e2e8f0;padding:6px 8px;text-align:left;vertical-align:middle}}
th{{background:#2f4870;color:#fff;position:sticky;top:0}}
.badge{{color:#fff;border-radius:10px;padding:2px 8px;font-size:11px;white-space:nowrap}}
.pbar{{background:#e2e8f0;border-radius:4px;height:8px;width:80px;display:inline-block;vertical-align:middle;margin-right:6px}}
.pbar div{{background:#4f6ef7;height:8px;border-radius:4px}}
details{{background:#fff;border:1px solid #e2e8f0;border-radius:8px;margin:8px 0;padding:10px 14px}}
summary{{cursor:pointer;font-weight:600}}
.cmt{{border-left:3px solid #cbd5e1;margin:8px 0;padding:4px 10px}}
.cmt span{{color:#94a3b8;font-size:12px;margin-left:8px}}
.cmt p{{margin:4px 0;white-space:pre-wrap}}
.desc{{white-space:pre-wrap;color:#475569}}
@media print{{.kanban{{overflow:visible}}body{{background:#fff}}}}
</style></head><body><div class="wrap">
<h1>{esc(p["name"])}</h1>
<p class="meta">期間: {esc(p["start_date"] or "-")} 〜 {esc(p["end_date"] or "-")} ／
出力日: {today} ／ タスク {total} 件・完了 {done} 件・平均進捗 {avg}%</p>
<p>{esc(p["description"])}</p>
<div class="stats">{status_cells}</div>
<h2>カンバン</h2><div class="kanban">{kanban_cols}</div>
<h2>タスク一覧（WBS）</h2>
<table><thead><tr><th>WBS</th><th>タスク名</th><th>ステータス</th><th>担当</th>
<th>優先度</th><th>開始</th><th>期限</th><th>進捗</th></tr></thead>
<tbody>{table_rows}</tbody></table>
<h2>タスク詳細・議論ログ</h2>{detail_sections}
<h2>QA（全 {len(d["qa"])} 件）</h2>
{'<table><thead><tr><th>No</th><th>分類</th><th>件名・質問</th><th>質問者</th><th>質問日</th><th>回答担当</th><th>回答期限</th><th>状態</th><th>回答</th><th>回答日</th><th>決定事項</th><th>備考</th><th>やり取り履歴</th></tr></thead><tbody>' + qa_rows + '</tbody></table>' if d["qa"] else '<p class="meta">なし</p>'}
<h2>課題（全 {len(d["issues"])} 件・オープン {sum(1 for q in d["issues"] if q["status"] != "closed")} 件）</h2>
{'<table><thead><tr><th>No</th><th>重要度</th><th>件名・内容</th><th>起票者</th><th>起票日</th><th>担当</th><th>期限</th><th>状態</th><th>方針</th><th>実行内容</th><th>関連タスク</th></tr></thead><tbody>' + issue_rows + '</tbody></table>' if d["issues"] else '<p class="meta">なし</p>'}
<h2>ノート</h2>{notes_sections or '<p class="meta">なし</p>'}
<h2>アクティビティ履歴（全 {len(d["activities"])} 件）</h2>
<details><summary>履歴を開く</summary>
<table><thead><tr><th>日時</th><th>操作者</th><th>操作</th><th>タスク</th><th>内容</th></tr></thead>
<tbody>{act_rows}</tbody></table></details>
<p class="meta" style="margin-top:40px">Generated by PJ Board</p>
</div></body></html>"""
    fname = f"report_{pid}_{today}.html"
    return Response(html_doc, media_type="text/html; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------------------------------------------------------------- AI向けAPI

AI_HELP = """# PJ Board — AIエージェント向けガイド

社内プロジェクト管理ツールの REST API。すべて JSON。日付は YYYY-MM-DD。

## まず読む
- GET /api/ai/context                 … 全プロジェクトの要約（軽量JSON）
- GET /api/ai/context?project_id=N    … 1プロジェクトの全データ（WBS・ノート・コメント込み）
- GET /openapi.json                   … 全エンドポイントのスキーマ

## 主要操作
- GET  /api/bootstrap                          … プロジェクト/ユーザー/組織の一覧
- GET  /api/projects/{pid}/data                … タスク・ステータス・メンバー一式
- POST /api/projects/{pid}/tasks               … タスク作成 {title, status_id?, assignee_id?, start_date?, due_date?, parent_id?, actor_id?}
- PATCH /api/tasks/{tid}                       … 部分更新（渡したフィールドのみ反映）
- GET  /api/tasks/{tid}/detail                 … 詳細＋コメント＋リンク＋履歴
- POST /api/tasks/{tid}/comments               … コメント {body, author_id}
- GET  /api/projects/{pid}/discussions         … コメントスレッド一覧
- GET  /api/projects/{pid}/notes               … PJの共有ノート（環境・体制・ルール）
- GET  /api/overview?user_id=N                 … ユーザー視点の横断サマリー
- GET  /api/projects/{pid}/export.json         … 全データダンプ

## 権限の約束事
- 書き込み時は actor_id（操作ユーザーのID）を渡すこと。権限は
  組織ロール(manager/professional=全PJ管理者) × PJロール(leader/member/advisor/external) で決まる。
- actor_id を省略するとチェックなしの管理操作になる（自動化スクリプト向け）。
- ステータスのID・名前は /api/projects/{pid}/data の statuses を参照。is_done=1 の列が完了扱い。

## MCP
同梱の mcp_server.py を MCP クライアントに登録すると、上記操作をツールとして呼べる。
環境変数: PJBOARD_URL（既定 http://localhost:8100）, PJBOARD_ACTOR_ID（操作ユーザーID）
"""


@app.get("/api/ai/help", response_class=Response)
def ai_help():
    """AIエージェント向けの使い方ガイド（プレーンテキスト）。"""
    return Response(AI_HELP, media_type="text/markdown; charset=utf-8")


@app.get("/api/ai/context")
def ai_context(project_id: Optional[int] = None):
    """LLMが読み込みやすい形の要約JSON。project_id 指定で1PJの全データを返す。"""
    today = date.today().isoformat()
    with db() as conn:
        if project_id is None:
            out = {"today": today, "projects": []}
            for pr in conn.execute("SELECT * FROM projects ORDER BY id"):
                p = project_row_to_dict(pr)
                smap = {r["id"]: r for r in conn.execute(
                    "SELECT * FROM statuses WHERE project_id=?", (p["id"],))}
                tasks = [task_row_to_dict(r) for r in conn.execute(
                    "SELECT * FROM tasks WHERE project_id=? AND deleted_at IS NULL", (p["id"],))]
                done = sum(1 for t in tasks
                           if smap.get(t["status_id"]) and smap[t["status_id"]]["is_done"])
                overdue = sum(1 for t in tasks if t["due_date"] and t["due_date"] < today
                              and not (smap.get(t["status_id"]) and smap[t["status_id"]]["is_done"]))
                out["projects"].append({
                    "id": p["id"], "name": p["name"], "status": p["status"],
                    "period": [p["start_date"], p["end_date"]],
                    "tasks_total": len(tasks), "tasks_done": done, "overdue": overdue,
                })
            out["users"] = [
                {"id": r["id"], "name": r["name"], "org_role": r["org_role"],
                 "account_type": r["account_type"]}
                for r in conn.execute("SELECT * FROM members WHERE active=1")]
            return out

        d = collect_export_data(project_id)
        smap = {s["id"]: s for s in d["statuses"]}
        mmap = {m["id"]: m["name"] for m in d["members"]}
        wbs = [{
            "wbs": t["wbs"], "id": t["id"], "title": t["title"],
            "status": smap.get(t["status_id"], {}).get("name"),
            "done": bool(smap.get(t["status_id"], {}).get("is_done")),
            "assignee": mmap.get(t["assignee_id"]) or t.get("assignee_label"),
            "priority": t["priority"], "start": t["start_date"], "due": t["due_date"],
            "progress": t["progress"], "milestone": t["milestone"],
            "parent_id": t["parent_id"], "deps": t["deps"],
            "description": t["description"],
        } for t in build_wbs_rows(d["tasks"])]
        with db() as conn2:
            notes = rows_to_dicts(conn2.execute(
                "SELECT category, title, content, updated_at FROM project_notes"
                " WHERE project_id=? ORDER BY pinned DESC, sort_order", (project_id,)))
        comments = [{
            "task_id": c["task_id"], "author": c.get("author_name"),
            "at": c["created_at"], "body": c["body"],
        } for c in d["comments"]]
        return {"today": today, "project": {
            "id": d["project"]["id"], "name": d["project"]["name"],
            "description": d["project"]["description"],
            "period": [d["project"]["start_date"], d["project"]["end_date"]],
            "settings": d["project"]["settings"],
        }, "statuses": [{"id": s["id"], "name": s["name"], "is_done": bool(s["is_done"])}
                        for s in d["statuses"]],
            "wbs": wbs, "notes": notes, "comments": comments}



# ---------------------------------------------------------------- API: ユーザー表示設定（担当者カラー・ダッシュボード配置など）

@app.get("/api/prefs")
def get_prefs(user_id: Optional[int] = None):
    uid = resolve_uid(user_id)
    if uid is None:
        return {}
    with db() as conn:
        rows = conn.execute("SELECT key, value FROM user_prefs WHERE member_id=?",
                            (uid,)).fetchall()
    out = {}
    for r in rows:
        try:
            out[r["key"]] = json.loads(r["value"])
        except ValueError:
            pass
    return out


@app.post("/api/prefs")
def set_pref(body: dict):
    uid = resolve_uid(body.get("user_id"))
    key = (body.get("key") or "").strip()
    if uid is None or not key:
        raise HTTPException(400, "user_id と key が必要です")
    with db() as conn:
        if body.get("value") is None:
            conn.execute("DELETE FROM user_prefs WHERE member_id=? AND key=?", (uid, key))
        else:
            conn.execute(
                "INSERT OR REPLACE INTO user_prefs(member_id, key, value) VALUES(?,?,?)",
                (uid, key, json.dumps(body["value"], ensure_ascii=False)))
    return {"ok": True}


@app.post("/api/me/webhook-test")
def webhook_test(body: dict):
    """ユーザー設定の通知転送Webhookへテスト送信（同期・結果を返す）。"""
    su = CURRENT_USER.get()
    uid = su["id"] if su else resolve_uid(body.get("user_id"))
    if uid is None:
        raise HTTPException(401, "ログインが必要です")
    url = (body.get("url") or "").strip()
    if not url.startswith("https://"):
        raise HTTPException(400, "https:// で始まる Incoming Webhook URL を指定してください")
    provider = body.get("provider") or "auto"
    if provider == "auto":
        provider = detect_webhook_provider(url)
    import urllib.request
    try:
        data = json.dumps(
            webhook_payload(provider,
                            "✅ PJ Board: 通知転送のテスト送信です。この通知が見えていれば設定は有効です"),
            ensure_ascii=False).encode()
        req = urllib.request.Request(url, data=data,
                                     headers={"Content-Type": "application/json"})
        urllib.request.urlopen(req, timeout=10)
    except Exception as e:
        raise HTTPException(400, f"送信に失敗しました: {e}")
    return {"ok": True, "provider": provider}


# ---------------------------------------------------------------- API: 通知

@app.get("/api/notifications")
def list_notifications(user_id: Optional[int] = None, unread_only: int = 0):
    uid = resolve_uid(user_id)
    if uid is None:
        return {"items": [], "unread": 0}
    with db() as conn:
        q = ("SELECT n.*, m.name actor_name, t.title task_title, p.name project_name"
             " FROM notifications n"
             " LEFT JOIN members m ON m.id=n.actor_id"
             " LEFT JOIN tasks t ON t.id=n.task_id"
             " LEFT JOIN projects p ON p.id=n.project_id"
             " WHERE n.user_id=?")
        if unread_only:
            q += " AND n.read=0"
        items = rows_to_dicts(conn.execute(q + " ORDER BY n.id DESC LIMIT 30", (uid,)))
        unread = conn.execute(
            "SELECT COUNT(*) c FROM notifications WHERE user_id=? AND read=0",
            (uid,)).fetchone()["c"]
    return {"items": items, "unread": unread}


@app.post("/api/notifications/read")
def read_notifications(body: dict):
    uid = resolve_uid(body.get("user_id"))
    ids = body.get("ids")
    with db() as conn:
        if ids:
            ph = ",".join("?" * len(ids))
            conn.execute(f"UPDATE notifications SET read=1 WHERE user_id=? AND id IN ({ph})",
                         [uid] + list(ids))
        else:
            conn.execute("UPDATE notifications SET read=1 WHERE user_id=?", (uid,))
    return {"ok": True}


# ---------------------------------------------------------------- API: ウォッチ

@app.post("/api/tasks/{tid}/watch")
def watch_task(tid: int, body: dict):
    uid = resolve_uid(body.get("user_id"))
    with db() as conn:
        get_task_or_404(conn, tid)
        if uid:
            conn.execute("INSERT OR IGNORE INTO watchers(task_id, member_id) VALUES(?,?)",
                         (tid, uid))
    return {"ok": True}


@app.delete("/api/tasks/{tid}/watch")
def unwatch_task(tid: int, user_id: Optional[int] = None):
    uid = resolve_uid(user_id)
    with db() as conn:
        conn.execute("DELETE FROM watchers WHERE task_id=? AND member_id=?", (tid, uid))
    return {"ok": True}


# ---------------------------------------------------------------- API: ファイル添付

@app.post("/api/upload")
async def upload_file(request: Request, target_type: str, target_id: int,
                      filename: str, actor_id: Optional[int] = None):
    """添付アップロード（本文=ファイルのバイト列。multipart不要の素朴方式）。"""
    actor = resolve_uid(actor_id)
    if target_type not in ("task", "note"):
        raise HTTPException(400, "target_type は task / note")
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_EXTS:
        raise HTTPException(400, f"この拡張子は添付できません ({ext})")
    data = await request.body()
    if len(data) > MAX_FILE_SIZE:
        raise HTTPException(413, "ファイルサイズは20MBまでです")
    if not data:
        raise HTTPException(400, "ファイルが空です")
    stored = f"{secrets.token_hex(12)}{ext}"
    with open(os.path.join(FILES_DIR, stored), "wb") as f:
        f.write(data)
    with db() as conn:
        # 権限: task はコメント可否と同等の緩い基準、note は編集権限
        if target_type == "task":
            t = get_task_or_404(conn, target_id)
            if actor and not can_comment_in(conn, t["project_id"], actor):
                raise HTTPException(403, "添付する権限がありません")
        cur = conn.execute(
            "INSERT INTO attachments(target_type, target_id, filename, stored_name,"
            " size, content_type, uploaded_by, created_at) VALUES(?,?,?,?,?,?,?,?)",
            (target_type, target_id, filename, stored, len(data),
             request.headers.get("content-type", "application/octet-stream"),
             actor, now()))
        r = conn.execute("SELECT * FROM attachments WHERE id=?", (cur.lastrowid,)).fetchone()
    return dict(r)


@app.get("/api/files/{aid}")
def download_file(aid: int):
    with db() as conn:
        r = conn.execute("SELECT * FROM attachments WHERE id=?", (aid,)).fetchone()
    if not r:
        raise HTTPException(404, "file not found")
    # 外部ユーザー: 添付の所属プロジェクトのメンバーでなければ拒否
    with db() as conn:
        if r["target_type"] == "task":
            p0 = conn.execute("SELECT project_id FROM tasks WHERE id=?",
                              (r["target_id"],)).fetchone()
        else:
            p0 = conn.execute("SELECT project_id FROM project_notes WHERE id=?",
                              (r["target_id"],)).fetchone()
    if p0:
        external_guard(p0["project_id"],
                       "notes" if r["target_type"] == "note" else None)
    path = os.path.join(FILES_DIR, r["stored_name"])
    if not os.path.exists(path):
        raise HTTPException(404, "file missing on disk")
    from urllib.parse import quote
    with open(path, "rb") as f:
        data = f.read()
    return Response(data, media_type="application/octet-stream",
                    headers={"Content-Disposition":
                             f"attachment; filename*=UTF-8''{quote(r['filename'])}"})


@app.delete("/api/attachments/{aid}")
def delete_attachment(aid: int, actor_id: Optional[int] = None):
    actor = resolve_uid(actor_id)
    with db() as conn:
        r = conn.execute("SELECT * FROM attachments WHERE id=?", (aid,)).fetchone()
        if r:
            if actor and r["uploaded_by"] not in (None, actor):
                # 管理者はどこのPJか特定して判定（taskのみ厳密化・緩め運用）
                pass
            try:
                os.remove(os.path.join(FILES_DIR, r["stored_name"]))
            except OSError:
                pass
            conn.execute("DELETE FROM attachments WHERE id=?", (aid,))
    return {"ok": True}


def attachments_for(conn, target_type, target_id):
    return rows_to_dicts(conn.execute(
        "SELECT a.*, m.name uploaded_by_name FROM attachments a"
        " LEFT JOIN members m ON m.id=a.uploaded_by"
        " WHERE a.target_type=? AND a.target_id=? ORDER BY a.id",
        (target_type, target_id)))


# ---------------------------------------------------------------- API: 全文検索

@app.get("/api/search")
def search(q: str, user_id: Optional[int] = None, limit: int = 30):
    """タスク・コメント・ノートの横断検索（LIKE。外部ユーザーは参加PJのみ）。"""
    uid = resolve_uid(user_id)
    q = q.strip()
    if not q:
        return {"tasks": [], "comments": [], "notes": []}
    like = f"%{q}%"
    with db() as conn:
        pid_filter = ""
        params_base = []
        if uid:
            m = conn.execute("SELECT * FROM members WHERE id=?", (uid,)).fetchone()
            if m and m["account_type"] == "external":
                pids = [str(r["project_id"]) for r in conn.execute(
                    "SELECT project_id FROM project_members WHERE member_id=?", (uid,))]
                pid_filter = f" AND project_id IN ({','.join(pids) or '0'})"
        tasks = rows_to_dicts(conn.execute(
            "SELECT t.id, t.title, t.project_id, t.status_id, p.name project_name"
            " FROM tasks t JOIN projects p ON p.id=t.project_id"
            " WHERE t.deleted_at IS NULL AND (t.title LIKE ? OR t.description LIKE ?)"
            + pid_filter.replace("project_id", "t.project_id") +
            " ORDER BY t.updated_at DESC LIMIT ?", (like, like, limit)))
        comments = rows_to_dicts(conn.execute(
            "SELECT c.id, c.task_id, c.body, c.created_at, t.title task_title,"
            " t.project_id, m.name author_name"
            " FROM comments c JOIN tasks t ON t.id=c.task_id"
            " LEFT JOIN members m ON m.id=c.author_id"
            " WHERE t.deleted_at IS NULL AND c.body LIKE ?"
            + pid_filter.replace("project_id", "t.project_id") +
            " ORDER BY c.id DESC LIMIT ?", (like, limit)))
        notes = rows_to_dicts(conn.execute(
            "SELECT n.id, n.title, n.category, n.project_id, p.name project_name"
            " FROM project_notes n JOIN projects p ON p.id=n.project_id"
            " WHERE n.deleted_at IS NULL AND (n.title LIKE ? OR n.content LIKE ?)"
            + pid_filter.replace("project_id", "n.project_id") +
            " ORDER BY n.updated_at DESC LIMIT ?", (like, like, limit)))
    return {"tasks": tasks, "comments": comments, "notes": notes}


# ---------------------------------------------------------------- API: ゴミ箱

@app.get("/api/projects/{pid}/trash")
def list_trash(pid: int):
    with db() as conn:
        tasks = rows_to_dicts(conn.execute(
            "SELECT id, title, deleted_at FROM tasks"
            " WHERE project_id=? AND deleted_at IS NOT NULL ORDER BY deleted_at DESC",
            (pid,)))
        notes = rows_to_dicts(conn.execute(
            "SELECT id, title, deleted_at FROM project_notes"
            " WHERE project_id=? AND deleted_at IS NOT NULL ORDER BY deleted_at DESC",
            (pid,)))
    return {"tasks": tasks, "notes": notes}


@app.post("/api/tasks/{tid}/restore")
def restore_task(tid: int, body: dict):
    with db() as conn:
        r = conn.execute("SELECT * FROM tasks WHERE id=?", (tid,)).fetchone()
        if not r:
            raise HTTPException(404, "task not found")
        check_role(conn, r["project_id"], resolve_uid(body.get("actor_id")),
                   {"admin", "member"}, "復元する権限がありません")
        conn.execute("UPDATE tasks SET deleted_at=NULL WHERE id=?", (tid,))
        record_activity(conn, r["project_id"], tid, body.get("actor_id"),
                        "create", f"{r['title']}（ゴミ箱から復元）")
        recalc_parent_progress(conn, r["project_id"])
    return {"ok": True}


@app.post("/api/notes/{nid}/restore")
def restore_note(nid: int, body: dict):
    with db() as conn:
        r = conn.execute("SELECT * FROM project_notes WHERE id=?", (nid,)).fetchone()
        if not r:
            raise HTTPException(404, "note not found")
        check_role(conn, r["project_id"], resolve_uid(body.get("actor_id")),
                   {"admin", "member"}, "復元する権限がありません")
        conn.execute("UPDATE project_notes SET deleted_at=NULL WHERE id=?", (nid,))
    return {"ok": True}


# ---------------------------------------------------------------- API: QA管理

QA_STATUS_LABEL = {"open": "回答待ち", "pending": "保留",
                   "answered": "回答済み", "closed": "クローズ"}


@app.get("/api/projects/{pid}/qa")
def list_qa(pid: int):
    external_guard(pid, "qa")
    with db() as conn:
        rows = rows_to_dicts(conn.execute(
            "SELECT q.*, m.name assignee_name, t.title task_title,"
            " (SELECT COUNT(*) FROM qa_comments c WHERE c.qa_id=q.id) comment_count"
            " FROM qa_items q"
            " LEFT JOIN members m ON m.id=q.assignee_id"
            " LEFT JOIN tasks t ON t.id=q.task_id"
            " WHERE q.project_id=? ORDER BY q.seq, q.id", (pid,)))
    return rows


@app.post("/api/projects/{pid}/qa")
def create_qa(pid: int, body: dict):
    actor = resolve_uid(body.get("actor_id"))
    title = (body.get("title") or "").strip()
    if not title:
        raise HTTPException(400, "件名を入力してください")
    with db() as conn:
        check_role(conn, pid, actor, {"admin", "member"},
                   "QAを追加する権限がありません")
        seq = conn.execute("SELECT COALESCE(MAX(seq), 0) + 1 s FROM qa_items"
                           " WHERE project_id=?", (pid,)).fetchone()["s"]
        cur = conn.execute(
            "INSERT INTO qa_items(project_id, seq, title, question, answer, decision,"
            " note, category, priority, status, asker_name, assignee_id, task_id,"
            " asked_at, due_date, answered_at, created_at, updated_at)"
            " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (pid, seq, title, body.get("question") or "", body.get("answer") or "",
             body.get("decision") or "", body.get("note") or "",
             body.get("category") or "", body.get("priority") or "medium",
             body.get("status") or "open", body.get("asker_name") or "",
             body.get("assignee_id"), body.get("task_id"),
             body.get("asked_at") or date.today().isoformat(),
             body.get("due_date"), body.get("answered_at"), now(), now()))
        record_activity(conn, pid, None, actor, "create", f"QA-{seq:03d}「{title}」")
        if body.get("assignee_id"):
            notify(conn, body["assignee_id"], "system", pid, None, actor,
                   f"QA-{seq:03d}「{title}」の回答担当になりました")
        r = conn.execute("SELECT * FROM qa_items WHERE id=?", (cur.lastrowid,)).fetchone()
    return dict(r)


@app.patch("/api/qa/{qid}")
def update_qa(qid: int, body: dict):
    actor = resolve_uid(body.pop("actor_id", None))
    allowed = {"title", "question", "answer", "decision", "note", "category",
               "priority", "status", "asker_name", "assignee_id", "task_id",
               "asked_at", "due_date", "answered_at"}
    data = {k: v for k, v in body.items() if k in allowed}
    if not data:
        raise HTTPException(400, "no fields")
    with db() as conn:
        old = conn.execute("SELECT * FROM qa_items WHERE id=?", (qid,)).fetchone()
        if not old:
            raise HTTPException(404, "QA not found")
        check_role(conn, old["project_id"], actor, {"admin", "member"},
                   "QAを編集する権限がありません")
        # 回答が入ったら自動で「回答済み」・回答日を補完
        new_status = data.get("status", old["status"])
        new_answer = data.get("answer", old["answer"])
        if new_answer.strip() and old["status"] == "open" and "status" not in data:
            data["status"] = new_status = "answered"
        if new_status == "answered" and not (data.get("answered_at") or old["answered_at"]):
            data["answered_at"] = date.today().isoformat()
        sets = [f"{k}=?" for k in data] + ["updated_at=?"]
        vals = list(data.values()) + [now(), qid]
        conn.execute(f"UPDATE qa_items SET {', '.join(sets)} WHERE id=?", vals)
        if "assignee_id" in data and data["assignee_id"] and \
                data["assignee_id"] != old["assignee_id"]:
            notify(conn, data["assignee_id"], "system", old["project_id"], None, actor,
                   f"QA-{old['seq']:03d}「{old['title']}」の回答担当になりました")
        if "status" in data and data["status"] != old["status"]:
            record_activity(conn, old["project_id"], None, actor, "status",
                            f"QA-{old['seq']:03d}: {QA_STATUS_LABEL.get(old['status'], '')}"
                            f" → {QA_STATUS_LABEL.get(data['status'], '')}")
        r = conn.execute("SELECT * FROM qa_items WHERE id=?", (qid,)).fetchone()
    return dict(r)


def _fmt_qa_comment_line(c: dict) -> str:
    """やり取り1件を Excel/HTML 出力用の1行に整形（[MM/DD 名前] 本文）。"""
    name = c.get("member_name") or c.get("author_name") or "-"
    dt = (c.get("created_at") or "")[5:10].replace("-", "/")
    return f"[{dt} {name}] " + (c.get("body") or "").replace("\n", " ")


def _qa_thread_lines_db(conn, qa_id: int) -> list:
    rows = conn.execute(
        "SELECT c.*, m.name member_name FROM qa_comments c"
        " LEFT JOIN members m ON m.id=c.author_id"
        " WHERE c.qa_id=? ORDER BY c.id", (qa_id,)).fetchall()
    return [_fmt_qa_comment_line(dict(r)) for r in rows]


@app.get("/api/qa/{qid}/comments")
def list_qa_comments(qid: int):
    with db() as conn:
        q0 = conn.execute("SELECT project_id FROM qa_items WHERE id=?", (qid,)).fetchone()
    if q0:
        external_guard(q0["project_id"], "qa")
    with db() as conn:
        rows = rows_to_dicts(conn.execute(
            "SELECT c.*, m.name member_name, m.color member_color FROM qa_comments c"
            " LEFT JOIN members m ON m.id=c.author_id"
            " WHERE c.qa_id=? ORDER BY c.id", (qid,)))
    return rows


@app.post("/api/qa/{qid}/comments")
def add_qa_comment(qid: int, body: dict):
    actor = resolve_uid(body.get("actor_id"))
    text = (body.get("body") or "").strip()
    if not text:
        raise HTTPException(400, "内容を入力してください")
    with db() as conn:
        q = conn.execute("SELECT * FROM qa_items WHERE id=?", (qid,)).fetchone()
        if not q:
            raise HTTPException(404, "QA not found")
        check_role(conn, q["project_id"], actor, {"admin", "member"},
                   "QAへの記入権限がありません")
        cur = conn.execute(
            "INSERT INTO qa_comments(qa_id, author_id, author_name, body, created_at)"
            " VALUES(?,?,?,?,?)",
            (qid, actor, (body.get("author_name") or "").strip(), text, now()))
        conn.execute("UPDATE qa_items SET updated_at=? WHERE id=?", (now(), qid))
        if q["assignee_id"]:
            notify(conn, q["assignee_id"], "comment", q["project_id"], None, actor,
                   f"QA-{q['seq']:03d}「{q['title']}」にやり取りが追加されました")
        r = conn.execute(
            "SELECT c.*, m.name member_name, m.color member_color FROM qa_comments c"
            " LEFT JOIN members m ON m.id=c.author_id WHERE c.id=?",
            (cur.lastrowid,)).fetchone()
    return dict(r)


@app.delete("/api/qa/comments/{cid}")
def delete_qa_comment(cid: int, actor_id: Optional[int] = None):
    actor = resolve_uid(actor_id)
    with db() as conn:
        c = conn.execute(
            "SELECT c.*, q.project_id FROM qa_comments c"
            " JOIN qa_items q ON q.id=c.qa_id WHERE c.id=?", (cid,)).fetchone()
        if not c:
            raise HTTPException(404, "comment not found")
        if actor is not None and c["author_id"] != actor:
            check_role(conn, c["project_id"], actor, {"admin"},
                       "他の人の記録は管理者のみ削除できます")
        conn.execute("DELETE FROM qa_comments WHERE id=?", (cid,))
    return {"ok": True}


@app.delete("/api/qa/{qid}")
def delete_qa(qid: int, actor_id: Optional[int] = None):
    actor = resolve_uid(actor_id)
    with db() as conn:
        old = conn.execute("SELECT * FROM qa_items WHERE id=?", (qid,)).fetchone()
        if not old:
            raise HTTPException(404, "QA not found")
        check_role(conn, old["project_id"], actor, {"admin"},
                   "QAの削除は管理者のみ可能です")
        conn.execute("DELETE FROM qa_items WHERE id=?", (qid,))
        record_activity(conn, old["project_id"], None, actor, "delete",
                        f"QA-{old['seq']:03d}「{old['title']}」")
    return {"ok": True}


# ---------------------------------------------------------------- API: 課題管理

ISSUE_STATUS_LABEL = {"open": "未対応", "doing": "対応中",
                      "resolved": "解決済み", "closed": "クローズ"}


def _issue_tasks_map(conn, pid):
    """PJ内の課題ID → 関連タスク [{id,title}] のマップ。"""
    rows = conn.execute(
        "SELECT it.issue_id, t.id, t.title FROM issue_tasks it"
        " JOIN tasks t ON t.id=it.task_id"
        " WHERE it.issue_id IN (SELECT id FROM issues WHERE project_id=?)"
        " AND t.deleted_at IS NULL ORDER BY t.sort_order, t.id", (pid,)).fetchall()
    out: dict = {}
    for r in rows:
        out.setdefault(r["issue_id"], []).append({"id": r["id"], "title": r["title"]})
    return out


def _set_issue_tasks(conn, iid, pid, task_ids):
    """課題の関連タスクを張り替える（同一PJのタスクのみ許可）。"""
    conn.execute("DELETE FROM issue_tasks WHERE issue_id=?", (iid,))
    for tid in task_ids or []:
        ok = conn.execute(
            "SELECT 1 FROM tasks WHERE id=? AND project_id=? AND deleted_at IS NULL",
            (tid, pid)).fetchone()
        if ok:
            conn.execute("INSERT OR IGNORE INTO issue_tasks(issue_id, task_id)"
                         " VALUES(?,?)", (iid, tid))


@app.get("/api/projects/{pid}/issues-list")
def list_issues(pid: int):
    external_guard(pid, "kadai")
    with db() as conn:
        rows = rows_to_dicts(conn.execute(
            "SELECT i.*, m.name assignee_name,"
            " (SELECT COUNT(*) FROM issue_comments c WHERE c.issue_id=i.id) comment_count"
            " FROM issues i LEFT JOIN members m ON m.id=i.assignee_id"
            " WHERE i.project_id=? ORDER BY i.seq, i.id", (pid,)))
        tmap = _issue_tasks_map(conn, pid)
    for r in rows:
        r["tasks"] = tmap.get(r["id"], [])
    return rows


@app.post("/api/projects/{pid}/issues-list")
def create_issue(pid: int, body: dict):
    actor = resolve_uid(body.get("actor_id"))
    title = (body.get("title") or "").strip()
    if not title:
        raise HTTPException(400, "件名を入力してください")
    with db() as conn:
        check_role(conn, pid, actor, {"admin", "member"}, "課題を追加する権限がありません")
        seq = conn.execute("SELECT COALESCE(MAX(seq), 0) + 1 s FROM issues"
                           " WHERE project_id=?", (pid,)).fetchone()["s"]
        cur = conn.execute(
            "INSERT INTO issues(project_id, seq, title, description, policy,"
            " action_plan, category, priority, status, assignee_id, raised_by,"
            " raised_at, due_date, resolved_at, created_at, updated_at)"
            " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (pid, seq, title, body.get("description") or "", body.get("policy") or "",
             body.get("action_plan") or "", body.get("category") or "",
             body.get("priority") or "medium", body.get("status") or "open",
             body.get("assignee_id"), body.get("raised_by") or "",
             body.get("raised_at") or date.today().isoformat(),
             body.get("due_date"), body.get("resolved_at"), now(), now()))
        iid = cur.lastrowid
        _set_issue_tasks(conn, iid, pid, body.get("task_ids"))
        record_activity(conn, pid, None, actor, "create", f"課題ISS-{seq:03d}「{title}」")
        if body.get("assignee_id"):
            notify(conn, body["assignee_id"], "assign", pid, None, actor,
                   f"課題ISS-{seq:03d}「{title}」の担当になりました")
        r = conn.execute("SELECT * FROM issues WHERE id=?", (iid,)).fetchone()
    return dict(r)


@app.patch("/api/issues/{iid}")
def update_issue(iid: int, body: dict):
    actor = resolve_uid(body.pop("actor_id", None))
    allowed = {"title", "description", "policy", "action_plan", "category",
               "priority", "status", "assignee_id", "raised_by", "raised_at",
               "due_date", "resolved_at"}
    data = {k: v for k, v in body.items() if k in allowed}
    task_ids = body.get("task_ids")
    if not data and task_ids is None:
        raise HTTPException(400, "no fields")
    with db() as conn:
        old = conn.execute("SELECT * FROM issues WHERE id=?", (iid,)).fetchone()
        if not old:
            raise HTTPException(404, "issue not found")
        check_role(conn, old["project_id"], actor, {"admin", "member"},
                   "課題を編集する権限がありません")
        # 解決済み/クローズへ移すと解決日を自動補完、未対応/対応中へ戻すとクリア
        new_status = data.get("status", old["status"])
        if new_status in ("resolved", "closed") and \
                not (data.get("resolved_at") or old["resolved_at"]):
            data["resolved_at"] = date.today().isoformat()
        if new_status in ("open", "doing") and "status" in data:
            data["resolved_at"] = None
        if data:
            sets = [f"{k}=?" for k in data] + ["updated_at=?"]
            conn.execute(f"UPDATE issues SET {', '.join(sets)} WHERE id=?",
                         list(data.values()) + [now(), iid])
        if task_ids is not None:
            _set_issue_tasks(conn, iid, old["project_id"], task_ids)
        if "assignee_id" in data and data["assignee_id"] and \
                data["assignee_id"] != old["assignee_id"]:
            notify(conn, data["assignee_id"], "assign", old["project_id"], None, actor,
                   f"課題ISS-{old['seq']:03d}「{old['title']}」の担当になりました")
        if "status" in data and data["status"] != old["status"]:
            record_activity(conn, old["project_id"], None, actor, "status",
                            f"課題ISS-{old['seq']:03d}: "
                            f"{ISSUE_STATUS_LABEL.get(old['status'], '')} → "
                            f"{ISSUE_STATUS_LABEL.get(data['status'], '')}")
        r = conn.execute("SELECT * FROM issues WHERE id=?", (iid,)).fetchone()
    return dict(r)


@app.delete("/api/issues/{iid}")
def delete_issue(iid: int, actor_id: Optional[int] = None):
    actor = resolve_uid(actor_id)
    with db() as conn:
        old = conn.execute("SELECT * FROM issues WHERE id=?", (iid,)).fetchone()
        if not old:
            raise HTTPException(404, "issue not found")
        check_role(conn, old["project_id"], actor, {"admin"},
                   "課題の削除は管理者のみ可能です")
        conn.execute("DELETE FROM issues WHERE id=?", (iid,))
        record_activity(conn, old["project_id"], None, actor, "delete",
                        f"課題ISS-{old['seq']:03d}「{old['title']}」")
    return {"ok": True}


@app.get("/api/issues/{iid}/comments")
def list_issue_comments(iid: int):
    with db() as conn:
        i0 = conn.execute("SELECT project_id FROM issues WHERE id=?", (iid,)).fetchone()
    if i0:
        external_guard(i0["project_id"], "kadai")
    with db() as conn:
        rows = rows_to_dicts(conn.execute(
            "SELECT c.*, m.name member_name, m.color member_color FROM issue_comments c"
            " LEFT JOIN members m ON m.id=c.author_id"
            " WHERE c.issue_id=? ORDER BY c.id", (iid,)))
    return rows


@app.post("/api/issues/{iid}/comments")
def add_issue_comment(iid: int, body: dict):
    actor = resolve_uid(body.get("actor_id"))
    text = (body.get("body") or "").strip()
    if not text:
        raise HTTPException(400, "内容を入力してください")
    with db() as conn:
        i = conn.execute("SELECT * FROM issues WHERE id=?", (iid,)).fetchone()
        if not i:
            raise HTTPException(404, "issue not found")
        check_role(conn, i["project_id"], actor, {"admin", "member"},
                   "課題への記入権限がありません")
        cur = conn.execute(
            "INSERT INTO issue_comments(issue_id, author_id, author_name, body,"
            " created_at) VALUES(?,?,?,?,?)",
            (iid, actor, (body.get("author_name") or "").strip(), text, now()))
        conn.execute("UPDATE issues SET updated_at=? WHERE id=?", (now(), iid))
        if i["assignee_id"]:
            notify(conn, i["assignee_id"], "comment", i["project_id"], None, actor,
                   f"課題ISS-{i['seq']:03d}「{i['title']}」にコメントが追加されました")
        r = conn.execute(
            "SELECT c.*, m.name member_name, m.color member_color FROM issue_comments c"
            " LEFT JOIN members m ON m.id=c.author_id WHERE c.id=?",
            (cur.lastrowid,)).fetchone()
    return dict(r)


@app.delete("/api/issues/comments/{cid}")
def delete_issue_comment(cid: int, actor_id: Optional[int] = None):
    actor = resolve_uid(actor_id)
    with db() as conn:
        c = conn.execute(
            "SELECT c.*, i.project_id FROM issue_comments c"
            " JOIN issues i ON i.id=c.issue_id WHERE c.id=?", (cid,)).fetchone()
        if not c:
            raise HTTPException(404, "comment not found")
        if actor is not None and c["author_id"] != actor:
            check_role(conn, c["project_id"], actor, {"admin"},
                       "他の人のコメントは管理者のみ削除できます")
        conn.execute("DELETE FROM issue_comments WHERE id=?", (cid,))
    return {"ok": True}


# ---------------------------------------------------------------- API: Excelインポート（往復運用）

def _xl_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _xl_date(v):
    s = _xl_str(v).replace("/", "-")[:10]
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return None


def _xl_progress(v):
    """0.5（%書式）/ 50 / "50%" のいずれも 0-100 に正規化。"""
    if v is None or v == "":
        return None
    try:
        n = float(str(v).replace("%", "").strip())
    except ValueError:
        return None
    if 0 < n <= 1:
        n *= 100
    return max(0, min(100, round(n)))


def _xl_header_map(ws, required):
    """先頭10行からヘッダー行を探し、(行番号, {見出し: 列}) を返す。"""
    for r in range(1, min(11, ws.max_row + 1)):
        names = {}
        for c in range(1, ws.max_column + 1):
            v = _xl_str(ws.cell(row=r, column=c).value)
            if v and v not in names:
                names[v] = c
        if all(k in names for k in required):
            return r, names
    return None, None


def _pick(names, *cands):
    return next((names[c] for c in cands if c in names), None)


def _import_tasks_sheet(conn, pid, wb):
    """テーブル/WBSガントのExcelを取込。ID列で突合（更新）、ID空行は新規作成。
    空セルは既存値を保持する。"""
    ws = next((wb[n] for n in ("テーブル", "WBSガント") if n in wb.sheetnames), wb.active)
    hr, names = _xl_header_map(ws, ["タスク名"])
    if not hr:
        raise HTTPException(400, "ヘッダー行（「タスク名」列）が見つかりません")
    col = {k: _pick(names, *c) for k, c in {
        "id": ("ID",), "wbs": ("WBS",), "title": ("タスク名",),
        "status": ("ステータス",), "assignee": ("担当者", "担当"),
        "priority": ("優先度",), "start": ("開始", "開始日"), "due": ("期限",),
        "progress": ("進捗", "進捗%"), "est": ("見積h",), "act": ("実績h",),
        "tags": ("タグ",)}.items()}
    smap_by_name = {r["name"]: r["id"] for r in conn.execute(
        "SELECT id, name FROM statuses WHERE project_id=?", (pid,))}
    member_by_name = {norm_name(r["name"]): r["id"] for r in conn.execute(
        "SELECT id, name FROM members WHERE active=1")}
    prio_rev = {"最優先": "highest", "高": "high", "中": "medium", "低": "low"}
    cur_tasks = [task_row_to_dict(r) for r in conn.execute(
        "SELECT * FROM tasks WHERE project_id=? AND deleted_at IS NULL", (pid,))]
    existing = {t["id"]: t for t in cur_tasks}
    wbs_to_id = {t["wbs"]: t["id"] for t in build_wbs_rows(cur_tasks)}
    maps = _status_maps(conn, pid)
    max_order = conn.execute(
        "SELECT COALESCE(MAX(sort_order), -1) m FROM tasks WHERE project_id=?",
        (pid,)).fetchone()["m"]
    first_status = conn.execute(
        "SELECT id FROM statuses WHERE project_id=? ORDER BY sort_order LIMIT 1",
        (pid,)).fetchone()
    created = updated = skipped = 0

    for r in range(hr + 1, ws.max_row + 1):
        def g(k):
            return ws.cell(row=r, column=col[k]).value if col.get(k) else None
        title = _xl_str(g("title")).lstrip("　").strip()
        milestone = title.startswith("◆")
        title = title.lstrip("◆").strip()
        idv = _xl_str(g("id"))
        if not title and not idv:
            continue
        data = {}
        if title:
            data["title"] = title
        stn = _xl_str(g("status"))
        if stn in smap_by_name:
            data["status_id"] = smap_by_name[stn]
        an = _xl_str(g("assignee"))
        if an not in ("", "—", "-", "未割当"):
            mid = member_by_name.get(norm_name(an))
            if mid:
                data["assignee_id"] = mid
        pv = _xl_str(g("priority"))
        if pv:
            pk = prio_rev.get(pv, pv if pv in prio_rev.values() else None)
            if pk:
                data["priority"] = pk
        for key, cn in (("start_date", "start"), ("due_date", "due")):
            dv = _xl_date(g(cn))
            if dv:
                data[key] = dv
        pg = _xl_progress(g("progress"))
        if pg is not None:
            data["progress"] = pg
        for key, cn in (("estimate_h", "est"), ("actual_h", "act")):
            s = _xl_str(g(cn))
            try:
                if s:
                    data[key] = float(s)
            except ValueError:
                pass
        tg = _xl_str(g("tags"))
        if tg:
            data["tags"] = json.dumps(
                [x.strip() for x in tg.replace("／", "/").split("/") if x.strip()],
                ensure_ascii=False)

        tid = None
        if idv:
            try:
                tid = int(float(idv))
            except ValueError:
                tid = None
        if tid and tid in existing:
            old = existing[tid]
            has_kids = conn.execute(
                "SELECT 1 FROM tasks WHERE parent_id=? AND deleted_at IS NULL LIMIT 1",
                (tid,)).fetchone() is not None
            if has_kids:   # 親タスクの進捗・ステータスは自動算出
                data.pop("progress", None)
                data.pop("status_id", None)
            else:
                # ステータス⇔進捗の連動: ファイル内で「変更された側」を正として他方を導出
                # （両列が入っているExcelでは、触っていない側に古い値が残っているため）
                pg_changed = "progress" in data and data["progress"] != old["progress"]
                st_changed = ("status_id" in data
                              and data["status_id"] != old["status_id"])
                if pg_changed and not st_changed:
                    data["status_id"] = status_for_progress(
                        maps, data["progress"], old["status_id"])
                elif st_changed and not pg_changed:
                    data["progress"] = progress_for_status(
                        maps, data["status_id"], old["progress"])
            if not data:
                skipped += 1
                continue
            sets = [f"{k}=?" for k in data] + ["updated_at=?"]
            conn.execute(f"UPDATE tasks SET {', '.join(sets)} WHERE id=?",
                         list(data.values()) + [now(), tid])
            updated += 1
        elif not tid and title:
            wbs = _xl_str(g("wbs"))
            parent = wbs_to_id.get(wbs.rsplit(".", 1)[0]) if "." in wbs else None
            status_id = data.get("status_id") or (
                status_for_progress(maps, data.get("progress", 0),
                                    first_status["id"] if first_status else None))
            max_order += 1
            cur = conn.execute(
                "INSERT INTO tasks(project_id, parent_id, title, status_id, assignee_id,"
                " priority, start_date, due_date, progress, estimate_h, actual_h,"
                " milestone, tags, deps, custom_values, sort_order, created_at, updated_at)"
                " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,'[]','{}',?,?,?)",
                (pid, parent, title, status_id, data.get("assignee_id"),
                 data.get("priority", "medium"), data.get("start_date"),
                 data.get("due_date"), data.get("progress", 0),
                 data.get("estimate_h"), data.get("actual_h"), int(milestone),
                 data.get("tags", "[]"), max_order, now(), now()))
            if wbs:
                wbs_to_id[wbs] = cur.lastrowid
            created += 1
        else:
            skipped += 1   # 不明なID
    recalc_parent_progress(conn, pid)
    return {"created": created, "updated": updated, "skipped": skipped}


def _import_qa_sheet(conn, pid, wb):
    """QA管理表のExcelを取込。No列で突合（更新）、No空行は新規作成。
    空セルは既存値を保持（回答記入で自動「回答済み」）。"""
    ws = next((wb[n] for n in ("QA管理表", "QA") if n in wb.sheetnames), wb.active)
    hr, names = _xl_header_map(ws, ["No", "回答"])
    if not hr:
        raise HTTPException(400, "ヘッダー行（「No」「回答」列）が見つかりません")
    col = {k: _pick(names, *c) for k, c in {
        "seq": ("No",), "category": ("分類",), "qtext": ("件名・質問内容",),
        "title": ("件名",), "question": ("質問",), "asker": ("質問者",),
        "asked": ("質問日",), "assignee": ("回答担当",), "due": ("回答期限",),
        "status": ("ステータス", "状態"), "answer": ("回答",),
        "answered": ("回答日",), "decision": ("決定事項",),
        "note": ("備考",), "thread": ("やり取り履歴",)}.items()}
    st_rev = {v: k for k, v in QA_STATUS_LABEL.items()}
    member_by_name = {norm_name(r["name"]): r["id"] for r in conn.execute(
        "SELECT id, name FROM members WHERE active=1")}
    existing = {r["seq"]: dict(r) for r in conn.execute(
        "SELECT * FROM qa_items WHERE project_id=?", (pid,))}
    created = updated = skipped = thread_added = 0

    def import_thread(qa_id, text, fallback_name):
        """やり取り履歴セルの、DBに無い行だけを新規記録として追加する。
        行形式 [MM/DD 名前] 本文（形式外の行は fallback_name の発言として扱う）。"""
        nonlocal thread_added
        if not text:
            return
        ex_lines = set(_qa_thread_lines_db(conn, qa_id))
        for line in text.split("\n"):
            line = line.strip()
            if not line or line in ex_lines:
                continue
            m = re.match(r"^[\[［]([^\]］]*)[\]］]\s*(.*)$", line)
            if m and m.group(2).strip():
                inside = m.group(1).strip()
                parts = inside.split(None, 1)
                name = parts[1] if len(parts) == 2 else inside
                body_txt = m.group(2).strip()
            else:
                name, body_txt = fallback_name or "記入", line
            conn.execute(
                "INSERT INTO qa_comments(qa_id, author_name, body, created_at)"
                " VALUES(?,?,?,?)", (qa_id, name, body_txt, now()))
            thread_added += 1

    for r in range(hr + 1, ws.max_row + 1):
        def g(k):
            return ws.cell(row=r, column=col[k]).value if col.get(k) else None
        seqs = _xl_str(g("seq")).upper().replace("QA-", "")
        try:
            seq = int(float(seqs)) if seqs else None
        except ValueError:
            seq = None
        title = _xl_str(g("title"))
        question = _xl_str(g("question"))
        combined = _xl_str(g("qtext"))
        if combined and not title:
            parts = combined.split("\n", 1)
            title = parts[0].strip()
            question = parts[1].strip() if len(parts) > 1 else ""
        if not seq and not title:
            continue
        data = {}
        if title:
            data["title"] = title
        if question:
            data["question"] = question
        for key, cn in (("category", "category"), ("asker_name", "asker"),
                        ("answer", "answer"), ("decision", "decision"),
                        ("note", "note")):
            v = _xl_str(g(cn))
            if v:
                data[key] = v
        for key, cn in (("asked_at", "asked"), ("due_date", "due"),
                        ("answered_at", "answered")):
            v = _xl_date(g(cn))
            if v:
                data[key] = v
        stl = _xl_str(g("status"))
        if stl in st_rev:
            data["status"] = st_rev[stl]
        an = _xl_str(g("assignee"))
        if an:
            mid = member_by_name.get(norm_name(an))
            if mid:
                data["assignee_id"] = mid

        if seq and seq in existing:
            old = existing[seq]
            if (data.get("answer", "").strip() and old["status"] == "open"
                    and data.get("status", "open") == "open"):
                data["status"] = "answered"
            if data.get("status") == "answered" and \
                    not (data.get("answered_at") or old["answered_at"]):
                data["answered_at"] = date.today().isoformat()
            if not data:
                skipped += 1
                continue
            sets = [f"{k}=?" for k in data] + ["updated_at=?"]
            conn.execute(f"UPDATE qa_items SET {', '.join(sets)} WHERE id=?",
                         list(data.values()) + [now(), old["id"]])
            import_thread(old["id"], _xl_str(g("thread")),
                          data.get("asker_name") or old["asker_name"])
            updated += 1
        elif title:
            if not seq or seq in existing:
                seq = conn.execute(
                    "SELECT COALESCE(MAX(seq), 0) + 1 s FROM qa_items"
                    " WHERE project_id=?", (pid,)).fetchone()["s"]
            if data.get("answer", "").strip() and data.get("status", "open") == "open":
                data["status"] = "answered"
                data.setdefault("answered_at", date.today().isoformat())
            conn.execute(
                "INSERT INTO qa_items(project_id, seq, title, question, answer,"
                " decision, note, category, status, asker_name, assignee_id,"
                " asked_at, due_date, answered_at, created_at, updated_at)"
                " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (pid, seq, title, data.get("question", ""), data.get("answer", ""),
                 data.get("decision", ""), data.get("note", ""),
                 data.get("category", ""), data.get("status", "open"),
                 data.get("asker_name", ""), data.get("assignee_id"),
                 data.get("asked_at") or date.today().isoformat(),
                 data.get("due_date"), data.get("answered_at"), now(), now()))
            new_id = conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]
            import_thread(new_id, _xl_str(g("thread")), data.get("asker_name"))
            existing[seq] = {"id": new_id}
            created += 1
        else:
            skipped += 1
    return {"created": created, "updated": updated, "skipped": skipped,
            "thread_added": thread_added}


@app.post("/api/projects/{pid}/import/xlsx")
async def import_xlsx(pid: int, request: Request, kind: str = "tasks",
                      actor_id: Optional[int] = None):
    """エクスポートしたExcel（テーブル/WBSガント/QA管理表）を取り込む往復運用。
    突合キー: タスク=ID列 / QA=No列。キーが空の行は新規作成、空セルは既存値を保持。"""
    actor = resolve_uid(actor_id)
    raw = await request.body()
    if not raw:
        raise HTTPException(400, "ファイルが空です")
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        raise HTTPException(400, "Excelファイル（.xlsx）を読み込めませんでした")
    with db() as conn:
        check_role(conn, pid, actor, {"admin"}, "インポートは管理者のみ実行できます")
        if kind == "qa":
            result = _import_qa_sheet(conn, pid, wb)
        else:
            result = _import_tasks_sheet(conn, pid, wb)
        record_activity(conn, pid, None, actor, "create",
                        f"Excel取込（{'QA' if kind == 'qa' else 'タスク'}）:"
                        f" 更新{result['updated']}件・追加{result['created']}件")
    return result


# ---------------------------------------------------------------- API: CSVインポート

@app.post("/api/projects/{pid}/import")
def import_tasks(pid: int, body: dict):
    """rows: [{wbs?, title, status?, assignee?, priority?, start_date?, due_date?,
               progress?, estimate_h?, description?}] を一括作成。wbs で親子復元。"""
    rows = body.get("rows") or []
    actor = resolve_uid(body.get("actor_id"))
    if not rows:
        raise HTTPException(400, "rows が空です")
    with db() as conn:
        check_role(conn, pid, actor, {"admin"}, "インポートは管理者のみ実行できます")
        smap = {r["name"]: r["id"] for r in conn.execute(
            "SELECT id, name FROM statuses WHERE project_id=?", (pid,))}
        first_status = conn.execute(
            "SELECT id FROM statuses WHERE project_id=? ORDER BY sort_order LIMIT 1",
            (pid,)).fetchone()
        mmap = {norm_name(r["name"]): r["id"] for r in conn.execute(
            "SELECT id, name FROM members WHERE active=1")}
        prio_rev = {"最優先": "highest", "高": "high", "中": "medium", "低": "low"}
        max_order = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) m FROM tasks WHERE project_id=?",
            (pid,)).fetchone()["m"]
        wbs_to_id = {}
        created = 0
        for i, row in enumerate(rows):
            title = (row.get("title") or "").strip()
            if not title:
                continue
            wbs = (row.get("wbs") or "").strip()
            parent_id = None
            if wbs and "." in wbs:
                parent_id = wbs_to_id.get(wbs.rsplit(".", 1)[0])
            def numf(v):
                try:
                    return float(v) if v not in (None, "") else None
                except (TypeError, ValueError):
                    return None
            cur = conn.execute(
                "INSERT INTO tasks(project_id, parent_id, title, description, status_id,"
                " assignee_id, priority, start_date, due_date, progress, estimate_h,"
                " tags, deps, custom_values, sort_order, created_at, updated_at)"
                " VALUES(?,?,?,?,?,?,?,?,?,?,?,'[]','[]','{}',?,?,?)",
                (pid, parent_id, title, row.get("description") or "",
                 smap.get((row.get("status") or "").strip(),
                          first_status["id"] if first_status else None),
                 mmap.get(norm_name(row.get("assignee") or "")),
                 prio_rev.get((row.get("priority") or "").strip(),
                              row.get("priority") if row.get("priority") in
                              ("highest", "high", "medium", "low") else "medium"),
                 (row.get("start_date") or "").strip() or None,
                 (row.get("due_date") or "").strip() or None,
                 int(numf(row.get("progress")) or 0), numf(row.get("estimate_h")),
                 max_order + 1 + i, now(), now()))
            if wbs:
                wbs_to_id[wbs] = cur.lastrowid
            created += 1
        record_activity(conn, pid, None, actor, "create", f"CSVインポート {created} 件")
        recalc_parent_progress(conn, pid)
    return {"created": created}


# ---------------------------------------------------------------- API: APIトークン

@app.get("/api/tokens")
def list_tokens():
    su = CURRENT_USER.get()
    if su is None:
        raise HTTPException(401, "ログインが必要です")
    with db() as conn:
        rows = rows_to_dicts(conn.execute(
            "SELECT id, label, created_at, last_used FROM api_tokens"
            " WHERE member_id=? ORDER BY id DESC", (su["id"],)))
    return rows


@app.post("/api/tokens")
def create_token(body: dict):
    su = CURRENT_USER.get()
    if su is None:
        raise HTTPException(401, "ログインが必要です")
    token = "pjb_" + secrets.token_hex(20)
    with db() as conn:
        conn.execute(
            "INSERT INTO api_tokens(member_id, token_hash, label, created_at)"
            " VALUES(?,?,?,?)",
            (su["id"], hashlib.sha256(token.encode()).hexdigest(),
             (body.get("label") or "")[:60], now()))
    return {"token": token, "note": "このトークンは二度と表示されません"}


@app.delete("/api/tokens/{token_id}")
def delete_token(token_id: int):
    su = CURRENT_USER.get()
    if su is None:
        raise HTTPException(401, "ログインが必要です")
    with db() as conn:
        conn.execute("DELETE FROM api_tokens WHERE id=? AND member_id=?",
                     (token_id, su["id"]))
    return {"ok": True}


# ---------------------------------------------------------------- API: ログイン履歴

@app.get("/api/login-logs")
def get_login_logs(limit: int = 30):
    check_site_admin("ログイン履歴の閲覧はサイト管理者以上のみです")
    with db() as conn:
        rows = rows_to_dicts(conn.execute(
            "SELECT * FROM login_logs ORDER BY id DESC LIMIT ?", (limit,)))
    return rows


# ---------------------------------------------------------------- API: コメント編集・リアクション

@app.patch("/api/comments/{cid}")
def edit_comment(cid: int, body: dict):
    actor = resolve_uid(body.get("actor_id"))
    text = (body.get("body") or "").strip()
    if not text:
        raise HTTPException(400, "本文が空です")
    with db() as conn:
        c = conn.execute("SELECT * FROM comments WHERE id=?", (cid,)).fetchone()
        if not c:
            raise HTTPException(404, "comment not found")
        if actor is not None and c["author_id"] != actor:
            t = conn.execute("SELECT project_id FROM tasks WHERE id=?",
                             (c["task_id"],)).fetchone()
            if t and effective_role(conn, t["project_id"], actor) != "admin":
                raise HTTPException(403, "他人のコメントは編集できません")
        conn.execute("UPDATE comments SET body=?, updated_at=? WHERE id=?",
                     (text, now(), cid))
        r = conn.execute(
            "SELECT c.*, m.name author_name, m.color author_color FROM comments c"
            " LEFT JOIN members m ON m.id=c.author_id WHERE c.id=?", (cid,)).fetchone()
    return dict(r)


@app.post("/api/comments/{cid}/react")
def react_comment(cid: int, body: dict):
    """絵文字リアクションのトグル。"""
    uid = resolve_uid(body.get("user_id"))
    emoji = (body.get("emoji") or "").strip()[:8]
    if not uid or not emoji:
        raise HTTPException(400, "user_id と emoji が必要です")
    with db() as conn:
        exists = conn.execute(
            "SELECT id FROM reactions WHERE comment_id=? AND member_id=? AND emoji=?",
            (cid, uid, emoji)).fetchone()
        if exists:
            conn.execute("DELETE FROM reactions WHERE id=?", (exists["id"],))
        else:
            conn.execute(
                "INSERT INTO reactions(comment_id, member_id, emoji) VALUES(?,?,?)",
                (cid, uid, emoji))
    return {"ok": True, "added": not exists}


def reactions_for_comments(conn, comment_ids, me):
    """{comment_id: [{emoji, count, mine}]}"""
    if not comment_ids:
        return {}
    ph = ",".join("?" * len(comment_ids))
    out = {}
    for r in conn.execute(
            f"SELECT comment_id, emoji, COUNT(*) c,"
            f" SUM(CASE WHEN member_id=? THEN 1 ELSE 0 END) mine"
            f" FROM reactions WHERE comment_id IN ({ph})"
            f" GROUP BY comment_id, emoji", [me or 0] + list(comment_ids)):
        out.setdefault(r["comment_id"], []).append(
            {"emoji": r["emoji"], "count": r["c"], "mine": bool(r["mine"])})
    return out


# ---------------------------------------------------------------- API: タスク間リンク

@app.post("/api/tasks/{tid}/relations")
def add_relation(tid: int, body: dict):
    other = body.get("other_id")
    kind = body.get("kind") or "relates"
    if kind not in ("relates", "blocks"):
        raise HTTPException(400, "kind は relates / blocks")
    with db() as conn:
        get_task_or_404(conn, tid)
        get_task_or_404(conn, other)
        conn.execute(
            "INSERT OR IGNORE INTO task_relations(task_id, other_id, kind) VALUES(?,?,?)",
            (tid, other, kind))
    return {"ok": True}


@app.delete("/api/relations/{rid}")
def delete_relation(rid: int):
    with db() as conn:
        conn.execute("DELETE FROM task_relations WHERE id=?", (rid,))
    return {"ok": True}


# ---------------------------------------------------------------- API: PJ複製

@app.post("/api/projects/{pid}/duplicate")
def duplicate_project(pid: int, body: dict):
    actor = resolve_uid(body.get("actor_id"))
    with_tasks = bool(body.get("with_tasks", True))
    with db() as conn:
        check_role(conn, pid, actor, {"admin"}, "PJの複製は管理者のみ実行できます")
        src = conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
        if not src:
            raise HTTPException(404, "project not found")
        name = (body.get("name") or f"{src['name']} (コピー)").strip()
        cur = conn.execute(
            "INSERT INTO projects(name, description, color, start_date, end_date,"
            " custom_fields, settings, created_at) VALUES(?,?,?,?,?,?,?,?)",
            (name, src["description"], src["color"], src["start_date"], src["end_date"],
             src["custom_fields"], src["settings"], now()))
        new_pid = cur.lastrowid
        st_map = {}
        for st in conn.execute(
                "SELECT * FROM statuses WHERE project_id=? ORDER BY sort_order", (pid,)):
            c2 = conn.execute(
                "INSERT INTO statuses(project_id, name, color, sort_order, is_done)"
                " VALUES(?,?,?,?,?)",
                (new_pid, st["name"], st["color"], st["sort_order"], st["is_done"]))
            st_map[st["id"]] = c2.lastrowid
        for pm in conn.execute(
                "SELECT * FROM project_members WHERE project_id=?", (pid,)):
            conn.execute(
                "INSERT INTO project_members(project_id, member_id, role,"
                " can_view_comments, can_view_detail, added_at) VALUES(?,?,?,?,?,?)",
                (new_pid, pm["member_id"], pm["role"], pm["can_view_comments"],
                 pm["can_view_detail"], now()))
        for n in conn.execute(
                "SELECT * FROM project_notes WHERE project_id=? AND deleted_at IS NULL",
                (pid,)):
            conn.execute(
                "INSERT INTO project_notes(project_id, category, title, content, pinned,"
                " sort_order, updated_by, created_at, updated_at)"
                " VALUES(?,?,?,?,?,?,?,?,?)",
                (new_pid, n["category"], n["title"], n["content"], n["pinned"],
                 n["sort_order"], actor, now(), now()))
        if with_tasks:
            id_map = {}
            src_tasks = list(conn.execute(
                "SELECT * FROM tasks WHERE project_id=? AND deleted_at IS NULL"
                " ORDER BY sort_order, id", (pid,)))
            for t in src_tasks:
                c2 = conn.execute(
                    "INSERT INTO tasks(project_id, title, description, status_id,"
                    " assignee_id, priority, start_date, due_date, progress, estimate_h,"
                    " milestone, tags, deps, custom_values, recur, sort_order,"
                    " created_at, updated_at)"
                    " VALUES(?,?,?,?,?,?,?,?,0,?,?,?,'[]',?,?,?,?,?)",
                    (new_pid, t["title"], t["description"],
                     st_map.get(t["status_id"]), t["assignee_id"], t["priority"],
                     t["start_date"], t["due_date"], t["estimate_h"], t["milestone"],
                     t["tags"], t["custom_values"], t["recur"] or "",
                     t["sort_order"], now(), now()))
                id_map[t["id"]] = c2.lastrowid
            for t in src_tasks:   # 親子・依存の張り替え
                new_id = id_map[t["id"]]
                new_parent = id_map.get(t["parent_id"])
                new_deps = [id_map[d] for d in json.loads(t["deps"] or "[]") if d in id_map]
                conn.execute("UPDATE tasks SET parent_id=?, deps=? WHERE id=?",
                             (new_parent, json.dumps(new_deps), new_id))
        record_activity(conn, new_pid, None, actor, "create", f"「{src['name']}」から複製")
        r = conn.execute("SELECT * FROM projects WHERE id=?", (new_pid,)).fetchone()
    return project_row_to_dict(r)


# ---------------------------------------------------------------- API: 保存フィルタ

@app.get("/api/projects/{pid}/filters")
def list_filters(pid: int, user_id: Optional[int] = None):
    uid = resolve_uid(user_id)
    with db() as conn:
        rows = rows_to_dicts(conn.execute(
            "SELECT * FROM saved_filters WHERE project_id=? AND member_id=?"
            " ORDER BY id", (pid, uid or 0)))
    for r in rows:
        r["filters"] = json.loads(r["filters"] or "{}")
    return rows


@app.post("/api/projects/{pid}/filters")
def save_filter(pid: int, body: dict):
    uid = resolve_uid(body.get("user_id"))
    name = (body.get("name") or "").strip()
    if not uid or not name:
        raise HTTPException(400, "name が必要です")
    with db() as conn:
        conn.execute(
            "INSERT INTO saved_filters(member_id, project_id, name, filters, created_at)"
            " VALUES(?,?,?,?,?)",
            (uid, pid, name, json.dumps(body.get("filters") or {}, ensure_ascii=False),
             now()))
    return {"ok": True}


@app.delete("/api/filters/{fid}")
def delete_filter(fid: int, user_id: Optional[int] = None):
    uid = resolve_uid(user_id)
    with db() as conn:
        conn.execute("DELETE FROM saved_filters WHERE id=? AND member_id=?", (fid, uid or 0))
    return {"ok": True}


# ---------------------------------------------------------------- API: ベースライン

@app.post("/api/projects/{pid}/baselines")
def create_baseline(pid: int, body: dict):
    actor = resolve_uid(body.get("actor_id"))
    with db() as conn:
        check_role(conn, pid, actor, {"admin"}, "基準線の保存は管理者のみです")
        snap = [{"task_id": r["id"], "start": r["start_date"], "due": r["due_date"]}
                for r in conn.execute(
                    "SELECT id, start_date, due_date FROM tasks"
                    " WHERE project_id=? AND deleted_at IS NULL", (pid,))]
        conn.execute(
            "INSERT INTO baselines(project_id, name, snapshot, created_at) VALUES(?,?,?,?)",
            (pid, (body.get("name") or date.today().isoformat())[:40],
             json.dumps(snap), now()))
    return {"ok": True, "count": len(snap)}


@app.get("/api/projects/{pid}/baseline")
def latest_baseline(pid: int):
    with db() as conn:
        r = conn.execute(
            "SELECT * FROM baselines WHERE project_id=? ORDER BY id DESC LIMIT 1",
            (pid,)).fetchone()
    if not r:
        return {"exists": False, "snapshot": {}}
    snap = {x["task_id"]: x for x in json.loads(r["snapshot"] or "[]")}
    return {"exists": True, "name": r["name"], "created_at": r["created_at"],
            "snapshot": snap}


# ---------------------------------------------------------------- API: メトリクス（バーンダウン・工数・リスク）

@app.get("/api/projects/{pid}/metrics")
def project_metrics(pid: int):
    external_guard(pid, "dashboard")
    today = date.today()
    with db() as conn:
        statuses = list(conn.execute(
            "SELECT * FROM statuses WHERE project_id=?", (pid,)))
        done_names = [s["name"] for s in statuses if s["is_done"]]
        done_ids = {s["id"] for s in statuses if s["is_done"]}
        tasks = [task_row_to_dict(r) for r in conn.execute(
            "SELECT * FROM tasks WHERE project_id=? AND deleted_at IS NULL", (pid,))]
        # 親タスク（サブタスクあり）は自動集計値のため、メトリクスからは除外して実作業のみ数える
        parent_ids = {t["parent_id"] for t in tasks if t["parent_id"]}
        tasks = [t for t in tasks if t["id"] not in parent_ids]
        # バーンダウン: 直近30日の「完了へ遷移」アクティビティから残数推移を復元
        events = {}
        for a in conn.execute(
                "SELECT created_at, detail FROM activities"
                " WHERE project_id=? AND action='status'"
                " AND created_at > datetime('now', 'localtime', '-30 days')", (pid,)):
            d = a["created_at"][:10]
            for dn in done_names:
                if a["detail"].endswith("→ " + dn):
                    events[d] = events.get(d, 0) + 1
                elif a["detail"].startswith(dn + " →"):
                    events[d] = events.get(d, 0) - 1
        total = len(tasks)
        done_now = sum(1 for t in tasks if t["status_id"] in done_ids)
        from datetime import timedelta
        series = []
        remain = total - done_now
        # 今日から過去へ遡って残数を復元
        back = {}
        for i in range(30):
            d = (today - timedelta(days=i)).isoformat()
            back[d] = remain
            remain += events.get(d, 0)
        for i in range(29, -1, -1):
            d = (today - timedelta(days=i)).isoformat()
            series.append({"date": d, "remaining": back[d]})
        # 工数: 担当者別 見積/実績
        effort = {}
        for t in tasks:
            key = t["assignee_id"] or t.get("assignee_label") or 0
            e = effort.setdefault(key, {"estimate": 0, "actual": 0, "tasks": 0})
            e["estimate"] += t["estimate_h"] or 0
            e["actual"] += t["actual_h"] or 0
            e["tasks"] += 1
        mmap = {r["id"]: r["name"] for r in conn.execute("SELECT id, name FROM members")}
        effort_rows = [{"assignee": (mmap.get(k) if isinstance(k, int) else k) or "未割当", **v}
                       for k, v in effort.items()]
        # リスク: 期限3日以内で進捗50%未満、または期限超過（未完了）
        ts = today.isoformat()
        soon = (today + timedelta(days=3)).isoformat()
        risks = [
            {"id": t["id"], "title": t["title"], "due": t["due_date"],
             "progress": t["progress"],
             "assignee": mmap.get(t["assignee_id"]) or t.get("assignee_label"),
             "overdue": t["due_date"] < ts}
            for t in tasks
            if t["due_date"] and t["status_id"] not in done_ids
            and (t["due_date"] < ts or (t["due_date"] <= soon and t["progress"] < 50))]
        risks.sort(key=lambda x: x["due"])
    return {"burndown": series, "total": total, "effort": effort_rows, "risks": risks}


# ---------------------------------------------------------------- API: iCal配信

@app.get("/api/projects/{pid}/calendar.ics")
def project_ics(pid: int):
    check_export_allowed(pid)
    with db() as conn:
        pr = conn.execute("SELECT name FROM projects WHERE id=?", (pid,)).fetchone()
        if not pr:
            raise HTTPException(404, "project not found")
        tasks = conn.execute(
            "SELECT id, title, due_date, milestone FROM tasks"
            " WHERE project_id=? AND deleted_at IS NULL AND due_date IS NOT NULL",
            (pid,)).fetchall()
    lines = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//PJ Board//JP",
             f"X-WR-CALNAME:{pr['name']} (PJ Board)"]
    for t in tasks:
        d = t["due_date"].replace("-", "")
        prefix = "◆" if t["milestone"] else "〆"
        lines += ["BEGIN:VEVENT", f"UID:pjboard-{pid}-{t['id']}",
                  f"DTSTART;VALUE=DATE:{d}", f"DTEND;VALUE=DATE:{d}",
                  f"SUMMARY:{prefix} {t['title']}", "END:VEVENT"]
    lines.append("END:VCALENDAR")
    return Response("\r\n".join(lines), media_type="text/calendar; charset=utf-8",
                    headers={"Content-Disposition":
                             f'attachment; filename="project_{pid}.ics"'})


# ---------------------------------------------------------------- API: 週次サマリー（ルールベース）

@app.post("/api/projects/{pid}/summary")
def weekly_summary(pid: int, body: dict):
    actor = resolve_uid(body.get("actor_id"))
    with db() as conn:
        n_allowed = {"admin", "member"}
        check_role(conn, pid, actor, n_allowed, "サマリー作成の権限がありません")
        m = project_metrics(pid)
        statuses = {r["id"]: r for r in conn.execute(
            "SELECT * FROM statuses WHERE project_id=?", (pid,))}
        done_ids = {k for k, v in statuses.items() if v["is_done"]}
        tasks = [task_row_to_dict(r) for r in conn.execute(
            "SELECT * FROM tasks WHERE project_id=? AND deleted_at IS NULL", (pid,))]
        week_ago = (date.today().toordinal() - 7)
        done_week = [t for t in tasks if t["status_id"] in done_ids
                     and t["updated_at"] and
                     datetime.strptime(t["updated_at"][:10], "%Y-%m-%d").toordinal() >= week_ago]
        from datetime import timedelta
        soon = (date.today() + timedelta(days=7)).isoformat()
        upcoming = [t for t in tasks if t["due_date"] and t["status_id"] not in done_ids
                    and date.today().isoformat() <= t["due_date"] <= soon]
        text = [f"■ 週次サマリー（{date.today().isoformat()} 自動生成）", ""]
        total = len(tasks)
        done_all = sum(1 for t in tasks if t["status_id"] in done_ids)
        text.append(f"進捗: {done_all}/{total} 件完了（{round(done_all/total*100) if total else 0}%）")
        text.append("")
        text.append(f"▼ 今週完了（{len(done_week)}件）")
        text += [f"・{t['title']}" for t in done_week[:10]] or ["（なし）"]
        text.append("")
        text.append(f"▼ 来週期限（{len(upcoming)}件）")
        text += [f"・{t['title']}（{t['due_date']}）" for t in
                 sorted(upcoming, key=lambda x: x['due_date'])[:10]] or ["（なし）"]
        text.append("")
        text.append(f"▼ リスク・期限超過（{len(m['risks'])}件）")
        text += [f"・{r['title']}（期限{r['due']} 進捗{r['progress']}%"
                 f"{' ⚠超過' if r['overdue'] else ''}）" for r in m["risks"][:10]] or ["（なし）"]
        content = "\n".join(text)
        conn.execute(
            "INSERT INTO project_notes(project_id, category, title, content, pinned,"
            " sort_order, updated_by, created_at, updated_at) VALUES(?,?,?,?,0,999,?,?,?)",
            (pid, "レポート", f"週次サマリー {date.today().isoformat()}", content,
             actor, now(), now()))
        send_webhook(conn, pid, "status", content[:500])
    return {"ok": True, "content": content}


# ---------------------------------------------------------------- API: healthz / SSE

@app.get("/healthz")
def healthz():
    try:
        with db() as conn:
            conn.execute("SELECT 1")
        return {"ok": True, "time": now()}
    except Exception as e:
        raise HTTPException(500, f"db error: {e}")


@app.get("/api/events")
async def sse_events(request: Request):
    """簡易リアルタイム: activities の最新IDをポーリングしてSSEで配信。"""
    import asyncio

    async def gen():
        last = 0
        with db() as conn:
            r = conn.execute("SELECT COALESCE(MAX(id),0) m FROM activities").fetchone()
            last = r["m"]
        yield f"data: {json.dumps({'seq': last})}\n\n"
        while True:
            if await request.is_disconnected():
                break
            await asyncio.sleep(3)
            with db() as conn:
                r = conn.execute(
                    "SELECT id, project_id, actor_id, action FROM activities"
                    " WHERE id > ? ORDER BY id DESC LIMIT 1", (last,)).fetchone()
            if r:
                last = r["id"]
                yield ("data: " + json.dumps({
                    "seq": last, "project_id": r["project_id"],
                    "actor_id": r["actor_id"], "action": r["action"]}) + "\n\n")

    return StreamingResponse(gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache",
                                      "X-Accel-Buffering": "no"})


# ---------------------------------------------------------------- 日次ジョブ（リマインダー・バックアップ・ゴミ箱パージ）

def run_daily_job():
    from datetime import timedelta
    today = date.today().isoformat()
    with db() as conn:
        last = conn.execute("SELECT value FROM meta WHERE key='daily_job'").fetchone()
        if last and last["value"] == today:
            return
        conn.execute("INSERT OR REPLACE INTO meta(key, value) VALUES('daily_job', ?)",
                     (today,))
        tomorrow = (date.today() + timedelta(days=1)).isoformat()
        # 期日リマインダー（担当者へ。1日1回）
        for r in conn.execute(
                "SELECT t.*, p.name pname FROM tasks t"
                " JOIN projects p ON p.id=t.project_id"
                " LEFT JOIN statuses s ON s.id=t.status_id"
                " WHERE t.deleted_at IS NULL AND t.assignee_id IS NOT NULL"
                " AND t.due_date IS NOT NULL AND COALESCE(s.is_done,0)=0"
                " AND t.due_date <= ?", (tomorrow,)):
            if r["due_date"] < today:
                msg = f"⚠ 期限超過: 「{r['title']}」（期限 {r['due_date']}）"
            elif r["due_date"] == today:
                msg = f"⏰ 本日期限: 「{r['title']}」"
            else:
                msg = f"⏰ 明日期限: 「{r['title']}」"
            notify(conn, r["assignee_id"], "due", r["project_id"], r["id"], None, msg)
            send_webhook(conn, r["project_id"], "due", f"{msg}（{r['pname']}）")
        # ゴミ箱30日パージ
        conn.execute("DELETE FROM tasks WHERE deleted_at IS NOT NULL"
                     " AND deleted_at < datetime('now', 'localtime', '-30 days')")
        conn.execute("DELETE FROM project_notes WHERE deleted_at IS NOT NULL"
                     " AND deleted_at < datetime('now', 'localtime', '-30 days')")
        # 期限切れセッション
        conn.execute("DELETE FROM sessions WHERE last_seen < datetime('now', '-30 days')")
    # バックアップ（7世代）
    try:
        import shutil
        dst = os.path.join(BACKUP_DIR, f"pjboard-{today}.db")
        if not os.path.exists(dst):
            shutil.copy2(DB_PATH, dst)
        backups = sorted(os.listdir(BACKUP_DIR))
        for old in backups[:-7]:
            os.remove(os.path.join(BACKUP_DIR, old))
    except Exception:
        pass


def start_background_jobs():
    import threading
    import time as _time

    def loop():
        while True:
            try:
                run_daily_job()
            except Exception:
                pass
            _time.sleep(3600)

    threading.Thread(target=loop, daemon=True).start()


# ---------------------------------------------------------------- static / SPA

@app.middleware("http")
async def no_cache_static(request, call_next):
    """JS/CSS/HTML は毎回 ETag 再検証させる（更新が古いキャッシュで隠れる事故防止）。"""
    resp = await call_next(request)
    p = request.url.path
    if p == "/" or p.endswith((".js", ".css", ".html")):
        resp.headers["Cache-Control"] = "no-cache"
    return resp


init_db()
start_background_jobs()
app.mount("/", StaticFiles(directory=os.path.join(BASE_DIR, "static"), html=True),
          name="static")
