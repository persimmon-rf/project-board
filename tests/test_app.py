# -*- coding: utf-8 -*-
"""PJ Board 回帰テスト。実行: python -m pytest tests/ -q
一時DBを使うため既存データには影響しない。"""
import os
import sys
import tempfile

TMP = tempfile.mkdtemp()
os.environ["PJBOARD_DB"] = os.path.join(TMP, "test.db")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from fastapi.testclient import TestClient  # noqa: E402
import app as appmod  # noqa: E402

client = TestClient(appmod.app)
H = {"X-Requested-With": "fetch"}   # CSRFヘッダー（Cookieセッション時に必須）


def login(member_id, password=""):
    # メールログイン化に伴い、シードのダミーメール（user{id}@example.com）で認証する
    r = client.post("/api/auth/login",
                    json={"email": f"user{member_id}@example.com", "password": password},
                    headers=H)
    assert r.status_code == 200, r.text
    return r


def logout():
    client.post("/api/auth/logout", json={}, headers=H)
    client.cookies.clear()


def test_healthz():
    assert client.get("/healthz").json()["ok"] is True


def test_bootstrap_and_seed():
    d = client.get("/api/bootstrap").json()
    assert len(d["projects"]) >= 1
    assert len(d["users"]) >= 4
    assert any(u["org_role"] == "manager" for u in d["users"])


def test_login_flow_and_me():
    assert client.get("/api/auth/me").status_code == 401
    login(1)
    me = client.get("/api/auth/me").json()
    assert me["id"] == 1 and "password_hash" not in me
    logout()
    assert client.get("/api/auth/me").status_code == 401


def test_csrf_required_with_session():
    login(1)
    # ヘッダーなしの書き込みは403
    r = client.post("/api/projects/1/tasks", json={"title": "x"})
    assert r.status_code == 403
    # ヘッダーありは通る
    r = client.post("/api/projects/1/tasks", json={"title": "csrf-ok", "actor_id": 1},
                    headers=H)
    assert r.status_code == 200
    logout()


def test_permission_matrix():
    """member=自分のタスクのみ・日程不可 / 未アサイン社内=閲覧+コメント / manager=暗黙admin"""
    boot = client.get("/api/bootstrap").json()
    users = {u["name"]: u for u in boot["users"]}
    suzuki = users["鈴木 一郎"]["id"]        # PJ1 member
    d = client.get("/api/projects/1/data").json()
    own = next(t for t in d["tasks"] if t["assignee_id"] == suzuki)
    other = next(t for t in d["tasks"] if t["assignee_id"] not in (suzuki, None))

    login(suzuki)
    ok = client.patch(f"/api/tasks/{own['id']}",
                      json={"progress": own["progress"], "actor_id": suzuki}, headers=H)
    assert ok.status_code == 200
    ng = client.patch(f"/api/tasks/{own['id']}",
                      json={"due_date": "2099-01-01", "actor_id": suzuki}, headers=H)
    assert ng.status_code == 403
    ng2 = client.patch(f"/api/tasks/{other['id']}",
                       json={"progress": 1, "actor_id": suzuki}, headers=H)
    assert ng2.status_code == 403
    # 他ユーザーへの偽装は403（一般ユーザー）
    assert client.get("/api/projects/1/data?user_id=1").status_code == 403
    logout()

    # manager（田中）は暗黙admin
    login(1)
    ok2 = client.patch(f"/api/tasks/{own['id']}",
                       json={"due_date": own["due_date"], "actor_id": 1}, headers=H)
    assert ok2.status_code == 200
    logout()


def test_optimistic_lock():
    login(1)
    d = client.get("/api/projects/1/data").json()
    t = d["tasks"][0]
    r = client.patch(f"/api/tasks/{t['id']}",
                     json={"progress": t["progress"], "actor_id": 1,
                           "expected_updated_at": "1999-01-01 00:00:00"}, headers=H)
    assert r.status_code == 409
    logout()


def test_soft_delete_and_restore():
    login(1)
    t = client.post("/api/projects/1/tasks",
                    json={"title": "trash-me", "actor_id": 1}, headers=H).json()
    client.delete(f"/api/tasks/{t['id']}?actor_id=1", headers=H)
    ids = [x["id"] for x in client.get("/api/projects/1/data").json()["tasks"]]
    assert t["id"] not in ids
    trash = client.get("/api/projects/1/trash").json()
    assert any(x["id"] == t["id"] for x in trash["tasks"])
    client.post(f"/api/tasks/{t['id']}/restore", json={"actor_id": 1}, headers=H)
    ids = [x["id"] for x in client.get("/api/projects/1/data").json()["tasks"]]
    assert t["id"] in ids
    logout()


def test_notifications_on_mention_and_assign():
    login(1)
    d = client.get("/api/projects/1/data").json()
    suzuki = next(m for m in d["members"] if "鈴木" in m["name"])
    t = next(x for x in d["tasks"] if x["assignee_id"] == suzuki["id"])
    client.post(f"/api/tasks/{t['id']}/comments",
                json={"body": f"@{suzuki['name'].replace(' ', '')} 確認お願いします",
                      "author_id": 1}, headers=H)
    logout()
    login(suzuki["id"])
    n = client.get(f"/api/notifications?user_id={suzuki['id']}").json()
    assert n["unread"] >= 1
    assert any(x["type"] == "mention" for x in n["items"])
    logout()


def test_search():
    login(1)
    r = client.get("/api/search?q=設計&user_id=1").json()
    assert "tasks" in r and "comments" in r and "notes" in r
    logout()


def test_api_token_auth():
    login(1)
    tok = client.post("/api/tokens", json={"label": "test"}, headers=H).json()["token"]
    logout()
    me = client.get("/api/auth/me", headers={"Authorization": f"Bearer {tok}"})
    assert me.status_code == 200 and me.json()["id"] == 1


def test_login_throttle():
    login(1)
    client.post("/api/auth/password", json={"current_password": "", "new_password": "pw1"},
                headers=H)
    logout()
    for _ in range(5):
        client.post("/api/auth/login",
                    json={"email": "user1@example.com", "password": "bad"}, headers=H)
    r = client.post("/api/auth/login",
                    json={"email": "user1@example.com", "password": "pw1"}, headers=H)
    assert r.status_code == 429
    # 後始末（DBを直接触ってロック解除・パスワード解除）
    with appmod.db() as conn:
        conn.execute("DELETE FROM login_logs")
        conn.execute("UPDATE members SET password_hash=NULL WHERE id=1")


def test_metrics_and_baseline():
    login(1)
    m = client.get("/api/projects/1/metrics").json()
    assert len(m["burndown"]) == 30 and "effort" in m and "risks" in m
    r = client.post("/api/projects/1/baselines", json={"actor_id": 1}, headers=H)
    assert r.status_code == 200
    b = client.get("/api/projects/1/baseline").json()
    assert b["exists"] is True and len(b["snapshot"]) > 0
    logout()


def test_export_endpoints():
    for ext in ("csv", "xlsx", "html", "json"):
        assert client.get(f"/api/projects/1/export.{ext}").status_code == 200
    assert client.get("/api/projects/1/calendar.ics").status_code == 200


def test_email_login_and_debug_login():
    # 存在しないメールは401（存在有無を悟らせない共通メッセージ）
    r = client.post("/api/auth/login", json={"email": "nobody@example.com"}, headers=H)
    assert r.status_code == 401
    # デバッグログインは未ログイン＋名前指定で可（開発時のみ）
    r = client.post("/api/auth/debug-login", json={"name": "田中 太郎"}, headers=H)
    assert r.status_code == 200 and r.json()["user"]["name"] == "田中 太郎"
    logout()


def test_parent_progress_autocalc():
    """サブタスクを持つタスクの進捗は子の平均から自動算出（直接指定は無視）。"""
    login(1)
    mk = lambda body: client.post("/api/projects/1/tasks",
                                  json={**body, "actor_id": 1}, headers=H).json()
    parent = mk({"title": "進捗自動テスト親"})
    c1 = mk({"title": "子1", "parent_id": parent["id"]})
    mk({"title": "子2", "parent_id": parent["id"]})
    client.patch(f"/api/tasks/{c1['id']}",
                 json={"progress": 100, "actor_id": 1}, headers=H)
    d = client.get(f"/api/tasks/{parent['id']}/detail").json()
    assert d["task"]["progress"] == 50          # (100 + 0) / 2
    # 親タスクへの直接指定は無視され自動算出値が維持される
    r = client.patch(f"/api/tasks/{parent['id']}",
                     json={"progress": 10, "actor_id": 1}, headers=H)
    assert r.status_code == 200
    d = client.get(f"/api/tasks/{parent['id']}/detail").json()
    assert d["task"]["progress"] == 50
    logout()


def test_status_progress_sync():
    """進捗0=未着手 / 1-99=進行中 / 100=完了 の双方向連動。"""
    login(1)
    t = client.post("/api/projects/1/tasks",
                    json={"title": "連動テスト", "actor_id": 1}, headers=H).json()
    sts = client.get("/api/projects/1/data").json()["statuses"]
    first = sts[0]
    mid = next(s for s in sts if "進行" in s["name"])
    done = next(s for s in sts if s["is_done"])
    patch = lambda body: client.patch(f"/api/tasks/{t['id']}",
                                      json={**body, "actor_id": 1}, headers=H).json()
    assert patch({"progress": 40})["status_id"] == mid["id"]
    assert patch({"progress": 100})["status_id"] == done["id"]
    assert patch({"status_id": first["id"]})["progress"] == 0
    assert patch({"status_id": mid["id"]})["progress"] == 50
    r = patch({"progress": 0})
    assert r["status_id"] == first["id"] and r["progress"] == 0
    # 中間ステータス（レビュー中等）は 1-99 の進捗変更で上書きしない
    review = next((s for s in sts if not s["is_done"] and s["id"] != first["id"]
                   and "進行" not in s["name"]), None)
    if review:
        patch({"status_id": review["id"]})
        assert patch({"progress": 80})["status_id"] == review["id"]
    logout()


def test_view_xlsx_export():
    login(1)
    for view in ("table", "wbs", "board", "calendar"):
        r = client.post("/api/projects/1/export/view.xlsx",
                        json={"view": view}, headers=H)
        assert r.status_code == 200 and len(r.content) > 2000, view
    logout()


def test_user_prefs_and_webhook_test():
    login(1)
    client.post("/api/prefs", json={"user_id": 1, "key": "theme", "value": "dark"},
                headers=H)
    assert client.get("/api/prefs?user_id=1").json()["theme"] == "dark"
    # 通知転送設定があってもアプリ内通知（担当割当）は正常に動く
    client.post("/api/prefs", json={"user_id": 3, "key": "notify_webhook",
                "value": {"enabled": True, "url": "https://invalid.example/hook",
                          "events": ["assign"]}}, headers=H)
    t = client.post("/api/projects/1/tasks",
                    json={"title": "通知転送テスト", "assignee_id": 3, "actor_id": 1},
                    headers=H)
    assert t.status_code == 200
    # テスト送信は https 以外を 400 で弾く
    r = client.post("/api/me/webhook-test", json={"url": "http://example.com"}, headers=H)
    assert r.status_code == 400
    logout()


def test_webhook_provider_and_payload():
    d = appmod.detect_webhook_provider
    assert d("https://hooks.slack.com/services/T00/B00/xxx") == "slack"
    assert d("https://chat.googleapis.com/v1/spaces/AAA/messages?key=k") == "googlechat"
    assert d("https://discord.com/api/webhooks/123/abc") == "discord"
    assert d("https://discordapp.com/api/webhooks/123/abc") == "discord"
    assert d("https://contoso.webhook.office.com/webhookb2/xxx") == "teams"
    assert d("https://prod-01.japaneast.logic.azure.com/workflows/xxx") == "teams"
    assert d("https://example.com/hook") == "text"
    assert appmod.webhook_payload("slack", "hi") == {"text": "hi"}
    assert appmod.webhook_payload("googlechat", "hi") == {"text": "hi"}
    assert appmod.webhook_payload("discord", "hi") == {"content": "hi"}
    card = appmod.webhook_payload("teams", "hi")
    assert card["type"] == "message"
    assert card["attachments"][0]["content"]["body"][0]["text"] == "hi"


def test_qa_crud_and_export():
    login(1)
    r = client.post("/api/projects/1/qa",
                    json={"title": "検証環境のIPアドレスは？", "question": "接続先の確認",
                          "category": "環境", "asker_name": "顧客A", "assignee_id": 3,
                          "due_date": "2099-01-01", "actor_id": 1}, headers=H)
    assert r.status_code == 200
    qa = r.json()
    assert qa["seq"] == 1 and qa["status"] == "open"
    # 回答を書くと自動で「回答済み」＋回答日が入る。決定事項・備考も保存できる
    d = client.patch(f"/api/qa/{qa['id']}",
                     json={"answer": "10.0.0.5 です", "decision": "10.0.0.5 で確定",
                           "note": "次回定例で共有", "actor_id": 1}, headers=H).json()
    assert d["status"] == "answered" and d["answered_at"]
    assert d["decision"] == "10.0.0.5 で確定" and d["note"] == "次回定例で共有"
    assert any(x["id"] == qa["id"] for x in client.get("/api/projects/1/qa").json())
    # QA管理表Excel（ビュー単位）と全体ExcelのQAシート
    r = client.post("/api/projects/1/export/view.xlsx", json={"view": "qa"}, headers=H)
    assert r.status_code == 200 and len(r.content) > 2000
    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(client.get("/api/projects/1/export.xlsx").content))
    assert "QA" in wb.sheetnames
    # アーカイブHTMLにもQAセクション
    assert "QA（全" in client.get("/api/projects/1/export.html").text
    # 削除（管理者のみ）
    assert client.delete(f"/api/qa/{qa['id']}?actor_id=1", headers=H).status_code == 200
    logout()


def test_excel_roundtrip_import():
    """出力→顧客が記入→取込 の往復（QA: No突合 / タスク: ID突合）。"""
    import io as _io
    from openpyxl import load_workbook
    login(1)
    # --- QA: 回答を記入して取込（自動で回答済み）＋ No空行の新規追加
    qa = client.post("/api/projects/1/qa",
                     json={"title": "往復テスト質問", "asker_name": "顧客", "actor_id": 1},
                     headers=H).json()
    wb = load_workbook(_io.BytesIO(client.post(
        "/api/projects/1/export/view.xlsx", json={"view": "qa"}, headers=H).content))
    ws = wb.active
    for row in ws.iter_rows(min_row=5):
        if row[0].value == qa["seq"]:
            row[8].value = "回答しました（Excel経由）"
            row[10].value = "Excelで決定した内容"     # 決定事項列
    nr = ws.max_row + 1
    ws.cell(row=nr, column=3, value="Excelで追加した質問\n詳細です")
    ws.cell(row=nr, column=4, value="顧客B")
    bio = _io.BytesIO()
    wb.save(bio)
    d = client.post("/api/projects/1/import/xlsx?kind=qa&actor_id=1",
                    content=bio.getvalue(), headers=H).json()
    assert d["updated"] >= 1 and d["created"] == 1
    lst = client.get("/api/projects/1/qa").json()
    got = next(x for x in lst if x["id"] == qa["id"])
    assert got["answer"].startswith("回答しました") and got["status"] == "answered"
    assert got["decision"] == "Excelで決定した内容"
    assert any(x["title"] == "Excelで追加した質問" and x["question"] == "詳細です"
               for x in lst)
    for x in lst:
        if x["title"] in ("往復テスト質問", "Excelで追加した質問"):
            client.delete(f"/api/qa/{x['id']}?actor_id=1", headers=H)

    # --- タスク: テーブルExcelの進捗を書き換えて取込 → ステータスも連動
    data0 = client.get("/api/projects/1/data").json()
    kids = {t["parent_id"] for t in data0["tasks"] if t["parent_id"]}
    leaf = next(t for t in data0["tasks"]
                if t["id"] not in kids and 0 < t["progress"] < 100 or
                (t["id"] not in kids and t["progress"] == 0))
    wb = load_workbook(_io.BytesIO(client.post(
        "/api/projects/1/export/view.xlsx", json={"view": "table"}, headers=H).content))
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == leaf["id"]:
            row[8].value = 0.77   # 進捗77%（パーセント書式の値）
    bio = _io.BytesIO()
    wb.save(bio)
    d = client.post("/api/projects/1/import/xlsx?kind=tasks&actor_id=1",
                    content=bio.getvalue(), headers=H).json()
    assert d["updated"] >= 1
    d1 = client.get("/api/projects/1/data").json()
    t2 = next(t for t in d1["tasks"] if t["id"] == leaf["id"])
    smap = {s["id"]: s for s in d1["statuses"]}
    assert t2["progress"] == 77
    assert not smap[t2["status_id"]]["is_done"]     # 1-99 → 進行中系に連動
    # 元に戻す
    client.patch(f"/api/tasks/{leaf['id']}",
                 json={"progress": leaf["progress"], "actor_id": 1}, headers=H)
    logout()


def test_qa_thread_and_roundtrip():
    """QAやり取り履歴: 記録→Excel出力に整形→顧客追記行の差分取込。"""
    import io as _io
    from openpyxl import load_workbook
    login(1)
    qa = client.post("/api/projects/1/qa",
                     json={"title": "スレッドテスト", "asker_name": "顧客C", "actor_id": 1},
                     headers=H).json()
    c1 = client.post(f"/api/qa/{qa['id']}/comments",
                     json={"body": "一次回答を送付しました", "actor_id": 1}, headers=H).json()
    assert c1["member_name"] == "田中 太郎"
    assert len(client.get(f"/api/qa/{qa['id']}/comments").json()) == 1
    # 一覧に件数が乗る
    lst = client.get("/api/projects/1/qa").json()
    assert next(x for x in lst if x["id"] == qa["id"])["comment_count"] == 1
    # Excel出力の「やり取り履歴」列（13列目）に整形される
    wb = load_workbook(_io.BytesIO(client.post(
        "/api/projects/1/export/view.xlsx", json={"view": "qa"}, headers=H).content))
    ws = wb.active
    assert ws.cell(row=4, column=13).value == "やり取り履歴"
    target = None
    for row in ws.iter_rows(min_row=5):
        if row[0].value == qa["seq"]:
            target = row
    assert "一次回答を送付しました" in target[12].value
    assert "田中 太郎" in target[12].value
    # 顧客がセルに1行追記して返送 → 差分だけ新規記録に
    target[12].value += "\n[09/02 顧客C] 追加で仕様の確認をお願いします"
    bio = _io.BytesIO()
    wb.save(bio)
    d = client.post("/api/projects/1/import/xlsx?kind=qa&actor_id=1",
                    content=bio.getvalue(), headers=H).json()
    assert d["thread_added"] == 1
    cs = client.get(f"/api/qa/{qa['id']}/comments").json()
    assert len(cs) == 2
    assert cs[1]["author_name"] == "顧客C"
    assert cs[1]["body"] == "追加で仕様の確認をお願いします"
    # 同じファイルをもう一度取込しても重複しない（冪等）
    wb2 = load_workbook(_io.BytesIO(client.post(
        "/api/projects/1/export/view.xlsx", json={"view": "qa"}, headers=H).content))
    bio2 = _io.BytesIO()
    wb2.save(bio2)
    d2 = client.post("/api/projects/1/import/xlsx?kind=qa&actor_id=1",
                     content=bio2.getvalue(), headers=H).json()
    assert d2["thread_added"] == 0
    assert len(client.get(f"/api/qa/{qa['id']}/comments").json()) == 2
    client.delete(f"/api/qa/{qa['id']}?actor_id=1", headers=H)
    logout()


def test_admin_console_and_external_security():
    """管理画面API権限＋外部ユーザーの多層防御（アサイン確認・タブ制限・遮断）。"""
    # 管理画面API: staff は403、manager はOK
    login(3)
    assert client.get("/api/admin/projects").status_code == 403
    assert client.get("/api/admin/analytics").status_code == 403
    logout()
    login(1)
    pjs = client.get("/api/admin/projects").json()
    assert any(x["project"]["id"] == 1 for x in pjs)
    ana = client.get("/api/admin/analytics").json()
    assert "workload" in ana and "externals" in ana and "overdue_tasks" in ana
    # 外部アサイン: 氏名の確認入力（confirm_name）が無いと400
    r = client.post("/api/projects/1/members",
                    json={"member_id": 5, "actor_id": 1}, headers=H)
    assert r.status_code == 400
    r = client.post("/api/projects/1/members",
                    json={"member_id": 5, "confirm_name": "高橋 健", "actor_id": 1},
                    headers=H)
    assert r.status_code == 200
    client.delete("/api/projects/1/members/5", headers=H)
    # 外部公開設定の変更: プロジェクト名の確認入力（confirm_text）が無いと400
    cur = client.get("/api/projects/1/data").json()["project"]["settings"]
    r = client.patch("/api/projects/1?actor_id=1",
                     json={"settings": {**cur, "external_visible_tabs": ["qa"]}},
                     headers=H)
    assert r.status_code == 400
    r = client.patch("/api/projects/1?actor_id=1",
                     json={"settings": {**cur, "external_visible_tabs": ["qa"]},
                           "confirm_text": "社内ポータル刷新"}, headers=H)
    assert r.status_code == 200
    logout()
    # 未アサインの外部（高橋）は全面遮断
    login(5)
    assert client.get("/api/projects/1/data").status_code == 403
    assert client.get("/api/projects/1/qa").status_code == 403
    assert client.get("/api/projects/1/export.html").status_code == 403
    logout()
    # アサイン済みの外部（山田）: 公開タブ=qaのみ → タスクは返らずQAのみ閲覧可
    login(4)
    d = client.get("/api/projects/1/data").json()
    assert d["tasks"] == []
    assert client.get("/api/projects/1/qa").status_code == 200
    assert client.get("/api/projects/1/notes").status_code == 403
    assert client.get("/api/projects/1/metrics").status_code == 403
    logout()
    # 後始末: 公開タブを既定（全公開）へ戻す
    login(1)
    r = client.patch("/api/projects/1?actor_id=1",
                     json={"settings": cur, "confirm_text": "社内ポータル刷新"},
                     headers=H)
    assert r.status_code == 200
    logout()


def test_issue_management():
    """課題管理: CRUD・関連タスク・タスク側のオープン課題数・コメント・エクスポート・外部既定非公開。"""
    login(1)
    mk = lambda b: client.post("/api/projects/1/tasks",
                               json={**b, "actor_id": 1}, headers=H).json()
    t1 = mk({"title": "課題関連タスク1"})
    t2 = mk({"title": "課題関連タスク2"})
    r = client.post("/api/projects/1/issues-list",
                    json={"title": "一覧表示の性能が出ない", "description": "3秒かかる",
                          "policy": "インデックス追加で対応", "action_plan": "SQL見直し",
                          "assignee_id": 3, "due_date": "2099-01-01",
                          "task_ids": [t1["id"], t2["id"]], "actor_id": 1}, headers=H)
    assert r.status_code == 200
    iss = r.json()
    assert iss["seq"] == 1 and iss["status"] == "open"
    got = next(x for x in client.get("/api/projects/1/issues-list").json()
               if x["id"] == iss["id"])
    assert len(got["tasks"]) == 2 and got["assignee_name"] == "鈴木 一郎"
    # タスク側にオープン課題数が乗る
    d = client.get("/api/projects/1/data").json()
    assert next(t for t in d["tasks"] if t["id"] == t1["id"])["issue_count"] == 1
    # タスク詳細にも関連課題
    dd = client.get(f"/api/tasks/{t1['id']}/detail").json()
    assert any(x["id"] == iss["id"] for x in dd["issues"])
    # コメント
    c = client.post(f"/api/issues/{iss['id']}/comments",
                    json={"body": "調査を開始", "actor_id": 1}, headers=H).json()
    assert c["member_name"] == "田中 太郎"
    # クローズ → 解決日が自動で入り、タスクの件数から外れる
    u = client.patch(f"/api/issues/{iss['id']}",
                     json={"status": "closed", "actor_id": 1}, headers=H).json()
    assert u["resolved_at"]
    d = client.get("/api/projects/1/data").json()
    assert next(t for t in d["tasks"] if t["id"] == t1["id"])["issue_count"] == 0
    # エクスポート（ビューExcel・全体Excelの課題シート・アーカイブHTML）
    r = client.post("/api/projects/1/export/view.xlsx",
                    json={"view": "kadai"}, headers=H)
    assert r.status_code == 200 and len(r.content) > 2000
    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(client.get("/api/projects/1/export.xlsx").content))
    assert "課題" in wb.sheetnames
    assert "課題（全" in client.get("/api/projects/1/export.html").text
    logout()
    # 外部ユーザー（山田・PJ1メンバー）: 課題タブは既定で非公開 → 403
    login(4)
    assert client.get("/api/projects/1/issues-list").status_code == 403
    logout()
    # 後始末
    login(1)
    client.delete(f"/api/issues/{iss['id']}?actor_id=1", headers=H)
    client.delete(f"/api/tasks/{t1['id']}?actor_id=1", headers=H)
    client.delete(f"/api/tasks/{t2['id']}?actor_id=1", headers=H)
    logout()


def test_member_email_required_and_unique():
    login(1)
    base = {"name": "メール必須テスト", "role": "", "color": "#111111"}
    r = client.post("/api/members", json={**base, "email": ""}, headers=H)
    assert r.status_code == 400
    r = client.post("/api/members", json={**base, "email": "user1@example.com"}, headers=H)
    assert r.status_code == 400          # 既存と重複
    r = client.post("/api/members", json={**base, "email": "New.User@Example.com"},
                    headers=H)
    assert r.status_code == 200 and r.json()["email"] == "new.user@example.com"
    logout()
